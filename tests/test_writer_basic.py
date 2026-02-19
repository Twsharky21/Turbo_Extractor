\
from openpyxl import Workbook

from core.planner import build_plan
from core.writer import apply_write_plan


def test_writer_writes_exact_rectangle():
    wb = Workbook()
    ws = wb.active

    shaped = [
        ["a", "b"],
        ["c", "d"],
    ]

    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    rows_written = apply_write_plan(ws, shaped, plan)

    assert rows_written == 2
    assert ws["C1"].value == "a"
    assert ws["D1"].value == "b"
    assert ws["C2"].value == "c"
    assert ws["D2"].value == "d"


def test_writer_appends_after_existing_data():
    wb = Workbook()
    ws = wb.active

    ws["C1"] = "existing"
    ws["D3"] = "also existing"

    shaped = [
        ["x", "y"],
    ]

    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    apply_write_plan(ws, shaped, plan)

    # max used row across C:D is 3 -> append at row 4
    assert ws["C4"].value == "x"
    assert ws["D4"].value == "y"
