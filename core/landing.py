"""
core/landing.py — Landing-zone scan and collision detection.

This module is the SOLE authority for:
  1. Extracting which columns in a shaped grid actually carry data (target cols).
  2. Determining where an extraction lands (append scan on target cols only).
  3. Deciding whether that zone is clear (collision probe on target cols only).

Design principles:
  - Scan and probe operate on TARGET COLUMNS ONLY — columns in the shaped grid
    that contain at least one non-None value. Gap columns (None throughout) are
    completely ignored for placement purposes. This enables automatic merge:
    two extractions whose bounding boxes overlap can still land side-by-side
    as long as their actual data columns don't conflict.
  - Uses ws.iter_rows() for reading — never ws.cell() in a scan loop — to avoid
    the openpyxl phantom-cell bug where reading an unwritten cell registers it
    and silently inflates ws.max_row.
  - Pure functions: all logic below read_zone() works on plain dicts and lists,
    making them fully testable without openpyxl.

Public API:
  find_target_col_offsets(shaped)                   -> List[int]  (0-based offsets)
  read_zone(ws, col_start, col_end, extra_rows)     -> CellMap
  scan_target_cols(cell_map, target_abs_cols)       -> int  (max occupied row)
  probe_target_cols(cell_map, row_start, row_end,
                    target_abs_cols)                -> (row,col,value)|None
  is_dest_cell_occupied(value)                      -> bool
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet


# ── Occupancy definition ──────────────────────────────────────────────────────

def is_dest_cell_occupied(value: Any) -> bool:
    """
    Destination occupancy per spec:
      OCCUPIED   — text (incl. whitespace/"N/A"), numbers (incl. 0),
                   dates, booleans, formula WITH a cached result.
      UNOCCUPIED — None, empty string "",
                   bare formula string with no cached result.
    """
    if value is None:
        return False
    if isinstance(value, str):
        if value == "":
            return False
        if value.startswith("="):
            return False
    return True


# ── Target column extraction ──────────────────────────────────────────────────

def find_target_col_offsets(shaped: List[List[Any]]) -> List[int]:
    """
    Return the 0-based column offsets within the shaped grid that contain
    at least one non-None, non-empty value across all rows.

    These are the columns that apply_write_plan will actually write data into.
    Gap columns (all-None) are excluded — they come from Keep Format bounding
    boxes and must not influence placement decisions.

    Example:
      shaped = [['a', None, 'c'], ['d', None, 'f']]
      → [0, 2]   (col 1 is a gap — all None)
    """
    if not shaped:
        return []
    width = max((len(r) for r in shaped), default=0)
    result = []
    for c in range(width):
        for row in shaped:
            if c < len(row) and is_dest_cell_occupied(row[c]):
                result.append(c)
                break
    return result


# ── Worksheet snapshot ────────────────────────────────────────────────────────

CellMap = Dict[Tuple[int, int], Any]


def read_zone(
    ws: Worksheet,
    col_start: int,
    col_end: int,
    extra_rows: int = 0,
) -> CellMap:
    """
    Read cells in col_start..col_end from the worksheet using iter_rows().

    iter_rows() is a safe read-only iterator that does NOT register phantom
    Cell objects, so ws.max_row is never inflated by this call.

    Only cells with non-None values are included in the returned map.
    extra_rows: extend the read beyond ws.max_row to cover the probe zone.
    """
    cell_map: CellMap = {}
    max_r = ws.max_row or 0
    upper = max_r + extra_rows

    if upper < 1 or col_start > col_end:
        return cell_map

    for row_cells in ws.iter_rows(
        min_row=1,
        max_row=upper,
        min_col=col_start,
        max_col=col_end,
    ):
        for cell in row_cells:
            if cell.value is not None:
                cell_map[(cell.row, cell.column)] = cell.value

    return cell_map


# ── Scan (target columns only) ────────────────────────────────────────────────

def scan_target_cols(
    cell_map: CellMap,
    target_abs_cols: List[int],
) -> int:
    """
    Find the highest occupied row across the target absolute columns only.

    Gap columns are not passed in — they never influence the scan result.
    Returns 0 if no occupied cells exist in the target columns.
    """
    if not target_abs_cols:
        return 0
    col_set = set(target_abs_cols)
    max_row = 0
    for (r, c), value in cell_map.items():
        if c in col_set and is_dest_cell_occupied(value):
            if r > max_row:
                max_row = r
    return max_row


# ── Collision probe (target columns only) ─────────────────────────────────────

def probe_target_cols(
    cell_map: CellMap,
    row_start: int,
    row_end: int,
    target_abs_cols: List[int],
) -> Optional[Tuple[int, int, Any]]:
    """
    Check whether any target column cell in row_start..row_end is occupied.

    Returns (row, col, value) of the first blocker found, or None if clear.
    Scans in row-major order so the reported blocker is the top-left one.

    Gap columns are not checked — it is valid for gap positions to have
    existing data (that's the whole point of merge mode).
    """
    if not target_abs_cols:
        return None
    col_set = set(target_abs_cols)
    for r in range(row_start, row_end + 1):
        for c in sorted(col_set):
            value = cell_map.get((r, c))
            if is_dest_cell_occupied(value):
                return (r, c, value)
    return None


# ── Backward-compat aliases (patch6 test_landing.py uses these names) ─────────

def scan_landing_zone(
    cell_map: CellMap,
    col_start: int,
    col_end: int,
) -> int:
    """
    Alias for scan_target_cols over a contiguous column range.
    Kept for backward compatibility with test_landing.py from patch6.
    """
    target_cols = list(range(col_start, col_end + 1))
    return scan_target_cols(cell_map, target_cols)


def probe_collision(
    cell_map: CellMap,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
) -> Optional[Tuple[int, int, Any]]:
    """
    Alias for probe_target_cols over a contiguous column range.
    Kept for backward compatibility with test_landing.py from patch6.
    """
    target_cols = list(range(col_start, col_end + 1))
    return probe_target_cols(cell_map, row_start, row_end, target_cols)
