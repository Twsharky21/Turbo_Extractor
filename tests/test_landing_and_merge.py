"""
tests/test_landing_and_merge.py

Tests for:
  1. core.landing — target col extraction, scan, probe (pure dict tests)
  2. Merge behaviour — side-by-side extractions with overlapping bounding boxes
  3. writer.py — None cells are never written (no phantom-cell poisoning)
  4. End-to-end: the exact screenshot scenario (CSV cols A,C then XLSX cols B,D)
"""
from __future__ import annotations

import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet, run_all
from core.errors import AppError, DEST_BLOCKED
from core.landing import (
    find_target_col_offsets,
    is_dest_cell_occupied,
    read_zone,
    scan_target_cols,
    probe_target_cols,
)
from core.models import Destination, SheetConfig
from core.planner import build_plan
from core.writer import apply_write_plan


# ── helpers ───────────────────────────────────────────────────────────────────

def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _cfg(dest, *, start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", columns="", rows="", mode="pack"):
    return SheetConfig(
        name=src_sheet, workbook_sheet=src_sheet,
        source_start_row="", columns_spec=columns, rows_spec=rows,
        paste_mode=mode, rules_combine="AND", rules=[],
        destination=Destination(
            file_path=dest, sheet_name=dest_sheet,
            start_col=start_col, start_row=start_row,
        ),
    )


def _ws():
    return Workbook().active


# ══════════════════════════════════════════════════════════════════════════════
# find_target_col_offsets
# ══════════════════════════════════════════════════════════════════════════════

class TestFindTargetColOffsets:
    def test_pack_mode_all_cols_are_targets(self):
        shaped = [["a", "b"], ["c", "d"]]
        assert find_target_col_offsets(shaped) == [0, 1]

    def test_keep_mode_gap_col_excluded(self):
        # Keep A,C → [data, None, data]
        shaped = [["a", None, "c"], ["d", None, "f"]]
        assert find_target_col_offsets(shaped) == [0, 2]

    def test_keep_mode_multiple_gaps(self):
        # Keep A,E → [data, None, None, None, data]
        shaped = [["a", None, None, None, "e"]]
        assert find_target_col_offsets(shaped) == [0, 4]

    def test_all_none_returns_empty(self):
        shaped = [[None, None], [None, None]]
        assert find_target_col_offsets(shaped) == []

    def test_empty_shaped_returns_empty(self):
        assert find_target_col_offsets([]) == []

    def test_zero_value_counts_as_target(self):
        shaped = [[0, None, False]]
        assert find_target_col_offsets(shaped) == [0, 2]

    def test_single_col_is_target(self):
        shaped = [["only"]]
        assert find_target_col_offsets(shaped) == [0]

    def test_gap_only_in_some_rows_still_target(self):
        # Col 1 has None in row 0 but data in row 1 → still a target
        shaped = [[None, "a"], ["b", "c"]]
        assert find_target_col_offsets(shaped) == [0, 1]


# ══════════════════════════════════════════════════════════════════════════════
# scan_target_cols
# ══════════════════════════════════════════════════════════════════════════════

class TestScanTargetCols:
    def test_empty_map_returns_zero(self):
        assert scan_target_cols({}, [1, 3]) == 0

    def test_finds_max_row_in_target_cols(self):
        cell_map = {(1, 1): "a", (5, 3): "b", (3, 2): "noise"}
        # target cols 1 and 3 — col 2 is gap
        assert scan_target_cols(cell_map, [1, 3]) == 5

    def test_gap_col_not_counted(self):
        cell_map = {(10, 2): "gap_data", (3, 1): "real"}
        # Only col 1 is target
        assert scan_target_cols(cell_map, [1]) == 3

    def test_zero_value_counts(self):
        cell_map = {(7, 3): 0}
        assert scan_target_cols(cell_map, [3]) == 7

    def test_none_value_not_counted(self):
        cell_map = {(9, 1): None, (4, 1): "real"}
        assert scan_target_cols(cell_map, [1]) == 4


# ══════════════════════════════════════════════════════════════════════════════
# probe_target_cols
# ══════════════════════════════════════════════════════════════════════════════

class TestProbeTargetCols:
    def test_clear_zone_returns_none(self):
        assert probe_target_cols({}, 1, 5, [1, 3]) is None

    def test_occupied_target_col_returns_blocker(self):
        cell_map = {(3, 1): "BLOCK"}
        result = probe_target_cols(cell_map, 1, 5, [1, 3])
        assert result == (3, 1, "BLOCK")

    def test_gap_col_data_does_not_block(self):
        # Col 2 is a gap — data there must NOT be a blocker
        cell_map = {(2, 2): "existing_from_other_extraction"}
        result = probe_target_cols(cell_map, 1, 5, [1, 3])
        assert result is None

    def test_returns_first_blocker_row_major(self):
        cell_map = {(2, 3): "second", (1, 1): "first"}
        result = probe_target_cols(cell_map, 1, 5, [1, 3])
        assert result == (1, 1, "first")

    def test_outside_row_range_not_reported(self):
        cell_map = {(10, 1): "outside"}
        assert probe_target_cols(cell_map, 1, 5, [1, 3]) is None


# ══════════════════════════════════════════════════════════════════════════════
# writer: None cells never written
# ══════════════════════════════════════════════════════════════════════════════

class TestWriterSkipsNone:
    def test_none_cells_not_written(self):
        ws = _ws()
        shaped = [["a", None, "c"]]
        plan = build_plan(ws, shaped, "A", "1")
        apply_write_plan(ws, shaped, plan)
        assert ws["A1"].value == "a"
        assert ws["C1"].value == "c"
        # B1 must not be registered — check via iter_rows
        b_vals = [cell.value for row in ws.iter_rows(1, 1, 2, 2) for cell in row]
        assert all(v is None for v in b_vals)

    def test_none_cells_dont_inflate_max_row(self):
        ws = _ws()
        # Write 5 rows with gap col in the middle
        shaped = [[f"a{i}", None, f"c{i}"] for i in range(1, 6)]
        plan = build_plan(ws, shaped, "A", "1")
        apply_write_plan(ws, shaped, plan)
        # max_col should be 3 (C), max_row should be 5
        assert ws.max_row == 5
        # B column should have NO registered cells
        b_cells = [cell.value for row in ws.iter_rows(1, 5, 2, 2) for cell in row]
        assert all(v is None for v in b_cells)

    def test_none_gap_does_not_poison_next_scan(self):
        """
        Core regression test: after writing Keep Format A,C (gap at B),
        scanning column B must return max_used=0 so next extraction starts at row 1.
        """
        ws = _ws()
        shaped = [[f"a{i}", None, f"c{i}"] for i in range(1, 527)]
        plan = build_plan(ws, shaped, "A", "1")
        apply_write_plan(ws, shaped, plan)

        # Scan B (col 2) only
        scan_map = read_zone(ws, 2, 2)
        max_used = scan_target_cols(scan_map, [2])
        assert max_used == 0, f"B column poisoned: max_used={max_used}"


# ══════════════════════════════════════════════════════════════════════════════
# read_zone: no phantom cells
# ══════════════════════════════════════════════════════════════════════════════

class TestReadZoneNoPhantoms:
    def test_reading_empty_cols_does_not_inflate_max_row(self):
        ws = _ws()
        for i in range(1, 8):
            ws.cell(row=i, column=1, value=f"name{i}")
            ws.cell(row=i, column=2, value=i * 1000.0)
        assert ws.max_row == 7
        read_zone(ws, 3, 4)           # read empty C:D
        assert ws.max_row == 7        # must not change


# ══════════════════════════════════════════════════════════════════════════════
# Merge: overlapping bounding boxes land side-by-side
# ══════════════════════════════════════════════════════════════════════════════

class TestMergeSideBySide:
    def test_keep_AC_then_keep_BD_merge_same_rows(self):
        """
        The exact screenshot scenario:
        - Source 1: Keep A,C → bounding box A:C, data in A and C, gap at B
        - Source 2: Keep B,D → bounding box B:D, data in B and D, gap at C
        - Both append to same dest sheet
        - Expected: data merges into rows 1-N, not stacked
        """
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "src1.xlsx")
            src2 = os.path.join(td, "src2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            # Source 1: 5 rows, cols A and C matter
            _xlsx(src1, [
                ["a1", "x", "c1", "y"],
                ["a2", "x", "c2", "y"],
                ["a3", "x", "c3", "y"],
                ["a4", "x", "c4", "y"],
                ["a5", "x", "c5", "y"],
            ])
            # Source 2: 5 rows, cols B and D matter
            _xlsx(src2, [
                ["p", "b1", "q", "d1"],
                ["p", "b2", "q", "d2"],
                ["p", "b3", "q", "d3"],
                ["p", "b4", "q", "d4"],
                ["p", "b5", "q", "d5"],
            ])

            r1 = run_sheet(src1, _cfg(dest, columns="A,C", mode="keep", start_col="A"))
            r2 = run_sheet(src2, _cfg(dest, columns="B,D", mode="keep", start_col="B"))

            assert r1.rows_written == 5
            assert r2.rows_written == 5

            ws = load_workbook(dest)["Out"]

            # All data in rows 1-5, not stacked
            for i in range(1, 6):
                assert ws.cell(row=i, column=1).value == f"a{i}", f"A{i} wrong"
                assert ws.cell(row=i, column=2).value == f"b{i}", f"B{i} wrong"
                assert ws.cell(row=i, column=3).value == f"c{i}", f"C{i} wrong"
                assert ws.cell(row=i, column=4).value == f"d{i}", f"D{i} wrong"

            # No data below row 5
            assert ws.cell(row=6, column=1).value is None
            assert ws.cell(row=6, column=2).value is None

    def test_keep_AC_then_keep_BD_via_run_all(self):
        """Same scenario via run_all (shared workbook cache)."""
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "src1.xlsx")
            src2 = os.path.join(td, "src2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(src1, [["a1", "x", "c1"], ["a2", "x", "c2"], ["a3", "x", "c3"]])
            _xlsx(src2, [["p", "b1", "q", "d1"], ["p", "b2", "q", "d2"], ["p", "b3", "q", "d3"]])

            report = run_all([
                (src1, "R1", _cfg(dest, columns="A,C", mode="keep", start_col="A")),
                (src2, "R2", _cfg(dest, columns="B,D", mode="keep", start_col="B")),
            ])

            assert report.ok
            ws = load_workbook(dest)["Out"]
            assert ws["A1"].value == "a1"
            assert ws["B1"].value == "b1"
            assert ws["C1"].value == "c1"
            assert ws["D1"].value == "d1"
            assert ws["A4"].value is None   # no row 4

    def test_merge_collision_on_target_col_raises_dest_blocked(self):
        """If a target col has existing data in the probe zone → DEST_BLOCKED."""
        with TemporaryDirectory() as td:
            src = os.path.join(td, "src.xlsx")
            dest = os.path.join(td, "dest.xlsx")
            _xlsx(src, [["x", "y"]])

            # Pre-populate dest: B1 has data (target col for second extraction)
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["B1"] = "EXISTING"
            wb.save(dest)

            # Keep B,D extraction with start_col=B, start_row=1 → B1 blocked
            with pytest.raises(AppError) as ei:
                run_sheet(src, _cfg(dest, columns="B,D", mode="keep",
                                    start_col="B", start_row="1"))
            assert ei.value.code == DEST_BLOCKED

    def test_gap_col_data_does_not_cause_dest_blocked(self):
        """
        C is a gap column for a B,D Keep extraction (bounding box B:D, data in B and D).
        Existing data in C must NOT trigger DEST_BLOCKED.
        """
        with TemporaryDirectory() as td:
            src = os.path.join(td, "src.xlsx")
            dest = os.path.join(td, "dest.xlsx")
            _xlsx(src, [["p", "b1", "q", "d1"]])

            # Pre-populate dest: C1 has data (gap col, not a target)
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["C1"] = "existing_from_other"
            wb.save(dest)

            # Should succeed: B and D are clear, C is a gap
            r = run_sheet(src, _cfg(dest, columns="B,D", mode="keep",
                                    start_col="B", start_row="1"))
            assert r.rows_written == 1
            ws2 = load_workbook(dest)["Out"]
            assert ws2["B1"].value == "b1"
            assert ws2["C1"].value == "existing_from_other"   # untouched
            assert ws2["D1"].value == "d1"

    def test_unequal_row_counts_shorter_leaves_blanks(self):
        """
        Source 1 has 5 rows, source 2 has 3 rows.
        Rows 4-5 in source 2's target cols stay blank — no error.
        """
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "src1.xlsx")
            src2 = os.path.join(td, "src2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(src1, [[f"a{i}", "x", f"c{i}"] for i in range(1, 6)])
            _xlsx(src2, [["p", f"b{i}", "q", f"d{i}"] for i in range(1, 4)])

            run_sheet(src1, _cfg(dest, columns="A,C", mode="keep", start_col="A"))
            run_sheet(src2, _cfg(dest, columns="B,D", mode="keep", start_col="B"))

            ws = load_workbook(dest)["Out"]
            # Rows 1-3: all four cols populated
            for i in range(1, 4):
                assert ws.cell(row=i, column=1).value == f"a{i}"
                assert ws.cell(row=i, column=2).value == f"b{i}"
            # Rows 4-5: A and C have data, B and D are blank
            assert ws.cell(row=4, column=1).value == "a4"
            assert ws.cell(row=4, column=2).value is None
            assert ws.cell(row=4, column=4).value is None

    def test_pack_side_by_side_unaffected(self):
        """Pack mode side-by-side still works correctly."""
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "src1.xlsx")
            src2 = os.path.join(td, "src2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(src1, [["name1", 100], ["name2", 200]])
            _xlsx(src2, [["date1", 111], ["date2", 222]])

            run_sheet(src1, _cfg(dest, start_col="A", mode="pack"))
            run_sheet(src2, _cfg(dest, start_col="C", mode="pack"))

            ws = load_workbook(dest)["Out"]
            assert ws["A1"].value == "name1"
            assert ws["C1"].value == "date1"
            assert ws["A2"].value == "name2"
            assert ws["C2"].value == "date2"
            assert ws["A3"].value is None
