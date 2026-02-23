"""
core/runner.py — Single-sheet extraction executor.

Responsible for:
  - Loading the source table (XLSX/CSV)
  - Applying source start row offset
  - Applying row selection, rules, column selection, and paste shaping
  - Opening or creating the destination workbook
  - Delegating placement to planner and write to writer
  - Returning a SheetResult

This module has NO knowledge of batch execution or progress callbacks.
"""
from __future__ import annotations

import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from .errors import (
    AppError,
    SOURCE_READ_FAILED, SHEET_NOT_FOUND,
    FILE_LOCKED, SAVE_FAILED, MISSING_DEST_PATH,
)
from .models import SheetConfig, SheetResult
from .parsing import parse_columns, parse_rows
from .io import load_csv, load_xlsx, compute_used_range, normalize_table
from .rules import apply_rules
from .transform import apply_row_selection, apply_column_selection, shape_pack, shape_keep
from .planner import build_plan
from .writer import apply_write_plan


# ── Private helpers ───────────────────────────────────────────────────────────

def _apply_source_start_row(table: List[List[Any]], source_start_row: str) -> List[List[Any]]:
    """Apply a 1-based source start row offset. Blank means no offset."""
    s = (source_start_row or "").strip()
    if not s:
        return table
    try:
        n = int(s)
    except ValueError:
        raise AppError(
            "BAD_SOURCE_START_ROW",
            f"Source Start Row must be a number (got '{source_start_row}')",
        )
    if n < 1:
        raise AppError(
            "BAD_SOURCE_START_ROW",
            f"Source Start Row must be >= 1 (got {n})",
        )
    offset = n - 1
    if offset <= 0:
        return table
    return table[offset:]


def _load_source_table(source_path: str, workbook_sheet: str) -> List[List[Any]]:
    """Load XLSX or CSV source into a raw table. Raises AppError on failure."""
    ext = os.path.splitext(source_path)[1].lower()
    try:
        if ext == ".csv":
            return load_csv(source_path)
        return load_xlsx(source_path, workbook_sheet)
    except PermissionError:
        raise AppError(
            FILE_LOCKED,
            f"Source file is locked: {source_path}",
            {"path": source_path},
        )
    except ValueError as e:
        raise AppError(SHEET_NOT_FOUND, str(e))
    except AppError:
        raise
    except Exception as e:
        raise AppError(SOURCE_READ_FAILED, f"Failed to read source: {e}")


def _open_or_create_dest(dest_path: str) -> Workbook:
    """Open an existing workbook or create a blank one."""
    try:
        if dest_path and os.path.exists(dest_path):
            return load_workbook(dest_path)
        return Workbook()
    except PermissionError:
        raise AppError(
            FILE_LOCKED,
            f"Destination file is locked: {dest_path}",
            {"path": dest_path},
        )
    except Exception as e:
        raise AppError(
            SOURCE_READ_FAILED,
            f"Could not open destination file: {e}",
        )


def _get_or_create_sheet(wb: Workbook, name: str):
    """Return the named sheet, creating it if absent. Cleans up the default blank sheet."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        default = wb["Sheet"]
        if default.max_row == 1 and default.max_column == 1 and default["A1"].value in (None, ""):
            wb.remove(default)
    return ws


def _apply_row_selection_indexed(
    table: List[List[Any]],
    row_indices: List[int],
) -> List[Tuple[int, List[Any]]]:
    """
    Like apply_row_selection but returns (original_index, row) pairs.
    This lets us track which absolute table indices survive rules filtering.
    """
    return [(i, table[i]) for i in row_indices if 0 <= i < len(table)]


# ── Public API ────────────────────────────────────────────────────────────────

def run_sheet(
    source_path: str,
    sheet_cfg: SheetConfig,
    recipe_name: str = "Recipe",
    _wb_cache: Optional[Dict[str, Workbook]] = None,
) -> SheetResult:
    """
    Execute a single sheet extraction.

    Pipeline:
      1. Load source table
      2. Apply source start row offset
      3. Apply row selection (rows_spec)
      4. Apply rules filtering — runs on full-width rows, absolute source cols
      5. Apply column selection
      6. Shape (pack or keep)
      7. Plan destination placement
      8. Collision check + write

    For keep mode, shape_keep receives only the post-rules absolute row indices
    so that rules correctly exclude rows from the spatial output.

    _wb_cache: optional dict keyed by dest_path. When provided (by batch.run_all),
               the workbook is NOT saved here. When None (standalone), saves here.
    """
    dest_path = sheet_cfg.destination.file_path
    if not (dest_path or "").strip():
        raise AppError(
            MISSING_DEST_PATH,
            "Destination file path is blank.",
        )

    table = _load_source_table(source_path, sheet_cfg.workbook_sheet)
    table = _apply_source_start_row(table, getattr(sheet_cfg, "source_start_row", ""))

    used_h, used_w = compute_used_range(table)
    table = normalize_table(table)

    # Step 3 — row selection: produce (abs_index, row) pairs
    row_indices = parse_rows(sheet_cfg.rows_spec)
    if not row_indices:
        row_indices = list(range(used_h))

    indexed_rows = _apply_row_selection_indexed(table, row_indices)
    # indexed_rows: [(abs_idx, row), ...]

    # Step 4 — rules filtering on full-width rows
    # Pass only the row values to apply_rules, then re-pair with indices.
    rows_only = [row for _, row in indexed_rows]
    filtered_rows = apply_rules(rows_only, sheet_cfg.rules, sheet_cfg.rules_combine)

    # Recover which absolute indices survived by walking indexed_rows in order.
    # apply_rules preserves order and returns a strict subset, so we can
    # consume filtered_rows sequentially while scanning indexed_rows.
    survived_abs_indices: List[int] = []
    filt_pos = 0
    for abs_idx, row in indexed_rows:
        if filt_pos < len(filtered_rows) and filtered_rows[filt_pos] is row:
            survived_abs_indices.append(abs_idx)
            filt_pos += 1

    # Step 5 — column selection
    col_indices = parse_columns(sheet_cfg.columns_spec)
    if not col_indices:
        col_indices = list(range(used_w))

    selected = apply_column_selection(filtered_rows, col_indices)

    # Step 6 — shape
    if sheet_cfg.paste_mode == "keep":
        shaped = shape_keep(table, survived_abs_indices, col_indices)
    else:
        shaped = shape_pack(selected)

    # Steps 7–8 — plan, collision check, write
    standalone = _wb_cache is None

    if standalone:
        wb = _open_or_create_dest(dest_path)
    else:
        if dest_path not in _wb_cache:
            _wb_cache[dest_path] = _open_or_create_dest(dest_path)
        wb = _wb_cache[dest_path]

    ws = _get_or_create_sheet(wb, sheet_cfg.destination.sheet_name)

    plan = build_plan(
        ws, shaped,
        sheet_cfg.destination.start_col,
        sheet_cfg.destination.start_row,
    )
    rows_written = 0
    if plan is not None:
        rows_written = apply_write_plan(ws, shaped, plan)

    if standalone and dest_path:
        try:
            wb.save(dest_path)
        except PermissionError:
            raise AppError(
                FILE_LOCKED,
                f"Destination file is open in another program: {dest_path}",
                {"path": dest_path},
            )
        except Exception as e:
            raise AppError(
                SAVE_FAILED,
                str(e),
                {"path": dest_path},
            )

    return SheetResult(
        source_path=source_path,
        recipe_name=recipe_name,
        sheet_name=sheet_cfg.name,
        dest_file=dest_path,
        dest_sheet=sheet_cfg.destination.sheet_name,
        rows_written=rows_written,
        message="OK" if rows_written > 0 else "0 rows written",
    )
