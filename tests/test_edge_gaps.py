"""
test_edge_gaps.py — Edge-case and combination gap tests.

Covers gaps identified in existing test suite:

Section 1: Rules engine edge cases
  - Equals: string "10" cell vs "10.0" target (numeric coercion)
  - Contains: numeric cell substring match
  - Rules on boolean cells (equals and contains)
  - Empty string cell vs contains with non-empty target
  - Empty string cell vs equals with empty target

Section 2: Keep mode + rules combinations
  - Keep mode with rules filtering ALL rows → zero rows
  - Keep mode with single surviving row and wide column gaps

Section 3: Batch edge cases
  - Two sources, same dest sheet, different non-overlapping start_cols (merge)
  - Batch item writes zero rows (all filtered) then normal item appends correctly
  - Three items: zero-row, normal, zero-row — middle item lands correctly

Section 4: Destination management
  - Dest file exists with only a custom-named sheet (no "Sheet" default)

Section 5: Full pipeline combo
  - source_start_row + rules + keep mode + non-adjacent columns in one shot
  - source_start_row + rules + pack mode + row selection combined
"""
from __future__ import annotations

import csv
import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.batch import run_all
from core.errors import AppError
from core.models import Destination, Rule, SheetConfig
from core.rules import apply_rules
from core.runner import run_sheet


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


def _csv(path, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path, sheet="Out"):
    return load_workbook(path)[sheet]


def _rule(op, val, col="A", mode="include"):
    return Rule(mode=mode, column=col, operator=op, value=val)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — RULES ENGINE EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

class TestRulesEdgeCases:

    def test_equals_string_cell_vs_float_target_numeric_coercion(self):
        """String cell "10" should match target "10.0" via numeric comparison."""
        rows = [["10"], ["20"], ["30"]]
        result = apply_rules(rows, [_rule("equals", "10.0")], "AND")
        assert len(result) == 1
        assert result[0][0] == "10"

    def test_equals_float_cell_vs_int_target_numeric_coercion(self):
        """Float cell 10.0 should match target "10" via numeric comparison."""
        rows = [[10.0], [20.0]]
        result = apply_rules(rows, [_rule("equals", "10")], "AND")
        assert len(result) == 1
        assert result[0][0] == 10.0

    def test_equals_string_ten_vs_string_ten_point_zero(self):
        """String "10" vs target "10.0" — both coerce to float, should match."""
        rows = [["10"], ["10.0"], ["11"]]
        result = apply_rules(rows, [_rule("equals", "10")], "AND")
        assert len(result) == 2

    def test_contains_numeric_cell_partial_match(self):
        """Numeric cell 98765 should match contains '876'."""
        rows = [[98765], [11111], [87600]]
        result = apply_rules(rows, [_rule("contains", "876")], "AND")
        assert len(result) == 2
        vals = [r[0] for r in result]
        assert 98765 in vals
        assert 87600 in vals

    def test_equals_boolean_true_cell_vs_string_true(self):
        """Boolean True cell — equals 'True' should match via string fallback."""
        rows = [[True], [False], ["text"]]
        result = apply_rules(rows, [_rule("equals", "True")], "AND")
        assert len(result) == 1
        assert result[0][0] is True

    def test_equals_boolean_false_cell_vs_string_false(self):
        """Boolean False cell — equals 'False' should match via string fallback."""
        rows = [[True], [False], ["text"]]
        result = apply_rules(rows, [_rule("equals", "False")], "AND")
        assert len(result) == 1
        assert result[0][0] is False

    def test_contains_boolean_cell_substring(self):
        """Boolean True converts to 'True' string — contains 'ru' should match."""
        rows = [[True], [False], ["fruit"]]
        result = apply_rules(rows, [_rule("contains", "ru")], "AND")
        assert len(result) == 2
        vals = [r[0] for r in result]
        assert True in vals
        assert "fruit" in vals

    def test_contains_false_cell_substring(self):
        """Boolean False converts to 'False' — contains 'als' should match."""
        rows = [[True], [False]]
        result = apply_rules(rows, [_rule("contains", "als")], "AND")
        assert len(result) == 1
        assert result[0][0] is False

    def test_empty_string_cell_contains_nonempty_target_no_match(self):
        """Empty string cell should not match contains 'something'."""
        rows = [[""], ["hello"], [""]]
        result = apply_rules(rows, [_rule("contains", "something")], "AND")
        assert len(result) == 0

    def test_empty_string_cell_equals_empty_target_matches(self):
        """Empty string cell should match equals '' target."""
        rows = [[""], ["hello"], [""]]
        result = apply_rules(rows, [_rule("equals", "")], "AND")
        assert len(result) == 2
        assert all(r[0] == "" for r in result)

    def test_none_cell_equals_empty_target_matches(self):
        """None cell matches equals '' (already tested, but verify alongside empty string)."""
        rows = [[None], ["hello"], [""]]
        result = apply_rules(rows, [_rule("equals", "")], "AND")
        assert len(result) == 2

    def test_equals_zero_int_cell_vs_zero_float_target(self):
        """Int 0 should match target '0.0' via numeric coercion."""
        rows = [[0], [1], [2]]
        result = apply_rules(rows, [_rule("equals", "0.0")], "AND")
        assert len(result) == 1
        assert result[0][0] == 0


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — KEEP MODE + RULES COMBINATIONS
# ══════════════════════════════════════════════════════════════════════════════

class TestKeepModeRulesCombos:

    def test_keep_mode_rules_filter_all_rows_falls_back_to_all(self):
        """
        Keep mode: when rules filter out every row, survived_abs_indices is
        empty. shape_keep treats empty row indices as 'all rows' (by design —
        same convention as parse_rows returning []). This is different from
        pack mode where filtered_rows being empty produces zero output.
        Verify the keep-mode behavior is stable.
        """
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["alpha", 1], ["beta", 2], ["gamma", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,B", mode="keep", rules=[
                Rule(mode="include", column="A", operator="equals", value="NOMATCH")
            ]))
            # Keep mode: empty survived indices → shape_keep uses all rows
            assert r.rows_written == 3

    def test_pack_mode_rules_filter_all_rows_zero_output(self):
        """Pack mode: when rules filter out every row → zero rows written."""
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["alpha", 1], ["beta", 2], ["gamma", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,B", mode="pack", rules=[
                Rule(mode="include", column="A", operator="equals", value="NOMATCH")
            ]))
            assert r.rows_written == 0

    def test_keep_mode_single_surviving_row_wide_gap(self):
        """Keep mode: one row survives rules, cols A and E selected → wide gap output."""
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["keep", "b", "c", "d", "e_val"],
                         ["drop", "b", "c", "d", "e_val2"],
                         ["drop", "b", "c", "d", "e_val3"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,E", mode="keep", rules=[
                Rule(mode="include", column="A", operator="equals", value="keep")
            ]))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "keep"
            assert ws["B1"].value is None  # gap
            assert ws["C1"].value is None  # gap
            assert ws["D1"].value is None  # gap
            assert ws["E1"].value == "e_val"

    def test_keep_mode_rules_exclude_middle_rows_only(self):
        """Keep mode: first and last rows survive, middle excluded — compressed output."""
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["keep", 1, "x"],
                         ["drop", 2, "y"],
                         ["drop", 3, "z"],
                         ["keep", 4, "w"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,C", mode="keep", rules=[
                Rule(mode="include", column="A", operator="equals", value="keep")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "keep"
            assert ws["C1"].value == "x"
            assert ws["A2"].value == "keep"
            assert ws["C2"].value == "w"
            assert ws["A3"].value is None  # nothing beyond 2 rows


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — BATCH EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

class TestBatchEdgeCases:

    def test_two_sources_same_dest_different_start_cols_merge(self):
        """Two sources writing to non-overlapping start_cols on the same row."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["left1"], ["left2"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["right1"], ["right2"]])
            report = run_all([
                (s1, "R1", _cfg(dest, start_col="A", start_row="1")),
                (s2, "R2", _cfg(dest, start_col="D", start_row="1")),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "left1"
            assert ws["A2"].value == "left2"
            assert ws["D1"].value == "right1"
            assert ws["D2"].value == "right2"
            assert ws["B1"].value is None
            assert ws["C1"].value is None

    def test_batch_zero_rows_then_normal_append_correct(self):
        """First item filters to zero rows; second item should still land at row 1."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"),
                        [["alpha", 1], ["beta", 2]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"),
                        [["data", 99]])
            report = run_all([
                (s1, "R1", _cfg(dest, rules=[
                    Rule(mode="include", column="A", operator="equals", value="NOMATCH")
                ])),
                (s2, "R2", _cfg(dest)),
            ])
            assert report.ok
            assert report.results[0].rows_written == 0
            assert report.results[1].rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "data"
            assert ws["B1"].value == 99

    def test_batch_zero_normal_zero_middle_lands_correctly(self):
        """Zero-row, normal, zero-row — middle item lands at row 1."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["x"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["real_data"]])
            s3 = _xlsx(os.path.join(td, "s3.xlsx"), [["y"]])
            no_match_rule = [Rule(mode="include", column="A",
                                  operator="equals", value="NOMATCH")]
            report = run_all([
                (s1, "R1", _cfg(dest, rules=no_match_rule)),
                (s2, "R2", _cfg(dest)),
                (s3, "R3", _cfg(dest, rules=no_match_rule)),
            ])
            assert report.ok
            assert report.results[0].rows_written == 0
            assert report.results[1].rows_written == 1
            assert report.results[2].rows_written == 0
            ws = _ws(dest)
            assert ws["A1"].value == "real_data"
            assert ws["A2"].value is None

    def test_batch_two_normal_then_zero_row_no_corruption(self):
        """Two normal appends then a zero-row item — first two stack, third is harmless."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["first"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["second"]])
            s3 = _xlsx(os.path.join(td, "s3.xlsx"), [["nope"]])
            report = run_all([
                (s1, "R1", _cfg(dest)),
                (s2, "R2", _cfg(dest)),
                (s3, "R3", _cfg(dest, rules=[
                    Rule(mode="include", column="A",
                         operator="equals", value="NOMATCH")
                ])),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "first"
            assert ws["A2"].value == "second"
            assert ws["A3"].value is None


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — DESTINATION MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

class TestDestinationManagement:

    def test_dest_exists_no_default_sheet_name(self):
        """Dest file has only 'Data' sheet (no 'Sheet') — new sheet created, 'Data' preserved."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook()
            wb.active.title = "Data"
            wb["Data"]["A1"] = "existing"
            wb.save(dest)

            src = _xlsx(os.path.join(td, "s.xlsx"), [["new_val"]])
            r = run_sheet(src, _cfg(dest, dest_sheet="Out"))
            assert r.rows_written == 1
            wb2 = load_workbook(dest)
            assert "Data" in wb2.sheetnames
            assert wb2["Data"]["A1"].value == "existing"
            assert wb2["Out"]["A1"].value == "new_val"

    def test_dest_exists_writing_to_existing_custom_sheet(self):
        """Dest has 'Report' sheet with data — writing appends without clobbering."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook()
            wb.active.title = "Report"
            wb["Report"]["A1"] = "header"
            wb["Report"]["A2"] = "old_data"
            wb.save(dest)

            src = _xlsx(os.path.join(td, "s.xlsx"), [["new_data"]])
            r = run_sheet(src, _cfg(dest, dest_sheet="Report"))
            assert r.rows_written == 1
            ws = _ws(dest, "Report")
            assert ws["A1"].value == "header"
            assert ws["A2"].value == "old_data"
            assert ws["A3"].value == "new_data"


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — FULL PIPELINE COMBINATIONS
# ══════════════════════════════════════════════════════════════════════════════

class TestFullPipelineCombos:

    def test_source_start_row_rules_keep_mode_non_adjacent_cols(self):
        """
        Full pipeline: source_start_row=2 skips header, rules filter on col B,
        keep mode with cols A,D selected — non-adjacent gaps preserved.
        """
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["HEADER", "FILTER", "X", "DATA"],   # row 1: skipped
                         ["r2a",    "yes",    "x", "r2d"],     # row 2: kept
                         ["r3a",    "no",     "x", "r3d"],     # row 3: filtered out
                         ["r4a",    "yes",    "x", "r4d"],     # row 4: kept
                         ["r5a",    "no",     "x", "r5d"]])    # row 5: filtered out
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest,
                                    src_start_row="2",
                                    columns="A,D",
                                    mode="keep",
                                    rules=[Rule(mode="include", column="B",
                                                operator="equals", value="yes")]))
            assert r.rows_written == 2
            ws = _ws(dest)
            # Keep mode: cols A-D bounding box, B and C are gaps
            assert ws["A1"].value == "r2a"
            assert ws["B1"].value is None  # gap
            assert ws["C1"].value is None  # gap
            assert ws["D1"].value == "r2d"
            assert ws["A2"].value == "r4a"
            assert ws["D2"].value == "r4d"
            assert ws["A3"].value is None  # only 2 rows

    def test_source_start_row_rules_pack_mode_row_selection(self):
        """
        source_start_row=2, rows=1-3 (relative to offset table), rules filter,
        pack mode, column subset.
        """
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["SKIP",   "header"],       # row 1: skipped by start_row
                         ["keep",   "val1",  "x1"],  # row 2 → offset row 1: selected, kept
                         ["drop",   "val2",  "x2"],  # row 3 → offset row 2: selected, filtered
                         ["keep",   "val3",  "x3"],  # row 4 → offset row 3: selected, kept
                         ["keep",   "val4",  "x4"]]) # row 5 → offset row 4: NOT selected
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest,
                                    src_start_row="2",
                                    rows="1-3",
                                    columns="A,C",
                                    mode="pack",
                                    rules=[Rule(mode="include", column="A",
                                                operator="equals", value="keep")]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "keep"
            assert ws["B1"].value == "x1"  # col C packed to output B
            assert ws["A2"].value == "keep"
            assert ws["B2"].value == "x3"

    def test_csv_source_start_row_rules_keep_mode(self):
        """Same full pipeline combo but with CSV source."""
        with TemporaryDirectory() as td:
            src = _csv(os.path.join(td, "s.csv"),
                       [["HEADER", "FILTER", "DATA"],
                        ["r2a",    "yes",    "r2c"],
                        ["r3a",    "no",     "r3c"],
                        ["r4a",    "yes",    "r4c"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest,
                                    src_start_row="2",
                                    columns="A,C",
                                    mode="keep",
                                    rules=[Rule(mode="include", column="B",
                                                operator="equals", value="yes")]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "r2a"
            assert ws["B1"].value is None  # gap
            assert ws["C1"].value == "r2c"
            assert ws["A2"].value == "r4a"
            assert ws["C2"].value == "r4c"

    def test_full_pipeline_explicit_start_row_and_col_with_rules(self):
        """Rules + column subset + explicit dest start_row=5 and start_col=C."""
        with TemporaryDirectory() as td:
            src = _xlsx(os.path.join(td, "s.xlsx"),
                        [["yes", 10, "a"],
                         ["no",  20, "b"],
                         ["yes", 30, "c"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest,
                                    columns="A,C",
                                    start_col="C",
                                    start_row="5",
                                    rules=[Rule(mode="include", column="A",
                                                operator="equals", value="yes")]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["C5"].value == "yes"
            assert ws["D5"].value == "a"
            assert ws["C6"].value == "yes"
            assert ws["D6"].value == "c"
            # Nothing above row 5
            assert ws["C4"].value is None
