"""
Integration tests: paste_mode x rules x destination combinations.

Covers every meaningful interaction between:
  - paste_mode:  pack | keep
  - rules:       none / include / exclude / AND / OR / rule-col-not-in-output
  - columns:     all / subset / single / range / non-contiguous
  - rows:        all / subset / non-contiguous
  - destination: append / explicit-row / new-file / stacked-runs
  - source:      xlsx / csv / source_start_row
"""
from __future__ import annotations

import csv
import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet, run_all
from core.errors import AppError, DEST_BLOCKED
from core.models import Destination, Rule, SheetConfig


# ── helpers ──────────────────────────────────────────────────────────────────

def _xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data or [], 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _csv(path: str, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)


def _cfg(dest_path, *, columns="", rows="", mode="pack",
         rules=None, combine="AND", start_col="A", start_row="",
         dest_sheet="Out", src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path: str, sheet: str = "Out"):
    return load_workbook(path)[sheet]


# ── 1. Baseline: pack, no rules, all cols, append ────────────────────────────

def test_pack_no_rules_all_cols_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", 1], ["beta", 2]])

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "alpha"
        assert ws["B2"].value == 2


# ── 2. Baseline: keep, no rules, subset cols/rows, append ────────────────────

def test_keep_no_rules_subset_cols_rows_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["a", "b", "c"],
            ["d", "e", "f"],
            ["g", "h", "i"],
        ])
        # rows 1 and 3 (indices 0,2), cols A and C (indices 0,2)
        cfg = _cfg(dest, columns="A,C", rows="1,3", mode="keep")
        result = run_sheet(src, cfg)
        assert result.rows_written == 3   # bounding box rows 1–3

        ws = _ws(dest)
        assert ws["A1"].value == "a"
        assert ws["B1"].value is None     # col B is gap
        assert ws["C1"].value == "c"
        assert ws["A2"].value is None     # row 2 is gap
        assert ws["B2"].value is None
        assert ws["C2"].value is None
        assert ws["A3"].value == "g"
        assert ws["C3"].value == "i"


# ── 3. Pack + include rule + subset cols + append ────────────────────────────

def test_pack_include_rule_subset_cols_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["keep",  10, "x"],
            ["drop",  20, "y"],
            ["keep",  30, "z"],
        ])
        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="include", column="A", operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2

        ws = _ws(dest)
        assert ws["A1"].value == "keep"
        assert ws["B1"].value == 10
        assert ws["A2"].value == "keep"
        assert ws["B2"].value == 30


# ── 4. Pack + exclude rule + subset cols + append ────────────────────────────

def test_pack_exclude_rule_subset_cols_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["keep",  10],
            ["drop",  20],
            ["keep",  30],
        ])
        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="exclude", column="A", operator="equals", value="drop")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "keep"
        assert ws["A2"].value == "keep"


# ── 5. Pack + rule on col NOT in output + append ─────────────────────────────

def test_pack_rule_col_not_in_output_append():
    """Filter on col C (hidden); output only A and B."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["alpha", 1, "YES"],
            ["beta",  2, "NO"],
            ["gamma", 3, "YES"],
        ])
        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="include", column="C", operator="equals", value="YES")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "alpha"
        assert ws["B1"].value == 1
        assert ws["A2"].value == "gamma"
        assert ws["B2"].value == 3
        assert ws["C1"].value is None   # col C never written


# ── 6. Keep + include rule + subset cols + explicit row ──────────────────────

def test_keep_include_rule_subset_cols_explicit_row():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["keep",  1, "tag"],
            ["drop",  2, "tag"],
            ["keep",  3, "tag"],
        ])
        cfg = _cfg(dest, columns="A,C", rows="1,3", mode="keep",
                   start_row="10",
                   rules=[Rule(mode="include", column="A", operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        # For keep mode, shape_keep uses the ORIGINAL table + row_indices + col_indices.
        # Rules filter table_rows but keep mode ignores the filtered result and
        # uses original_rows directly — preserving the original spacing contract.
        # rows 1,3 → bbox rows 1-3 → 3 rows written regardless of rule.
        assert result.rows_written == 3

        ws = _ws(dest)
        # Output starts at explicit row 10; cols A and C (indices 0,2 → 2-col gap)
        assert ws["A10"].value == "keep"
        assert ws["B10"].value is None   # col B is gap
        assert ws["C10"].value == "tag"
        assert ws["A11"].value is None   # row 2 of bbox is gap (not selected)
        assert ws["A12"].value == "keep"
        assert ws["C12"].value == "tag"


# ── 7. Pack + AND rules → 0 rows → nothing written ──────────────────────────

def test_pack_and_rules_zero_rows():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", 10], ["beta", 20]])

        # Impossible to satisfy both: A=="alpha" AND A=="beta"
        cfg = _cfg(dest, rules=[
            Rule(mode="include", column="A", operator="equals", value="alpha"),
            Rule(mode="include", column="A", operator="equals", value="beta"),
        ], combine="AND")
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


# ── 8. Pack + OR rules + subset cols + append ────────────────────────────────

def test_pack_or_rules_subset_cols_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["alpha", 10],
            ["beta",  20],
            ["gamma", 30],
        ])
        cfg = _cfg(dest, columns="A,B", combine="OR", rules=[
            Rule(mode="include", column="A", operator="equals", value="alpha"),
            Rule(mode="include", column="B", operator=">", value="25"),
        ])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "alpha"
        assert ws["A2"].value == "gamma"


# ── 9. Pack + no rules + collision → DEST_BLOCKED ────────────────────────────

def test_pack_no_rules_collision_blocked():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["data", 1]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A5"] = "BLOCK"
        wb.save(dest)

        cfg = _cfg(dest, start_row="5")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == DEST_BLOCKED


# ── 10. Keep + collision on row 2 of shape → DEST_BLOCKED ───────────────────

def test_keep_collision_on_interior_row_blocked():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["a", "b"], ["c", "d"], ["e", "f"]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"   # shape is 3 rows, row 2 blocked
        wb.save(dest)

        cfg = _cfg(dest, mode="keep", start_row="1")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == DEST_BLOCKED


# ── 11. Pack + rules → 0 rows → no collision check ──────────────────────────

def test_pack_rules_zero_rows_no_collision_even_if_dest_occupied():
    """0 rows means no WritePlan, so a blocked destination is irrelevant."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", 1]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)

        cfg = _cfg(dest, start_row="1",
                   rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")])
        result = run_sheet(src, cfg)  # must NOT raise
        assert result.rows_written == 0


# ── 12. Keep + rules — keep ignores rule filter, always writes full bbox ────────
# shape_keep() uses the ORIGINAL table + row_indices + col_indices directly.
# Rules filter table_rows but keep mode bypasses that entirely.
# Contract: keep always preserves original spacing, rules have no effect on output.

def test_keep_rules_does_not_filter_output():
    """keep mode writes the full bounding box regardless of rules."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", "b"], ["beta", "d"]])

        cfg = _cfg(dest, mode="keep",
                   rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")])
        result = run_sheet(src, cfg)
        # Rule matches nothing, but keep ignores it — full 2-row bbox is written
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "alpha"
        assert ws["A2"].value == "beta"


# ── 13. Pack + two runs stack correctly (append landing-zone aware) ───────────

def test_pack_two_runs_append_stack():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["val1", 1], ["val2", 2]])

        cfg = _cfg(dest, columns="A,B")
        run_sheet(src, cfg)
        run_sheet(src, cfg)

        ws = _ws(dest)
        assert ws["A1"].value == "val1"
        assert ws["A3"].value == "val1"   # second run appends at row 3
        assert ws["A4"].value == "val2"


# ── 14. Keep + two runs stack correctly ──────────────────────────────────────

def test_keep_two_runs_append_stack():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["a", "b"], [None, "d"], ["e", "f"]])

        # keep mode: 3-row bounding box (rows 1,3 selected → bbox rows 1-3)
        cfg = _cfg(dest, columns="A,B", rows="1,3", mode="keep")
        run_sheet(src, cfg)
        run_sheet(src, cfg)

        ws = _ws(dest)
        # First run: rows 1-3; second run appends at row 4
        assert ws["A1"].value == "a"
        assert ws["A2"].value is None    # gap row
        assert ws["A3"].value == "e"
        assert ws["A4"].value == "a"    # second run starts here
        assert ws["A6"].value == "e"


# ── 15. Pack + rule on hidden col → result 0 rows ────────────────────────────

def test_pack_rule_hidden_col_no_match_zero_rows():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", 1, "NO"], ["beta", 2, "NO"]])

        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="include", column="C", operator="equals", value="YES")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


# ── 16. Pack + column range spec (A-C) + rules + append ──────────────────────

def test_pack_column_range_spec_with_rules_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["keep", "x", 10, "extra"],
            ["drop", "y", 20, "extra"],
            ["keep", "z", 30, "extra"],
        ])
        # A-C outputs cols A,B,C; col D excluded
        cfg = _cfg(dest, columns="A-C",
                   rules=[Rule(mode="include", column="A", operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2

        ws = _ws(dest)
        assert ws["A1"].value == "keep"
        assert ws["B1"].value == "x"
        assert ws["C1"].value == 10
        assert ws["D1"].value is None   # col D not written


# ── 17. Keep + non-contiguous cols and rows → gaps preserved ─────────────────

def test_keep_non_contiguous_cols_rows_gaps_preserved():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["A1", "B1", "C1", "D1", "E1"],
            ["A2", "B2", "C2", "D2", "E2"],
            ["A3", "B3", "C3", "D3", "E3"],
            ["A4", "B4", "C4", "D4", "E4"],
            ["A5", "B5", "C5", "D5", "E5"],
        ])
        # Rows 1,3,5 — cols A,C,E — bounding box is 5×5 with gaps
        cfg = _cfg(dest, columns="A,C,E", rows="1,3,5", mode="keep")
        result = run_sheet(src, cfg)
        assert result.rows_written == 5  # bounding box rows 1-5

        ws = _ws(dest)
        # Row 1 (selected): A,C,E present; B,D are gaps
        assert ws["A1"].value == "A1"
        assert ws["B1"].value is None
        assert ws["C1"].value == "C1"
        assert ws["D1"].value is None
        assert ws["E1"].value == "E1"
        # Row 2 (not selected): all None
        for col in ["A","B","C","D","E"]:
            assert ws[f"{col}2"].value is None
        # Row 3 (selected)
        assert ws["A3"].value == "A3"
        assert ws["C3"].value == "C3"
        assert ws["E3"].value == "E3"
        # Row 5 (selected)
        assert ws["A5"].value == "A5"


# ── 18. Pack + source_start_row + rules + subset cols + append ───────────────

def test_pack_source_start_row_rules_subset_cols_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["HEADER", "HEADER"],   # row 1 — skip via source_start_row=2
            ["keep",   10],
            ["drop",   20],
            ["keep",   30],
        ])
        cfg = _cfg(dest, columns="A,B", src_start_row="2",
                   rules=[Rule(mode="include", column="A", operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2

        ws = _ws(dest)
        assert ws["A1"].value == "keep"
        assert ws["B1"].value == 10
        assert ws["A2"].value == "keep"
        assert ws["B2"].value == 30


# ── 19. CSV source + rules + pack + append ───────────────────────────────────

def test_csv_source_rules_pack_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")
        _csv(src, [["alpha", "10"], ["beta", "20"], ["gamma", "30"]])

        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="include", column="B", operator=">", value="15")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2

        ws = _ws(dest)
        assert ws["A1"].value == "beta"
        assert ws["A2"].value == "gamma"


# ── 20. Pack + rules + destination sheet auto-created (new file) ─────────────

def test_pack_rules_new_dest_file_auto_created():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "new_dest.xlsx")  # does not exist yet
        _xlsx(src, data=[["alpha", 1], ["beta", 2]])

        assert not os.path.exists(dest)
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A", operator="equals", value="alpha")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        assert os.path.exists(dest)
        assert _ws(dest)["A1"].value == "alpha"


# ── 21. Pack + single-row result + explicit large start row ──────────────────

def test_pack_single_row_explicit_large_start_row():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["only"]])

        cfg = _cfg(dest, start_row="999")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        assert _ws(dest)["A999"].value == "only"


# ── 22. Pack + include-all rule (always true) = same as no rules ─────────────

def test_pack_include_all_rule_same_as_no_rules():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest1 = os.path.join(td, "d1.xlsx")
        dest2 = os.path.join(td, "d2.xlsx")
        data = [["alpha", 1], ["beta", 2], ["gamma", 3]]
        _xlsx(src, data=data)

        # contains "" matches everything (every string contains empty string)
        cfg_rule = _cfg(dest1,
                        rules=[Rule(mode="include", column="A", operator="contains", value="")])
        cfg_none = _cfg(dest2)

        r1 = run_sheet(src, cfg_rule)
        r2 = run_sheet(src, cfg_none)
        assert r1.rows_written == r2.rows_written == 3


# ── 23. Pack + exclude-all rule → 0 rows ─────────────────────────────────────

def test_pack_exclude_all_rule_zero_rows():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["alpha", 1], ["beta", 2]])

        # Exclude rows where A contains "" = exclude everything
        cfg = _cfg(dest,
                   rules=[Rule(mode="exclude", column="A", operator="contains", value="")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


# ── 24. Pack + contains rule (case insensitive) + subset cols ────────────────

def test_pack_contains_case_insensitive_subset_cols():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["ALPHA_val", 1],
            ["beta_val",  2],
            ["Alpha_val", 3],
        ])
        cfg = _cfg(dest, columns="A",
                   rules=[Rule(mode="include", column="A", operator="contains", value="alpha")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert "ALPHA" in str(ws["A1"].value).upper()
        assert "ALPHA" in str(ws["A2"].value).upper()


# ── 25. Pack + numeric > rule on col in output + append ──────────────────────

def test_pack_numeric_greater_than_col_in_output():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["a", 5], ["b", 15], ["c", 25]])

        cfg = _cfg(dest, columns="A,B",
                   rules=[Rule(mode="include", column="B", operator=">", value="10")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["B1"].value == 15
        assert ws["B2"].value == 25


# ── 26. Pack + numeric < rule on col NOT in output ───────────────────────────

def test_pack_numeric_less_than_col_not_in_output():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[["a", 5, "x"], ["b", 15, "y"], ["c", 3, "z"]])

        # Filter by col B (< 10) but output only A and C
        cfg = _cfg(dest, columns="A,C",
                   rules=[Rule(mode="include", column="B", operator="<", value="10")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = _ws(dest)
        assert ws["A1"].value == "a"
        assert ws["B1"].value == "x"
        assert ws["A2"].value == "c"
        assert ws["B2"].value == "z"


# ── 27. Multi-run: pack then keep to same dest → correct stacking ─────────────

def test_pack_then_keep_same_dest_stack():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src1, data=[["pack1", 1], ["pack2", 2]])
        _xlsx(src2, data=[["keep1", "x", "A"], ["keep2", "y", "B"], ["keep3", "z", "C"]])

        cfg_pack = _cfg(dest, columns="A,B", mode="pack")
        cfg_keep = _cfg(dest, columns="A,C", rows="1,3", mode="keep")

        r1 = run_sheet(src1, cfg_pack)
        r2 = run_sheet(src2, cfg_keep)
        assert r1.rows_written == 2
        assert r2.rows_written == 3   # keep bounding box 3 rows

        ws = _ws(dest)
        # Pack run at rows 1-2
        assert ws["A1"].value == "pack1"
        assert ws["A2"].value == "pack2"
        # Keep run appends after max used row in A:C (keep width=3) = 2 → starts at row 3
        # shape_keep(src2, rows=[0,2], cols=[0,2]) → 3 rows, 3 cols (A..C)
        # row1=[keep1,None,A], row2=[None,None,None], row3=[keep3,None,C]
        assert ws["A3"].value == "keep1"
        assert ws["B3"].value is None   # gap (col B not selected)
        assert ws["C3"].value == "A"    # col C (index 2) = src2 row1 col3 = "A"
        assert ws["A4"].value is None   # gap row
        assert ws["A5"].value == "keep3"
        assert ws["C5"].value == "C"    # src2 row3 col3 = "C"


# ── 28. Multi-run: keep then pack to same dest → correct stacking ─────────────

def test_keep_then_pack_same_dest_stack():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src1, data=[["k1", "x"], [None, None], ["k3", "z"]])
        _xlsx(src2, data=[["p1", 1], ["p2", 2]])

        cfg_keep = _cfg(dest, columns="A,B", rows="1,3", mode="keep")
        cfg_pack = _cfg(dest, columns="A,B", mode="pack")

        r1 = run_sheet(src1, cfg_keep)
        r2 = run_sheet(src2, cfg_pack)
        assert r1.rows_written == 3   # bounding box rows 1-3
        assert r2.rows_written == 2

        ws = _ws(dest)
        assert ws["A1"].value == "k1"
        assert ws["A2"].value is None   # gap
        assert ws["A3"].value == "k3"
        # Pack appends after row 3
        assert ws["A4"].value == "p1"
        assert ws["A5"].value == "p2"


# ── 29. Multiple AND rules all satisfied ─────────────────────────────────────

def test_pack_multiple_and_rules_all_satisfied():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["alpha", 50, "YES"],
            ["alpha", 5,  "YES"],  # B too low
            ["beta",  50, "YES"],  # A wrong
            ["alpha", 60, "NO"],   # C wrong
            ["alpha", 55, "YES"],  # matches all
        ])
        cfg = _cfg(dest, combine="AND", rules=[
            Rule(mode="include", column="A", operator="equals",   value="alpha"),
            Rule(mode="include", column="B", operator=">",        value="40"),
            Rule(mode="include", column="C", operator="equals",   value="YES"),
        ])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2  # rows 0 and 4 (0-based)
        ws = _ws(dest)
        assert ws["B1"].value == 50
        assert ws["B2"].value == 55


# ── 30. Pack + exclude + include OR combo ────────────────────────────────────

def test_pack_mixed_include_exclude_or():
    """OR: include A==alpha OR exclude B>40 (i.e. keep rows where B<=40)."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        _xlsx(src, data=[
            ["alpha",  50],   # include (A==alpha) → kept
            ["beta",   10],   # exclude(B>40) → not excluded → kept
            ["gamma",  60],   # neither → not kept
            ["alpha",  80],   # include (A==alpha) → kept
        ])
        cfg = _cfg(dest, combine="OR", rules=[
            Rule(mode="include", column="A", operator="equals", value="alpha"),
            Rule(mode="exclude", column="B", operator=">",      value="40"),
        ])
        result = run_sheet(src, cfg)
        # Row 0: include matches → kept
        # Row 1: exclude(B>40)=False(10 not >40) → not excluded → True → kept
        # Row 2: include=False, exclude(60>40)=True so exclude fires → not kept
        # Row 3: include matches → kept
        assert result.rows_written == 3
        ws = _ws(dest)
        assert ws["A1"].value == "alpha"
        assert ws["A2"].value == "beta"
        assert ws["A3"].value == "alpha"
