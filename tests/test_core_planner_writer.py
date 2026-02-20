"""Tests for core.planner (write plan / collision detection) and core.writer (cell writing)."""
import pytest
from openpyxl import Workbook

from core.planner import build_plan
from core.writer import apply_write_plan
from core.errors import AppError, DEST_BLOCKED


# ---- Planner ----

def _fresh_ws():
    wb = Workbook()
    return wb, wb.active


def test_append_uses_max_used_row_across_landing_cols():
    wb, ws = _fresh_ws()
    ws["D5"] = "x"
    ws["E10"] = "y"
    ws["A100"] = "zzz"   # outside landing zone — must not affect result

    shaped = [["a", "b"], ["c", "d"]]
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 11
    assert plan.start_col == 4
    assert plan.landing_cols == (4, 5)
    assert plan.landing_rows == (11, 12)


def test_append_skips_past_any_used_cells_in_landing_zone():
    wb, ws = _fresh_ws()
    ws["D3"] = "x"
    ws["E4"] = "BLOCK"

    shaped = [["a", "b"], ["c", "d"]]
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 5


def test_explicit_start_row_collision_probe_blocks():
    wb, ws = _fresh_ws()
    ws["D50"] = "BLOCK"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"]], start_col_letters="D", start_row_str="50")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["target_start"] == "D50"


def test_build_plan_returns_none_for_empty_shaped():
    wb, ws = _fresh_ws()
    assert build_plan(ws, [], start_col_letters="A", start_row_str="") is None
    assert build_plan(ws, [[]], start_col_letters="A", start_row_str="") is None


# ---- Writer ----

def test_writer_writes_exact_rectangle():
    wb, ws = _fresh_ws()
    shaped = [["a", "b"], ["c", "d"]]

    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    rows_written = apply_write_plan(ws, shaped, plan)

    assert rows_written == 2
    assert ws["C1"].value == "a"
    assert ws["D1"].value == "b"
    assert ws["C2"].value == "c"
    assert ws["D2"].value == "d"


def test_writer_appends_after_existing_data():
    wb, ws = _fresh_ws()
    ws["C1"] = "existing"
    ws["D3"] = "also existing"

    shaped = [["x", "y"]]
    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    apply_write_plan(ws, shaped, plan)

    # max used row across C:D is 3 → append at 4
    assert ws["C4"].value == "x"
    assert ws["D4"].value == "y"


# ---- Errors ----

def test_app_error_str_includes_code_message():
    from core.errors import AppError
    e = AppError("X", "Nope")
    assert str(e).startswith("X: Nope")


def test_app_error_str_includes_details_when_present():
    from core.errors import AppError
    e = AppError("X", "Nope", {"a": 1})
    s = str(e)
    assert "X: Nope" in s
    assert "a" in s
