"""
test_runner.py — Unit tests for core.runner (single-sheet extraction).

Covers:
  - run_sheet: XLSX/CSV extraction, pack/keep modes, rules, append, error paths
  - _apply_source_start_row: offset logic, validation
  - _load_source_table: missing sheet, bad path
  - _open_or_create_dest / _get_or_create_sheet: workbook management
  - Pipeline ordering: rules use absolute source columns (pre-column-selection)
  - Append / landing-zone isolation
  - Collision detection end-to-end (run_sheet level)
  - Edge cases: wide sheets, unicode, mixed types, CSV quirks
  - Pack/Keep combination matrix
"""
from __future__ import annotations

import csv
import os
import time
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.runner import run_sheet
from core.errors import AppError, BAD_SPEC, DEST_BLOCKED, SHEET_NOT_FOUND
from core.models import Destination, Rule, SheetConfig


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _make_xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if data:
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)
    return path


def _make_csv(path: str, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path, sheet="Out"):
    return load_workbook(path)[sheet]


def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


_STD_DATA = [
    ["alpha", "x", 1],
    ["beta",  "y", 2],
    ["gamma", "z", 3],
    ["delta", "w", 4],
]


# ══════════════════════════════════════════════════════════════════════════════
# BASIC EXTRACTION — PACK MODE
# ══════════════════════════════════════════════════════════════════════════════

def test_run_sheet_basic_xlsx():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=_STD_DATA)
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,B"))
        assert result.rows_written == 4
        ws2 = _ws(dest)
        assert ws2["A1"].value == "alpha"
        assert ws2["B1"].value == "x"


def test_run_sheet_csv_source():
    with TemporaryDirectory() as td:
        src  = _make_csv(os.path.join(td, "src.csv"), _STD_DATA)
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,B"))
        assert result.rows_written == 4
        ws2 = _ws(dest)
        assert ws2["A1"].value == "alpha"
        assert ws2["B1"].value == "x"


def test_run_sheet_all_columns_when_blank_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b", "c"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns=""))
        assert result.rows_written == 1
        assert _ws(dest)["C1"].value == "c"


def test_run_sheet_result_message_ok():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.message == "OK"


def test_run_sheet_result_message_zero_rows():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, rules=[Rule(mode="include", column="A",
                                      operator="equals", value="NO_MATCH")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 0
        assert result.message == "0 rows written"


# ══════════════════════════════════════════════════════════════════════════════
# SOURCE START ROW
# ══════════════════════════════════════════════════════════════════════════════

def test_run_sheet_source_start_row_offset():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["header"], ["row1"], ["row2"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, src_start_row="2"))
        assert result.rows_written == 2
        assert _ws(dest)["A1"].value == "row1"


def test_run_sheet_source_start_row_1_same_as_no_offset():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"], ["b"]])
        dest  = os.path.join(td, "dest.xlsx")
        dest2 = os.path.join(td, "dest2.xlsx")
        r1 = run_sheet(src, _cfg(dest,  src_start_row="1"))
        r2 = run_sheet(src, _cfg(dest2, src_start_row=""))
        assert r1.rows_written == r2.rows_written == 2


def test_run_sheet_source_start_row_non_numeric_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="abc"))


def test_run_sheet_source_start_row_zero_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="0"))


def test_run_sheet_source_start_row_negative_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="-1"))


def test_run_sheet_source_start_row_past_end_zero_rows():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"], ["b"], ["c"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, src_start_row="10"))
        assert result.rows_written == 0


# ══════════════════════════════════════════════════════════════════════════════
# KEEP MODE
# ══════════════════════════════════════════════════════════════════════════════

def test_run_sheet_keep_mode_all_rows_all_cols():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["a", "b"], ["c", "d"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns="", rows="", mode="keep"))
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "a"
        assert ws2["B2"].value == "d"


def test_run_sheet_keep_non_adjacent_cols_preserves_gaps():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["alpha", "x", 1],
                                ["beta",  "y", 2],
                                ["gamma", "z", 3]])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="A", start_row=""),
        )
        result = run_sheet(src, cfg)
        assert result.rows_written == 3
        ws2 = _ws(dest)
        assert ws2["A1"].value == "alpha"
        assert ws2["B1"].value is None   # gap
        assert ws2["C1"].value == 1

def test_run_sheet_keep_mode_rules_filter_rows():
    """
    Rules must filter rows in keep mode. shape_keep returns a bounding box,
    so the filtered-out row becomes a None gap — but its data must not appear.
    """
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["keep",  "x", 1],
                                ["drop",  "y", 2],
                                ["keep",  "z", 3]])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="keep", rules_combine="AND",
            rules=[Rule(mode="include", column="A",
                        operator="equals", value="keep")],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="A", start_row=""),
        )
        result = run_sheet(src, cfg)
        ws2 = load_workbook(dest)["Out"]
        # Collect all non-None values from col A
        col_a = [ws2.cell(row=r, column=1).value for r in range(1, result.rows_written + 1)]
        # "drop" must not appear anywhere — it was filtered by the rule
        assert "drop" not in col_a
        # Both "keep" values must be present
        assert col_a.count("keep") == 2


def test_run_sheet_pack_mode_rules_filter_rows():
    """Sanity check: rules work in pack mode (regression guard)."""
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["keep", 1], ["drop", 2], ["keep", 3]])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, rules=[Rule(mode="include", column="A",
                                      operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "keep"
        assert ws2["A2"].value == "keep"


# ══════════════════════════════════════════════════════════════════════════════
# DESTINATION SHEET MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

def test_new_dest_file_gets_default_sheet_removed(tmp_path):
    src  = _make_xlsx(str(tmp_path / "src.xlsx"), data=[["a"]])
    dest = str(tmp_path / "dest.xlsx")
    run_sheet(src, _cfg(dest, dest_sheet="MyOutput"))
    wb = load_workbook(dest)
    assert "MyOutput" in wb.sheetnames
    assert "Sheet" not in wb.sheetnames or len(wb.sheetnames) == 1


def test_run_sheet_multiple_dest_sheets_preserved(tmp_path):
    wb = Workbook()
    wb.active.title = "Existing"
    wb["Existing"]["A1"] = "keep_me"
    wb.create_sheet("Other")
    wb["Other"]["A1"] = "also_keep"
    dest = str(tmp_path / "dest.xlsx")
    wb.save(dest)

    src = _make_xlsx(str(tmp_path / "src.xlsx"), data=[["new"]])
    run_sheet(src, _cfg(dest, dest_sheet="Existing", start_row="2"))

    wb2 = load_workbook(dest)
    assert wb2["Existing"]["A1"].value == "keep_me"
    assert wb2["Existing"]["A2"].value == "new"
    assert wb2["Other"]["A1"].value == "also_keep"


def test_run_sheet_two_calls_same_file_different_dest_sheets():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["v1"], ["v2"]])
        dest = os.path.join(td, "dest.xlsx")
        r1 = run_sheet(src, _cfg(dest, dest_sheet="Sheet1"))
        r2 = run_sheet(src, _cfg(dest, dest_sheet="Sheet2"))
        assert r1.rows_written == 2
        assert r2.rows_written == 2
        wb = load_workbook(dest)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames


# ══════════════════════════════════════════════════════════════════════════════
# ERROR PATHS
# ══════════════════════════════════════════════════════════════════════════════

def test_missing_sheet_raises_sheet_not_found():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"))
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, src_sheet="DoesNotExist")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == SHEET_NOT_FOUND


def test_bad_column_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, columns="A,??"))
        assert ei.value.code == BAD_SPEC


def test_bad_row_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, rows="nope"))
        assert ei.value.code == BAD_SPEC


def test_bad_start_row_zero_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="0"))
        assert ei.value.code == BAD_SPEC


def test_collision_raises_dest_blocked():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="1"))
        assert ei.value.code == DEST_BLOCKED


def test_collision_blocked_on_inner_row_of_output():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["r1"], ["r2"], ["r3"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"
        wb.save(dest)
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="1"))
        assert ei.value.code == DEST_BLOCKED


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE ORDERING
# ══════════════════════════════════════════════════════════════════════════════

def test_pipeline_rules_use_absolute_source_column_not_in_output():
    """Rules must run against original source columns, not post-selection columns."""
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[
            ["keep",  "x", 1, "YES"],
            ["drop",  "y", 2, "NO"],
            ["keep2", "z", 3, "YES"],
        ])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, columns="A,C",
                    rules=[Rule(mode="include", column="D",
                                operator="equals", value="YES")])
        result = run_sheet(src, cfg, recipe_name="R")
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "keep"
        assert ws2["B1"].value == 1
        assert ws2["A2"].value == "keep2"
        assert ws2["B2"].value == 3


def test_pipeline_rules_then_column_selection_order():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[
            ["alpha", "x", 1, "YES"],
            ["beta",  "y", 2, "NO"],
        ])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, columns="A,B",
                    rules=[Rule(mode="include", column="D",
                                operator="equals", value="YES")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        assert _ws(dest)["A1"].value == "alpha"


# ══════════════════════════════════════════════════════════════════════════════
# APPEND — LANDING ZONE ISOLATION
# ══════════════════════════════════════════════════════════════════════════════

def test_append_column_outside_landing_zone_does_not_affect_row():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["val1", "val2"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        for i in range(1, 101):
            ws[f"A{i}"] = f"noise_{i}"
        wb.save(dest)
        result = run_sheet(src, _cfg(dest, columns="A,B", start_col="B"))
        assert result.rows_written == 1
        ws2 = _ws(dest)
        assert ws2["B1"].value == "val1"
        assert ws2["C1"].value == "val2"


def test_append_formula_cell_treated_as_unoccupied():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["new_data"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "=SUM(B1:B10)"
        wb.save(dest)
        result = run_sheet(src, _cfg(dest, columns="A", start_col="A"))
        assert result.rows_written == 1


# ══════════════════════════════════════════════════════════════════════════════
# PACK / KEEP COMBINATION MATRIX
# ══════════════════════════════════════════════════════════════════════════════

def test_pack_all_cols_no_rules():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["a", 1], ["b", 2], ["c", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        assert _ws(dest)["A1"].value == "a"


def test_pack_subset_columns_no_rules():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["a", "b", "c"], ["d", "e", "f"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,C"))
        assert result.rows_written == 2
        assert _ws(dest)["A1"].value == "a"
        assert _ws(dest)["B1"].value == "c"


def test_pack_include_equals_rule():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["keep", 1], ["drop", 2], ["keep", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="include", column="A", operator="equals", value="keep")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["A2"].value == "keep"


def test_pack_exclude_rule():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["alpha", 1], ["beta", 2], ["gamma", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="exclude", column="A", operator="equals", value="beta")
        ]))
        assert result.rows_written == 2
        names = [_ws(dest)[f"A{i}"].value for i in range(1, 3)]
        assert "beta" not in names


def test_pack_row_range_selection():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["r1"], ["r2"], ["r3"], ["r4"], ["r5"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rows="2-4"))
        assert result.rows_written == 3
        assert _ws(dest)["A1"].value == "r2"


def test_pack_explicit_start_row():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["only"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, start_row="999"))
        assert result.rows_written == 1
        assert _ws(dest)["A999"].value == "only"


def test_pack_exclude_all_rule_zero_rows():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["alpha", 1], ["beta", 2]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="exclude", column="A", operator="contains", value="")
        ]))
        assert result.rows_written == 0


def test_rows_spec_beyond_used_range_ignored_gracefully():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["r1"], ["r2"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rows="1,3,5,100"))
        assert result.rows_written == 1   # only row 1 exists


# ══════════════════════════════════════════════════════════════════════════════
# EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_unicode_values_preserved():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["日本語", "中文", "한국어"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 1
        assert _ws(dest)["A1"].value == "日本語"


def test_very_long_string_cell_value_survives_roundtrip():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["x" * 10_000, "short"], ["normal", "val"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        assert len(_ws(dest)["A1"].value) == 10_000


def test_csv_quoted_fields_with_commas():
    with TemporaryDirectory() as td:
        src  = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")
        _make_csv(src, [["Smith, John", "New York, NY", 100],
                        ["Doe, Jane",   "Austin, TX",   200]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "Smith, John"
        assert out["B1"].value == "New York, NY"


def test_dest_sheet_name_with_spaces():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["v", 1]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, dest_sheet="My Sheet Name"))
        assert result.rows_written == 1
        assert load_workbook(dest)["My Sheet Name"]["A1"].value == "v"