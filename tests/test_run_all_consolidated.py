# Consolidated from: test_run_all.py, test_run_all_more.py, test_project_config.py
# Generated: 2026-02-19 20:40 UTC
# NOTE: Function renames applied only to avoid name collisions across original test modules.



# ---- BEGIN test_run_all.py ----

\
import os
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_all
from core.models import SheetConfig, Destination
from core.errors import DEST_BLOCKED


def runall_make_source_xlsx(path: str, sheet_name: str, tag: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = f"{tag}-alpha"
    ws["B1"] = "x"
    ws["C1"] = 1
    wb.save(path)


def test_run_all_two_items_success_same_dest_stacks():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "out.xlsx")

        runall_make_source_xlsx(src1, "Sheet1", "S1")
        runall_make_source_xlsx(src2, "Sheet1", "S2")

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (src1, "R1", cfg1),
            (src2, "R2", cfg2),
        ])

        assert report.ok is True
        assert len(report.results) == 2
        assert report.results[0].rows_written == 1
        assert report.results[1].rows_written == 1

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "S1-alpha"
        assert ws["C1"].value == 1
        assert ws["B2"].value == "S2-alpha"
        assert ws["C2"].value == 1


def test_run_all_fail_fast_records_error_and_stops():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "out.xlsx")

        runall_make_source_xlsx(src1, "Sheet1", "S1")
        runall_make_source_xlsx(src2, "Sheet1", "S2")

        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B1"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )

        cfg_should_not_run = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (src1, "R1", cfg_blocked),
            (src2, "R2", cfg_should_not_run),
        ])

        assert report.ok is False
        assert len(report.results) == 1
        assert report.results[0].error_code == DEST_BLOCKED

        wb2 = load_workbook(dest)
        ws2 = wb2["Out"]
        values = [ws2["B1"].value, ws2["B2"].value, ws2["B3"].value]
        assert "S2-alpha" not in values


# ---- END {f} ----



# ---- BEGIN test_run_all_more.py ----

\
import os
import csv
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_all
from core.models import SheetConfig, Destination, Rule
from core.errors import DEST_BLOCKED


def runall_more_make_source_xlsx(path: str, sheet_name: str, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)


def runall_more_make_source_csv(path: str, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(row)


def test_run_all_three_items_stacks_rows_in_order_same_dest():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")
        s3 = os.path.join(td, "s3.xlsx")

        runall_more_make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        runall_more_make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])
        runall_more_make_source_xlsx(s3, "Sheet1", [["A3", "x", 3]])

        cfg = lambda: SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg()),
            (s2, "R2", cfg()),
            (s3, "R3", cfg()),
        ])

        assert report.ok is True
        assert len(report.results) == 3

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "A1"
        assert ws["C1"].value == 1
        assert ws["B2"].value == "A2"
        assert ws["C2"].value == 2
        assert ws["B3"].value == "A3"
        assert ws["C3"].value == 3


def test_run_all_mixed_widths_append_uses_full_landing_zone_awareness():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        # s1 writes width=3 (A,B,C) -> landing zone B:D
        runall_more_make_source_xlsx(s1, "Sheet1", [["v1", "v2", "v3"]])
        # s2 writes width=2 (A,C) -> landing zone B:C
        runall_more_make_source_xlsx(s2, "Sheet1", [["w1", "x", "w3"]])

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,B,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg1),
            (s2, "R2", cfg2),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]
        # first row from cfg1 occupies B1:D1
        assert ws["B1"].value == "v1"
        assert ws["C1"].value == "v2"
        assert ws["D1"].value == "v3"
        # second run uses landing zone B:C; max used row across B:C is 1 => append at 2
        assert ws["B2"].value == "w1"
        assert ws["C2"].value == "w3"


def test_run_all_two_different_destinations_both_created():
    with TemporaryDirectory() as td:
        dest1 = os.path.join(td, "out1.xlsx")
        dest2 = os.path.join(td, "out2.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        runall_more_make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        runall_more_make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest1, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest2, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg1),
            (s2, "R2", cfg2),
        ])
        assert report.ok is True

        wb1 = load_workbook(dest1)
        wb2 = load_workbook(dest2)
        assert wb1["Out"]["B1"].value == "A1"
        assert wb2["Out"]["B1"].value == "A2"


def test_run_all_keep_mode_then_pack_mode_stacks():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        # 3 rows so keep-mode bounding box height=3 when rows_spec=1,3
        runall_more_make_source_xlsx(s1, "Sheet1", [
            ["alpha", "x", 1],
            ["beta",  "y", 2],
            ["gamma", "z", 3],
        ])
        runall_more_make_source_xlsx(s2, "Sheet1", [["delta", "q", 9]])

        cfg_keep = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1,3",
            paste_mode="keep",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_pack = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_keep),
            (s2, "R2", cfg_pack),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]

        # keep run occupies B1:D3 (width=3, height=3)
        assert ws["B1"].value == "alpha"
        assert ws["D1"].value == 1
        assert ws["B2"].value is None
        assert ws["B3"].value == "gamma"
        assert ws["D3"].value == 3

        # pack run uses landing zone B:C; max used row across B:C is 3 => append at 4
        assert ws["B4"].value == "delta"
        assert ws["C4"].value == 9


def test_run_all_fail_fast_on_second_item_dest_blocked():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")
        s3 = os.path.join(td, "s3.xlsx")

        runall_more_make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        runall_more_make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])
        runall_more_make_source_xlsx(s3, "Sheet1", [["A3", "x", 3]])

        cfg_ok = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        # Second item uses explicit start row that is blocked
        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B50"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="50"),
        )

        report = run_all([
            (s1, "R1", cfg_ok),
            (s2, "R2", cfg_blocked),
            (s3, "R3", cfg_ok),
        ])

        assert report.ok is False
        assert len(report.results) == 2
        assert report.results[1].error_code == DEST_BLOCKED

        wb2 = load_workbook(dest)
        ws2 = wb2["Out"]
        # Third item should not run
        colB = [ws2[f"B{i}"].value for i in range(1, 10)]
        assert "A3" not in colB


def test_run_all_csv_then_xlsx_same_dest_stacks():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.csv")
        s2 = os.path.join(td, "s2.xlsx")

        runall_more_make_source_csv(s1, [["csv1", "x", 1]])
        runall_more_make_source_xlsx(s2, "Sheet1", [["xls1", "y", 2]])

        cfg_csv = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_xls = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_csv),
            (s2, "R2", cfg_xls),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "csv1"
        assert ws["C1"].value in (1, "1")
        assert ws["B2"].value == "xls1"
        assert ws["C2"].value == 2


def test_run_all_rules_filter_zero_rows_then_next_item_runs_and_writes():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        runall_more_make_source_xlsx(s1, "Sheet1", [["alpha", "x", 1], ["beta", "y", 2]])
        runall_more_make_source_xlsx(s2, "Sheet1", [["gamma", "z", 3]])

        cfg_zero = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="",
            paste_mode="pack",
            rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_write = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_zero),
            (s2, "R2", cfg_write),
        ])
        assert report.ok is True
        assert report.results[0].rows_written == 0
        assert report.results[1].rows_written == 1

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "gamma"
        assert ws["C1"].value == 3


# ---- END {f} ----



# ---- BEGIN test_project_config.py ----

\
import os
from tempfile import TemporaryDirectory

from core.project import ProjectConfig, SourceConfig, RecipeConfig
from core.models import SheetConfig, Destination, Rule


def make_sample_project(dest_path: str):
    sheet = SheetConfig(
        name="Sheet1",
        workbook_sheet="Sheet1",
        columns_spec="A,C",
        rows_spec="1-1",
        paste_mode="pack",
        rules_combine="AND",
        rules=[Rule(mode="include", column="A", operator="equals", value="alpha")],
        destination=Destination(
            file_path=dest_path,
            sheet_name="Out",
            start_col="B",
            start_row="",
        ),
    )
    recipe = RecipeConfig(name="R1", sheets=[sheet])
    source = SourceConfig(path="dummy.xlsx", recipes=[recipe])
    return ProjectConfig(sources=[source])


def test_project_serialization_roundtrip():
    with TemporaryDirectory() as td:
        json_path = os.path.join(td, "proj.json")
        proj = make_sample_project("out.xlsx")

        proj.save_json(json_path)
        loaded = ProjectConfig.load_json(json_path)

        assert len(loaded.sources) == 1
        assert loaded.sources[0].recipes[0].name == "R1"
        assert loaded.sources[0].recipes[0].sheets[0].columns_spec == "A,C"


def test_project_build_run_items_order():
    proj = make_sample_project("out.xlsx")
    items = proj.build_run_items()

    assert len(items) == 1
    src_path, recipe_name, sheet = items[0]
    assert src_path == "dummy.xlsx"
    assert recipe_name == "R1"
    assert sheet.name == "Sheet1"


# ---- END {f} ----
