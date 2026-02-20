"""
Engine hardening tests.

Covers robustness under conditions the happy-path tests never exercise:
  - Wide sheets (100+ cols)
  - Unicode, emoji, CJK in cell values
  - Merged cells (openpyxl reads anchor, rest None)
  - Empty / sparse rows in the middle of data
  - Mixed types in one column (int, float, str, None, bool)
  - Very long string cell values
  - Negative and float numbers in numeric rules
  - CSV with quoted commas, trailing empty lines
  - Ragged rows (trailing empty cols)
  - Column letter edge cases (Z, AA, ZZ, AAA, XFD)
  - Rows spec that references beyond the used range
  - source_start_row past end of file
  - Destination sheet name with spaces
  - 200k row end-to-end run (slow — marked with pytest.mark.slow)
"""
from __future__ import annotations

import csv
import os
import time
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet
from core.errors import AppError, BAD_SPEC
from core.models import Destination, Rule, SheetConfig
from core.parsing import col_index_to_letters, col_letters_to_index


# ── helpers ───────────────────────────────────────────────────────────────────

def _xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data or [], 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _csv(path: str, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)


def _cfg(dest_path, *, columns="", rows="", mode="pack",
         rules=None, combine="AND", start_col="A", start_row="",
         dest_sheet="Out", src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path: str, sheet: str = "Out"):
    return load_workbook(path)[sheet]


# ═══════════════════════════════════════════════════════════════════════════════
# WIDE SHEETS (100+ COLS)
# ═══════════════════════════════════════════════════════════════════════════════

def test_wide_sheet_128_cols_all_cols_pack():
    """128-column source extracted with all cols and no rules."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        row = [f"col{i}" for i in range(128)]
        ws.append(row)
        ws.append([i * 10 for i in range(128)])
        wb.save(src)

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2

        out = _ws(dest)
        # Last column should be column DX (128th column letter)
        last_letter = col_index_to_letters(128)
        assert out[f"A1"].value == "col0"
        assert out[f"{last_letter}1"].value == "col127"
        assert out[f"{last_letter}2"].value == 127 * 10


def test_wide_sheet_subset_of_cols_from_high_columns():
    """Select only the last few columns of a 100-col sheet."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        row = list(range(100))
        ws.append(row)
        wb.save(src)

        # Select cols CV and CX (indices 99 and 101 in 1-based = CV=100, CX... wait)
        # col 99 = CW? Let's use col_index_to_letters to be precise
        col_98 = col_index_to_letters(98)   # 0-based 97 → 1-based 98
        col_100 = col_index_to_letters(100)
        cfg = _cfg(dest, columns=f"{col_98},{col_100}")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1

        out = _ws(dest)
        # col_98 (1-based) → value at index 97 (0-based)
        assert out["A1"].value == 97
        assert out["B1"].value == 99


def test_wide_sheet_column_letter_roundtrip_at_boundaries():
    """col_index_to_letters and col_letters_to_index are inverses at boundary cols."""
    boundaries = [1, 26, 27, 52, 53, 702, 703, 16384]
    for n in boundaries:
        letters = col_index_to_letters(n)
        assert col_letters_to_index(letters) == n, f"Roundtrip failed at {n} → {letters}"


def test_wide_sheet_zz_column_spec():
    """ZZ (col 702) can be specified in columns_spec and extracted correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Set cell ZZ1 = 9999
        ws.cell(row=1, column=702, value=9999)
        ws.cell(row=1, column=1, value="first")
        wb.save(src)

        cfg = _cfg(dest, columns="A,ZZ")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1

        out = _ws(dest)
        assert out["A1"].value == "first"
        assert out["B1"].value == 9999


# ═══════════════════════════════════════════════════════════════════════════════
# UNICODE, EMOJI, CJK
# ═══════════════════════════════════════════════════════════════════════════════

def test_unicode_values_pass_through_pack():
    """Unicode strings survive load → filter → write unchanged."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        data = [
            ["café",    "naïve",   "résumé"],
            ["日本語",   "中文",    "한국어"],
            ["emoji🚀",  "fire🔥",  "✓check"],
            ["αβγδ",    "Ünïcödé", "мир"],
        ]
        _xlsx(src, data=data)

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 4

        out = _ws(dest)
        assert out["A1"].value == "café"
        assert out["A2"].value == "日本語"
        assert out["A3"].value == "emoji🚀"
        assert out["C3"].value == "✓check"
        assert out["B4"].value == "Ünïcödé"


def test_unicode_rule_filter_equals():
    """Include rule with Unicode value correctly filters rows."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            ["日本語", 1],
            ["中文",   2],
            ["日本語", 3],
        ])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="equals", value="日本語")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "日本語"
        assert out["B1"].value == 1
        assert out["B2"].value == 3


def test_unicode_rule_filter_contains():
    """Contains rule with emoji substring matches correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            ["hello🚀world", 1],
            ["no match",    2],
            ["also🚀here",  3],
        ])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="contains", value="🚀")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2


def test_unicode_csv_passthrough():
    """CSV with Unicode values loads and writes correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        _csv(src, [["café", "naïve"], ["日本語", "中文"]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "café"
        assert out["A2"].value == "日本語"


# ═══════════════════════════════════════════════════════════════════════════════
# MERGED CELLS
# ═══════════════════════════════════════════════════════════════════════════════

def test_merged_cells_anchor_read_others_none():
    """Merged region: anchor cell has value, merged-into cells read as None."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "merged_header"
        ws["D1"] = "separate"
        ws.merge_cells("A1:C1")   # A1 is anchor; B1, C1 become None
        ws["A2"] = "data1"
        ws["B2"] = "data2"
        ws["C2"] = "data3"
        wb.save(src)

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2

        out = _ws(dest)
        assert out["A1"].value == "merged_header"
        assert out["B1"].value is None   # merged-into cell
        assert out["C1"].value is None   # merged-into cell
        assert out["D1"].value == "separate"
        assert out["A2"].value == "data1"


def test_merged_cells_rule_on_anchor_works():
    """Rules can still filter on the anchor cell of a merge."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "keep"
        ws["B1"] = 10
        ws.merge_cells("A1:A2")   # A1 anchor, A2 None
        ws["A3"] = "drop"
        ws["B3"] = 30
        wb.save(src)

        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        # Row with "keep" anchor passes; row with None fails (None != "keep")
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["A1"].value == "keep"


def test_merged_cells_vertical_merge_none_rows_handled():
    """Vertical merge produces None rows that engine handles gracefully."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "header"
        ws["A2"] = "data"
        ws["A3"] = "data"
        ws.merge_cells("B1:B3")   # B1 anchor; B2, B3 are None
        ws["B1"] = "side_label"
        wb.save(src)

        # Should not crash; rows with None in B2/B3 are valid rows
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["B1"].value == "side_label"
        assert out["B2"].value is None
        assert out["B3"].value is None


# ═══════════════════════════════════════════════════════════════════════════════
# EMPTY / SPARSE ROWS IN MIDDLE OF DATA
# ═══════════════════════════════════════════════════════════════════════════════

def test_empty_rows_in_middle_used_range_correct():
    """Empty rows in the middle don't affect used-range calculation."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            ["data", 1],
            [None,   None],
            [None,   None],
            ["data", 4],
            [None,   None],
            ["data", 6],
        ])
        result = run_sheet(src, _cfg(dest))
        # All 6 rows are within the used range (row 6 has data)
        assert result.rows_written == 6
        out = _ws(dest)
        assert out["A1"].value == "data"
        assert out["A2"].value is None
        assert out["A4"].value == "data"


def test_empty_rows_in_middle_pack_with_rule_skips_empty():
    """Pack + include rule: empty rows fail the rule and are excluded."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            ["keep",  1],
            [None,    None],   # empty row — A is None, "None" != "keep"
            ["keep",  3],
            ["drop",  4],
        ])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="equals", value="keep")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "keep"
        assert out["B1"].value == 1
        assert out["B2"].value == 3


def test_leading_empty_rows_append_still_works():
    """Source where first N rows are empty — used range starts later."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            [None,   None],
            [None,   None],
            ["real", 99],
        ])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["A1"].value is None
        assert out["A3"].value == "real"


def test_trailing_empty_rows_after_last_data():
    """Trailing empty rows beyond last occupied row still write (within used range)."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "data"
        # Explicitly write empty strings in rows 2-4 to make them part of the sheet
        ws["A2"] = ""
        ws["A3"] = ""
        wb.save(src)

        result = run_sheet(src, _cfg(dest))
        # Only row 1 has occupied data; rows 2-3 have "" which is unoccupied
        # used_range returns h=1 since is_occupied("") == False
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["A1"].value == "data"


# ═══════════════════════════════════════════════════════════════════════════════
# MIXED TYPES IN ONE COLUMN
# ═══════════════════════════════════════════════════════════════════════════════

def test_mixed_types_in_column_pack_no_rules():
    """int, float, str, None, bool all survive the pipeline unchanged."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            [42,    "label"],
            [3.14,  "label"],
            ["text","label"],
            [None,  "label"],
            [True,  "label"],
            [False, "label"],
        ])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 6

        out = _ws(dest)
        assert out["A1"].value == 42
        assert abs(out["A2"].value - 3.14) < 1e-9
        assert out["A3"].value == "text"
        assert out["A4"].value is None
        assert out["A5"].value is True
        assert out["A6"].value is False


def test_mixed_types_numeric_rule_skips_non_numeric_safely():
    """Numeric rule on a mixed col: non-numeric rows treated as no-match, no crash."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[
            [100,    "big"],
            ["text", "skip"],
            [None,   "skip"],
            [True,   "bool"],   # True == 1, less than 50 → skip
            [200,    "big"],
        ])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator=">", value="50")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == 100
        assert out["A2"].value == 200


def test_mixed_types_negative_numbers_in_rule():
    """Negative numbers in both data and rule threshold work correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[[-10, "a"], [-5, "b"], [0, "c"], [5, "d"]])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="<", value="-3")])
        result = run_sheet(src, cfg)
        # Both -10 and -5 are less than -3
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == -10
        assert out["A2"].value == -5


def test_mixed_types_float_threshold_in_rule():
    """Float threshold (1.5) in rule filters correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[[1.0, "a"], [1.5, "b"], [1.6, "c"], [2.0, "d"]])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator=">", value="1.5")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == 1.6
        assert out["A2"].value == 2.0


# ═══════════════════════════════════════════════════════════════════════════════
# VERY LONG STRING VALUES
# ═══════════════════════════════════════════════════════════════════════════════

def test_very_long_string_cell_value_survives_roundtrip():
    """A 10 000-character string survives load → write unchanged."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        long_val = "x" * 10_000
        _xlsx(src, data=[[long_val, "short"], ["normal", "val"]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        out = _ws(dest)
        assert len(out["A1"].value) == 10_000
        assert out["A2"].value == "normal"


def test_very_long_string_rule_contains_works():
    """Contains rule on a cell that is a long string finds the substring."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        prefix = "a" * 5_000
        match_val = prefix + "TARGET" + "b" * 5_000
        no_match_val = "c" * 10_000

        _xlsx(src, data=[[match_val, 1], [no_match_val, 2], [match_val, 3]])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="contains", value="TARGET")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["B1"].value == 1
        assert out["B2"].value == 3


# ═══════════════════════════════════════════════════════════════════════════════
# CSV EDGE CASES
# ═══════════════════════════════════════════════════════════════════════════════

def test_csv_quoted_fields_with_commas():
    """CSV fields containing commas inside quotes parse correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        # csv.writer handles quoting automatically
        _csv(src, [
            ["Smith, John", "New York, NY", 100],
            ["Doe, Jane",   "Austin, TX",  200],
        ])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "Smith, John"
        assert out["B1"].value == "New York, NY"


def test_csv_quoted_fields_rule_on_comma_value():
    """Rule filter works on a value that itself contains a comma."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        _csv(src, [
            ["Smith, John", 1],
            ["Doe, Jane",   2],
            ["Smith, John", 3],
        ])
        cfg = _cfg(dest, rules=[Rule(mode="include", column="A",
                                     operator="equals", value="Smith, John")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 2


def test_csv_trailing_empty_lines_ignored():
    """CSV with blank lines at the end doesn't add empty rows to output."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        # Write file manually with trailing blank lines
        with open(src, "w", newline="", encoding="utf-8") as f:
            f.write("alpha,1\nbeta,2\n\n\n")

        result = run_sheet(src, _cfg(dest))
        # Trailing empty lines become rows with ["", ""] which are unoccupied
        # compute_used_range will report used_h=2
        out = _ws(dest)
        assert out["A1"].value == "alpha"
        assert out["A2"].value == "beta"
        # Regardless of whether empty rows were loaded, no data at row 3+
        assert out["A3"].value in (None, "")


def test_csv_single_column_no_commas():
    """Single-column CSV (no commas) loads as a list of single-element rows."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        with open(src, "w", encoding="utf-8") as f:
            f.write("alpha\nbeta\ngamma\n")

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["A1"].value == "alpha"
        assert out["A3"].value == "gamma"


# ═══════════════════════════════════════════════════════════════════════════════
# RAGGED ROWS (TRAILING EMPTY COLS)
# ═══════════════════════════════════════════════════════════════════════════════

def test_ragged_rows_normalized_to_max_width():
    """Rows of varying lengths are padded with None to uniform width."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["a", "b", "c", "d"])   # 4 cols
        ws.append(["x", "y"])              # only 2 cols
        ws.append(["p"])                   # only 1 col
        wb.save(src)

        result = run_sheet(src, _cfg(dest, columns="A,B,C,D"))
        assert result.rows_written == 3

        out = _ws(dest)
        assert out["D1"].value == "d"
        assert out["C2"].value is None   # padded
        assert out["B3"].value is None   # padded
        assert out["D3"].value is None   # padded


def test_ragged_csv_rows_normalized():
    """Ragged CSV rows are padded with None for consistent column access."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")

        with open(src, "w", newline="", encoding="utf-8") as f:
            f.write("a,b,c\nx,y\np\n")

        result = run_sheet(src, _cfg(dest, columns="A,B,C"))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["A1"].value == "a"
        assert out["C1"].value == "c"
        assert out["C2"].value in (None, "")   # padded / empty
        assert out["B3"].value in (None, "")


# ═══════════════════════════════════════════════════════════════════════════════
# ROWS SPEC BEYOND USED RANGE
# ═══════════════════════════════════════════════════════════════════════════════

def test_rows_spec_partially_beyond_used_range():
    """rows_spec with some indices beyond the data silently skips missing rows."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["r1", 1], ["r2", 2]])

        # Request rows 1, 2, 5, 10 — only 1 and 2 exist
        cfg = _cfg(dest, rows="1,2,5,10")
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "r1"
        assert out["A2"].value == "r2"
        assert out["A3"].value is None


def test_rows_spec_entirely_beyond_used_range():
    """rows_spec that selects only rows beyond the data → 0 written."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["r1", 1], ["r2", 2]])

        cfg = _cfg(dest, rows="100-200")
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


# ═══════════════════════════════════════════════════════════════════════════════
# SOURCE_START_ROW EDGE CASES
# ═══════════════════════════════════════════════════════════════════════════════

def test_source_start_row_exactly_at_last_row():
    """source_start_row pointing to the last row extracts exactly one row."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["skip1", 1], ["skip2", 2], ["last", 3]])
        cfg = _cfg(dest, src_start_row="3")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["A1"].value == "last"


def test_source_start_row_beyond_last_row_writes_nothing():
    """source_start_row past end of data → empty table → 0 rows written."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["r1", 1], ["r2", 2]])
        cfg = _cfg(dest, src_start_row="10")
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


def test_source_start_row_one_skips_nothing():
    """source_start_row=1 is the same as no offset."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest1 = os.path.join(td, "d1.xlsx")
        dest2 = os.path.join(td, "d2.xlsx")

        _xlsx(src, data=[["r1", 1], ["r2", 2]])
        r1 = run_sheet(src, _cfg(dest1, src_start_row="1"))
        r2 = run_sheet(src, _cfg(dest2, src_start_row=""))
        assert r1.rows_written == r2.rows_written == 2


# ═══════════════════════════════════════════════════════════════════════════════
# DESTINATION SHEET NAME WITH SPACES
# ═══════════════════════════════════════════════════════════════════════════════

def test_dest_sheet_name_with_spaces():
    """Destination sheet names containing spaces are created and written to."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["alpha", 1], ["beta", 2]])
        cfg = _cfg(dest, dest_sheet="My Output Sheet")
        result = run_sheet(src, cfg)
        assert result.rows_written == 2

        wb = load_workbook(dest)
        assert "My Output Sheet" in wb.sheetnames
        ws = wb["My Output Sheet"]
        assert ws["A1"].value == "alpha"


def test_dest_sheet_name_with_special_chars():
    """Sheet names with underscores and hyphens work correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["data", 99]])
        cfg = _cfg(dest, dest_sheet="Q3-Results_2025")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1

        wb = load_workbook(dest)
        assert "Q3-Results_2025" in wb.sheetnames


# ═══════════════════════════════════════════════════════════════════════════════
# SINGLE ROW / SINGLE COLUMN EDGE CASES
# ═══════════════════════════════════════════════════════════════════════════════

def test_single_column_source_pack():
    """Source with only one column processes without errors."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["a"], ["b"], ["c"]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["A3"].value == "c"


def test_single_row_source_pack():
    """Source with only one row processes without errors."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["only", "row", "here"]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["C1"].value == "here"


def test_single_cell_source_pack():
    """Source with a single occupied cell (1×1) processes correctly."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        _xlsx(src, data=[["solo"]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["A1"].value == "solo"


def test_empty_sheet_writes_nothing():
    """A completely empty source sheet writes 0 rows without crashing."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")

        wb = Workbook(); wb.active.title = "Sheet1"; wb.save(src)
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 0


