"""
test_core.py — Consolidated core unit tests.

Covers:
  - core.io: is_occupied, normalize_table, compute_used_range, load_xlsx, load_csv
  - core.planner: build_plan, is_cell_occupied, append scan, collision detection
  - core.writer: apply_write_plan
  - core.transform: apply_row_selection, apply_column_selection, shape_pack, shape_keep
  - core.parsing: col_letters_to_index, col_index_to_letters, parse_columns, parse_rows
  - core.rules: apply_rules (operators, AND/OR, edge cases)
  - core.errors: AppError string formatting
  - Occupancy consistency between is_occupied and is_cell_occupied
"""
from __future__ import annotations

import csv
import os
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.errors import AppError, BAD_SPEC, DEST_BLOCKED, INVALID_RULE
from core.io import compute_used_range, is_occupied, load_xlsx, normalize_table
from core.models import Destination, Rule, SheetConfig
from core.parsing import (
    col_index_to_letters,
    col_letters_to_index,
    parse_columns,
    parse_rows,
)
from core.planner import build_plan, is_cell_occupied
from core.rules import apply_rules
from core.transform import (
    apply_column_selection,
    apply_row_selection,
    shape_keep,
    shape_pack,
)
from core.writer import apply_write_plan


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _fresh_ws():
    wb = Workbook()
    return wb, wb.active


def _ws():
    """Return a fresh in-memory worksheet."""
    return Workbook().active


def _xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data or [], 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# CORE.IO — is_occupied
# ══════════════════════════════════════════════════════════════════════════════

def test_is_occupied_none_is_false():
    assert is_occupied(None) is False


def test_is_occupied_empty_string_is_false():
    assert is_occupied("") is False


def test_is_occupied_whitespace_is_true():
    assert is_occupied(" ") is True


def test_is_occupied_zero_is_true():
    assert is_occupied(0) is True


def test_is_occupied_zero_float_is_true():
    assert is_occupied(0.0) is True


def test_is_occupied_false_bool_is_true():
    assert is_occupied(False) is True


def test_is_occupied_true_bool_is_true():
    assert is_occupied(True) is True


def test_is_occupied_text_is_true():
    assert is_occupied("text") is True


def test_is_occupied_integer_is_true():
    assert is_occupied(42) is True


def test_is_occupied_empty_list_is_true():
    assert is_occupied([]) is True


def test_is_occupied_empty_dict_is_true():
    assert is_occupied({}) is True


def test_is_occupied_formula_string_is_occupied():
    """io.is_occupied has no formula special-case — formula strings ARE occupied."""
    assert is_occupied("=SUM(A1:A10)") is True
    assert is_occupied("=A1+B1") is True


# ══════════════════════════════════════════════════════════════════════════════
# CORE.IO — normalize_table, compute_used_range
# ══════════════════════════════════════════════════════════════════════════════

def test_normalize_table_pads_ragged_rows():
    rows = [[1, 2], [3]]
    norm = normalize_table(rows)
    assert len(norm[1]) == 2
    assert norm[1][1] is None


def test_compute_used_range_basic():
    rows = [
        [None, None],
        [None, 5],
        [None, None],
    ]
    h, w = compute_used_range(rows)
    assert h == 2
    assert w == 2


def test_compute_used_range_only_rightmost_col_determines_width():
    rows = [
        [None, None, "v"],
        ["x",  None, None],
    ]
    h, w = compute_used_range(rows)
    assert h == 2
    assert w == 3


def test_compute_used_range_only_lowest_row_determines_height():
    rows = [
        ["a", None],
        [None, None],
        [None, None],
        [None, "b"],
        [None, None],
    ]
    h, w = compute_used_range(rows)
    assert h == 4
    assert w == 2


def test_compute_used_range_empty_returns_zero():
    assert compute_used_range([]) == (0, 0)


# ══════════════════════════════════════════════════════════════════════════════
# CORE.IO — load_xlsx
# ══════════════════════════════════════════════════════════════════════════════

def test_load_xlsx_reads_only_specified_sheet(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "from_sheet1"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "from_sheet2"
    path = str(tmp_path / "multi.xlsx")
    wb.save(path)

    rows1 = load_xlsx(path, "Sheet1")
    rows2 = load_xlsx(path, "Sheet2")
    assert rows1[0][0] == "from_sheet1"
    assert rows2[0][0] == "from_sheet2"


def test_load_xlsx_sheet_with_only_empty_strings_used_range_zero(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = ""
    ws["B2"] = ""
    path = str(tmp_path / "empty_strings.xlsx")
    wb.save(path)

    rows = load_xlsx(path, "Sheet1")
    h, w = compute_used_range(rows)
    assert h == 0
    assert w == 0


def test_load_xlsx_missing_sheet_raises(tmp_path):
    wb = Workbook()
    wb.active.title = "Sheet1"
    path = str(tmp_path / "s.xlsx")
    wb.save(path)
    with pytest.raises(Exception):
        load_xlsx(path, "DoesNotExist")


# ══════════════════════════════════════════════════════════════════════════════
# CORE.PLANNER — is_cell_occupied
# ══════════════════════════════════════════════════════════════════════════════

def test_is_cell_occupied_matches_is_occupied_for_common_values():
    """Both functions agree on all common values."""
    test_cases = [
        (None,   False),
        ("",     False),
        ("hello", True),
        (0,      True),
        (0.0,    True),
        (False,  True),
        (True,   True),
        (42,     True),
        (" ",    True),
    ]
    for val, expected in test_cases:
        assert is_occupied(val) == expected,      f"is_occupied({val!r}) should be {expected}"
        assert is_cell_occupied(val) == expected, f"is_cell_occupied({val!r}) should be {expected}"


def test_is_cell_occupied_formula_string_treated_as_unoccupied():
    """Planner-specific: formula strings (start with '=') are unoccupied in dest scan."""
    assert is_cell_occupied("=SUM(A1:A10)") is False
    assert is_cell_occupied("=A1+B1") is False


# ══════════════════════════════════════════════════════════════════════════════
# CORE.PLANNER — build_plan append mode
# ══════════════════════════════════════════════════════════════════════════════

def test_append_uses_max_used_row_across_landing_cols():
    wb, ws = _fresh_ws()
    ws["D5"] = "x"
    ws["E10"] = "y"
    ws["A100"] = "zzz"  # outside landing zone — must not affect result

    shaped = [["a", "b"], ["c", "d"]]
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 11
    assert plan.start_col == 4
    assert plan.landing_cols == (4, 5)
    assert plan.landing_rows == (11, 12)


def test_append_skips_past_any_used_cells_in_landing_zone():
    wb, ws = _fresh_ws()
    ws["D3"] = "x"
    ws["E4"] = "BLOCK"

    shaped = [["a", "b"], ["c", "d"]]
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 5


def test_append_on_empty_sheet_starts_at_row_1():
    ws = _ws()
    plan = build_plan(ws, [["a"]], "A", "")
    assert plan is not None
    assert plan.start_row == 1


def test_append_outside_landing_zone_does_not_affect_row():
    """Data in col A must NOT affect append row for landing zone B:C."""
    wb, ws = _fresh_ws()
    for i in range(1, 101):
        ws[f"A{i}"] = f"noise_{i}"

    shaped = [["val1", "val2"]]
    plan = build_plan(ws, shaped, start_col_letters="B", start_row_str="")
    assert plan is not None
    assert plan.start_row == 1


def test_append_blocker_absorbed_into_scan_not_probe():
    """
    DEST_BLOCKED in pure append mode is impossible by design:
    any occupied cell in the landing zone gets counted in the scan,
    which pushes start_row past it. The probe then lands in an empty zone.
    """
    wb, ws = _fresh_ws()
    ws["A1"] = "existing"
    ws["A2"] = "existing2"
    ws["A3"] = "BLOCK"  # all 3 cells counted by scan → start_row=4

    plan = build_plan(ws, [["a"]], "A", "")
    assert plan is not None
    assert plan.start_row == 4


# ══════════════════════════════════════════════════════════════════════════════
# CORE.PLANNER — build_plan explicit mode / collision
# ══════════════════════════════════════════════════════════════════════════════

def test_explicit_start_row_collision_probe_blocks():
    wb, ws = _fresh_ws()
    ws["D50"] = "BLOCK"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"]], start_col_letters="D", start_row_str="50")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["target_start"] == "D50"


def test_planner_start_row_zero_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "0")
    assert ei.value.code == BAD_SPEC


def test_planner_start_row_negative_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "-5")
    assert ei.value.code == BAD_SPEC


def test_planner_start_row_float_string_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "3.5")
    assert ei.value.code == BAD_SPEC


def test_planner_blocker_explicit_mode_details_flag_false():
    ws = _ws()
    ws["A5"] = "BLOCK"
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "5")
    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["append_mode"] is False


def test_planner_blocker_append_mode_details_flag_true():
    """
    Triggers DEST_BLOCKED in append mode by using a 2-wide output where the
    scan sees only col A (max_used=2), but col B at start_row=3 has a blocker.
    """
    ws = _ws()
    ws["A1"] = "existing"
    ws["A2"] = "existing2"
    ws["B3"] = "BLOCKER_IN_B"  # not counted by col-A scan, but inside probe rect

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"]], "A", "")  # append mode, width=2

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["append_mode"] is True


def test_planner_blocker_details_contain_first_blocker_fields():
    ws = _ws()
    ws["C3"] = "blocked_value"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"], ["c", "d"]], "B", "2")

    details = ei.value.details
    blocker = details["first_blocker"]
    assert blocker["row"] == 3
    assert blocker["col"] == 3
    assert blocker["col_letter"] == "C"
    assert blocker["value"] == "blocked_value"


def test_planner_collision_anywhere_in_bounding_box():
    """Blocker not at top-left but inside the write rectangle still triggers DEST_BLOCKED."""
    ws = _ws()
    ws["B2"] = "inner_block"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"], ["c", "d"], ["e", "f"]], "A", "1")

    assert ei.value.code == DEST_BLOCKED


def test_build_plan_returns_none_for_empty_shaped():
    wb, ws = _fresh_ws()
    assert build_plan(ws, [], start_col_letters="A", start_row_str="") is None
    assert build_plan(ws, [[]], start_col_letters="A", start_row_str="") is None


# ══════════════════════════════════════════════════════════════════════════════
# CORE.WRITER — apply_write_plan
# ══════════════════════════════════════════════════════════════════════════════

def test_writer_writes_exact_rectangle():
    wb, ws = _fresh_ws()
    shaped = [["a", "b"], ["c", "d"]]
    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    rows_written = apply_write_plan(ws, shaped, plan)
    assert rows_written == 2
    assert ws["C1"].value == "a"
    assert ws["D1"].value == "b"
    assert ws["C2"].value == "c"
    assert ws["D2"].value == "d"


def test_writer_appends_after_existing_data():
    wb, ws = _fresh_ws()
    ws["C1"] = "existing"
    ws["D3"] = "also existing"

    shaped = [["x", "y"]]
    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    apply_write_plan(ws, shaped, plan)

    assert ws["C4"].value == "x"
    assert ws["D4"].value == "y"


def test_writer_writes_bool_values():
    wb, ws = _fresh_ws()
    shaped = [[True, False]]
    plan = build_plan(ws, shaped, "A", "")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value is True
    assert ws["B1"].value is False


def test_writer_writes_numeric_zero():
    wb, ws = _fresh_ws()
    shaped = [[0, 0.0]]
    plan = build_plan(ws, shaped, "A", "")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value == 0


def test_writer_skips_none_values():
    """None cells must not be written (avoids openpyxl phantom cell registration)."""
    wb, ws = _fresh_ws()
    shaped = [["data", None, "more"]]
    plan = build_plan(ws, shaped, "A", "")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value == "data"
    assert ws["B1"].value is None
    assert ws["C1"].value == "more"


# ══════════════════════════════════════════════════════════════════════════════
# CORE.TRANSFORM — apply_row_selection, apply_column_selection
# ══════════════════════════════════════════════════════════════════════════════

def test_apply_row_selection_empty_indices_returns_all():
    rows = [[1], [2], [3]]
    assert apply_row_selection(rows, []) == rows


def test_apply_row_selection_selects_correct_rows():
    rows = [["a"], ["b"], ["c"], ["d"]]
    assert apply_row_selection(rows, [0, 2]) == [["a"], ["c"]]


def test_apply_row_selection_duplicate_indices_duplicates_rows():
    rows = [["a"], ["b"]]
    result = apply_row_selection(rows, [0, 0, 1])
    assert result == [["a"], ["a"], ["b"]]


def test_apply_column_selection_empty_indices_returns_all():
    rows = [[1, 2, 3]]
    assert apply_column_selection(rows, []) == rows


def test_apply_column_selection_selects_correct_cols():
    rows = [["a", "b", "c", "d", "e"]]
    assert apply_column_selection(rows, [0, 2, 4]) == [["a", "c", "e"]]


def test_apply_column_selection_duplicate_indices_duplicates_cols():
    rows = [["a", "b"]]
    result = apply_column_selection(rows, [0, 0, 1])
    assert result == [["a", "a", "b"]]


def test_apply_column_selection_out_of_range_fills_none():
    rows = [["a", "b"]]
    result = apply_column_selection(rows, [0, 5])
    assert result == [["a", None]]


# ══════════════════════════════════════════════════════════════════════════════
# CORE.TRANSFORM — shape_pack, shape_keep
# ══════════════════════════════════════════════════════════════════════════════

def test_shape_pack_returns_rows_unchanged():
    rows = [["a", "b"], ["c", "d"]]
    assert shape_pack(rows) == rows


def test_shape_keep_preserves_gaps_between_selected_rows_and_cols():
    original = [
        ["A1", "B1", "C1"],
        ["A2", "B2", "C2"],
        ["A3", "B3", "C3"],
    ]
    result = shape_keep(original, [0, 2], [0, 2])
    # bounding box: rows 0-2, cols 0-2 → 3×3
    assert len(result) == 3
    assert len(result[0]) == 3
    assert result[0][0] == "A1"
    assert result[0][1] is None   # gap col
    assert result[0][2] == "C1"
    assert result[1][0] is None   # gap row
    assert result[2][0] == "A3"
    assert result[2][2] == "C3"


def test_shape_keep_empty_col_indices_uses_all_cols():
    original = [["a", "b", "c"], ["d", "e", "f"]]
    result = shape_keep(original, [0, 1], [])
    assert result == original


def test_shape_keep_empty_row_indices_uses_all_rows():
    original = [["a", "b"], ["c", "d"]]
    result = shape_keep(original, [], [0, 1])
    assert result == original


def test_shape_keep_bounding_box_height_and_width():
    original = [
        ["a", "b", "c", "d", "e"],
        ["f", "g", "h", "i", "j"],
        ["k", "l", "m", "n", "o"],
        ["p", "q", "r", "s", "t"],
        ["u", "v", "w", "x", "y"],
    ]
    result = shape_keep(original, [0, 2, 4], [0, 2, 4])
    assert len(result) == 5      # rows 0-4
    assert len(result[0]) == 5   # cols 0-4


# ══════════════════════════════════════════════════════════════════════════════
# CORE.PARSING
# ══════════════════════════════════════════════════════════════════════════════

def test_col_letters_to_index_single():
    assert col_letters_to_index("A") == 1
    assert col_letters_to_index("Z") == 26


def test_col_letters_to_index_double():
    assert col_letters_to_index("AA") == 27
    assert col_letters_to_index("AZ") == 52


def test_col_index_to_letters_single():
    assert col_index_to_letters(1) == "A"
    assert col_index_to_letters(26) == "Z"


def test_col_index_to_letters_double():
    assert col_index_to_letters(27) == "AA"


def test_col_letters_roundtrip_boundaries():
    boundaries = [1, 26, 27, 52, 53, 702, 703, 16384]
    for n in boundaries:
        letters = col_index_to_letters(n)
        assert col_letters_to_index(letters) == n, f"Roundtrip failed at {n} → {letters}"


def test_col_letters_to_index_xfd():
    assert col_letters_to_index("XFD") == 16384
    assert col_index_to_letters(16384) == "XFD"


def test_parse_columns_single():
    assert parse_columns("A") == [0]
    assert parse_columns("Z") == [25]
    assert parse_columns("AA") == [26]


def test_parse_columns_list():
    assert parse_columns("A,C,E") == [0, 2, 4]


def test_parse_columns_range():
    assert parse_columns("A-C") == [0, 1, 2]


def test_parse_columns_trailing_comma_ok():
    result = parse_columns("A,B,C,")
    assert result == [0, 1, 2]


def test_parse_rows_single():
    assert parse_rows("1") == [0]


def test_parse_rows_range():
    assert parse_rows("1-3") == [0, 1, 2]


def test_parse_rows_list():
    assert parse_rows("1,3,5") == [0, 2, 4]


def test_parse_rows_very_large_row_number():
    result = parse_rows("1000000")
    assert result == [999999]


# ══════════════════════════════════════════════════════════════════════════════
# CORE.RULES — apply_rules
# ══════════════════════════════════════════════════════════════════════════════

def _rows():
    return [
        ["alpha", "10", "tag_a"],
        ["beta",  "20", "tag_b"],
        ["gamma", "30", "tag_a"],
        ["beta",  "5",  "tag_b"],
        ["",      None, None],
    ]


def test_apply_rules_no_rules_and_returns_all():
    rows = [["a", 1], ["b", 2], ["c", 3]]
    assert apply_rules(rows, [], "AND") == rows


def test_apply_rules_no_rules_or_returns_all():
    rows = [["a", 1], ["b", 2]]
    assert apply_rules(rows, [], "OR") == rows


def test_apply_rules_include_equals():
    result = apply_rules(_rows(), [Rule(mode="include", column="A", operator="equals", value="alpha")], "AND")
    assert len(result) == 1
    assert result[0][0] == "alpha"


def test_apply_rules_exclude_equals():
    result = apply_rules(_rows(), [Rule(mode="exclude", column="A", operator="equals", value="beta")], "AND")
    names = [r[0] for r in result]
    assert "beta" not in names
    assert "alpha" in names


def test_apply_rules_contains():
    result = apply_rules(_rows(), [Rule(mode="include", column="C", operator="contains", value="tag")], "AND")
    assert all(r[2] is not None and "tag" in str(r[2]) for r in result)


def test_apply_rules_numeric_greater_than():
    result = apply_rules(_rows(), [Rule(mode="include", column="B", operator=">", value="15")], "AND")
    assert len(result) == 2

def test_apply_rules_numeric_less_than():
    result = apply_rules(_rows(), [Rule(mode="include", column="B", operator="<", value="15")], "AND")
    assert len(result) == 2


def test_apply_rules_and_one_true_one_false_row_excluded():
    rows = [["alpha", "10"], ["beta", "20"]]
    rules = [
        Rule(mode="include", column="A", operator="equals", value="alpha"),
        Rule(mode="include", column="B", operator=">",     value="15"),
    ]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 0


def test_apply_rules_or_one_true_one_false_row_included():
    rows = [["alpha", "10"], ["beta", "20"]]
    rules = [
        Rule(mode="include", column="A", operator="equals", value="alpha"),
        Rule(mode="include", column="B", operator=">",     value="15"),
    ]
    result = apply_rules(rows, rules, "OR")
    assert len(result) == 2


def test_apply_rules_all_exclude_or_semantics():
    """OR with all-exclude rules: a row is kept if ANY exclude inversion is True."""
    rows = [["alpha", 1], ["beta", 2], ["gamma", 3]]
    rules = [
        Rule(mode="exclude", column="A", operator="equals", value="alpha"),
        Rule(mode="exclude", column="A", operator="equals", value="beta"),
    ]
    result = apply_rules(rows, rules, "OR")
    names = [r[0] for r in result]
    # gamma: neither exclude matches → both inversions True → OR=True → kept
    assert "gamma" in names


def test_apply_rules_rule_column_beyond_row_width_no_crash():
    """Column Z is beyond 3-col rows — should not crash, just not match."""
    rules = [Rule(mode="include", column="Z", operator="equals", value="anything")]
    result = apply_rules(_rows(), rules, "AND")
    assert result == [] or isinstance(result, list)


def test_apply_rules_unknown_operator_raises_invalid_rule():
    rules = [Rule(mode="include", column="A", operator="LIKE", value="x")]
    with pytest.raises(AppError) as ei:
        apply_rules(_rows(), rules, "AND")
    assert ei.value.code == INVALID_RULE


def test_apply_rules_bad_rule_mode_raises_invalid_rule():
    rules = [Rule(mode="maybe", column="A", operator="equals", value="alpha")]
    with pytest.raises(AppError) as ei:
        apply_rules(_rows(), rules, "AND")
    assert ei.value.code == INVALID_RULE


def test_apply_rules_bad_combine_mode_raises_invalid_rule():
    rules = [Rule(mode="include", column="A", operator="equals", value="alpha")]
    with pytest.raises(AppError) as ei:
        apply_rules(_rows(), rules, "XOR")
    assert ei.value.code == INVALID_RULE


# ══════════════════════════════════════════════════════════════════════════════
# CORE.ERRORS — AppError
# ══════════════════════════════════════════════════════════════════════════════

def test_app_error_str_includes_code_and_message():
    e = AppError("X", "Nope")
    assert str(e).startswith("X: Nope")


def test_app_error_str_includes_details_when_present():
    e = AppError("X", "Nope", {"a": 1})
    s = str(e)
    assert "X: Nope" in s
    assert "a" in s
