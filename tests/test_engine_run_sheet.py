"""Tests for core.engine.run_sheet — single sheet extraction end-to-end."""
import os
import csv
import pytest
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet
from core.models import SheetConfig, Destination, Rule
from core.errors import AppError, DEST_BLOCKED


def _make_xlsx(path: str, sheet: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    data = [
        ["alpha", "x", 1, "foo"],
        ["beta",  "y", 2, "bar"],
        ["gamma", "z", 3, "baz"],
        ["beta",  "y", 4, "qux"],
    ]
    for r, row in enumerate(data, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)


def _make_csv(path: str):
    data = [["alpha","x",1,"foo"],["beta","y",2,"bar"],["gamma","z",3,"baz"],["beta","y",4,"qux"]]
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)


def _read(path, sheet, cells):
    ws = load_workbook(path)[sheet]
    return [ws[a].value for a in cells]


# ---- XLSX ----

def test_run_sheet_pack_columns_rules_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B1"] = "existing"; ws["C1"] = "existing2"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="pack", rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="beta")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        vals = _read(dest, "Out", ["B2","C2","B3","C3"])
        assert vals[0] == "beta" and vals[1] == 2
        assert vals[2] == "beta" and vals[3] == 4


def test_run_sheet_keep_mode_preserves_gaps():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1,3",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 3  # bounding box rows 1–3

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "alpha"
        assert ws["C1"].value is None   # gap column
        assert ws["D1"].value == 1
        assert ws["B2"].value is None   # gap row
        assert ws["B3"].value == "gamma"
        assert ws["D3"].value == 3


def test_run_sheet_explicit_start_row_lands_exactly():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-2",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="D", start_row="50"),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        ws = load_workbook(dest)["Out"]
        assert ws["D50"].value == "alpha"
        assert ws["E50"].value == 1
        assert ws["D51"].value == "beta"
        assert ws["E51"].value == 2


def test_run_sheet_dest_blocked_raises():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["D50"] = "BLOCK"; wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="D", start_row="50"),
        )
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg, recipe_name="R1")
        assert ei.value.code == DEST_BLOCKED


def test_run_sheet_append_two_runs_stack():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        r1 = run_sheet(src, cfg, recipe_name="R1")
        r2 = run_sheet(src, cfg, recipe_name="R1")
        assert r1.rows_written == 1
        assert r2.rows_written == 1

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "alpha" and ws["C1"].value == 1
        assert ws["B2"].value == "alpha" and ws["C2"].value == 1


def test_run_sheet_zero_rows_writes_nothing():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="pack", rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 0
        assert load_workbook(dest)["Out"]["B1"].value in (None, "")


# ---- CSV ----

def test_run_sheet_csv_pack_rules_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.csv")
        dest = os.path.join(td, "dest.xlsx")
        _make_csv(src)

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B1"] = "existing"; ws["C1"] = "existing2"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="pack", rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="beta")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        vals = _read(dest, "Out", ["B2","C2","B3","C3"])
        assert vals[0] == "beta"
        assert vals[1] in (2, "2")
        assert vals[2] == "beta"
        assert vals[3] in (4, "4")


# ---- source_start_row ----

def test_source_start_row_offsets_data():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        out = os.path.join(td, "out.xlsx")

        wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=f"A{i}")
            ws.cell(row=i, column=2, value=f"B{i}")
        wb.save(src)

        sh = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            source_start_row="3", columns_spec="A", rows_spec="1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=str(out), sheet_name="Dest", start_col="A", start_row="1"),
        )
        res = run_sheet(str(src), sh, recipe_name="R")
        assert res.rows_written == 1
        assert load_workbook(out)["Dest"]["A1"].value == "A3"


def test_source_start_row_all_rows_trimmed():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        out = os.path.join(td, "out.xlsx")

        wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=f"A{i}")
        wb.save(src)

        sh = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            source_start_row="3", columns_spec="A", rows_spec="",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=str(out), sheet_name="Dest", start_col="A", start_row="1"),
        )
        res = run_sheet(str(src), sh, recipe_name="R")
        assert res.rows_written == 3
        ws2 = load_workbook(out)["Dest"]
        assert [ws2[f"A{i}"].value for i in range(1, 4)] == ["A3", "A4", "A5"]
