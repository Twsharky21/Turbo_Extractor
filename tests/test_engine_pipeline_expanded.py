"""
Expanded tests for core.engine.run_sheet — pipeline ordering, error paths,
edge cases, and append/collision behaviour not covered by existing tests.
"""
import os
import pytest
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet
from core.models import SheetConfig, Destination, Rule
from core.errors import AppError, DEST_BLOCKED, SHEET_NOT_FOUND, BAD_SPEC


# ---- Helpers ----

def _make_xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if data:
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)


def _cfg(dest, columns="", rows="", mode="pack", rules=None, start_col="A",
         start_row="1", sheet_name="Sheet1", src_sheet="Sheet1",
         src_start_row="", rules_combine="AND"):
    return SheetConfig(
        name=sheet_name,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=rules_combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest, sheet_name="Out",
            start_col=start_col, start_row=start_row,
        ),
    )


# ============================================================
# PIPELINE ORDER: rules use absolute source columns (pre-column-selection)
# ============================================================

def test_pipeline_rules_reference_absolute_source_column_not_in_output():
    """
    Filter by column D (not in output), output only A and C.
    Rules must run against the original source columns, not the post-selection ones.
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[
            ["keep",   "x", 1, "YES"],
            ["drop",   "y", 2, "NO"],
            ["keep2",  "z", 3, "YES"],
        ])

        cfg = _cfg(
            dest, columns="A,C", rows="",
            rules=[Rule(mode="include", column="D", operator="equals", value="YES")],
        )
        result = run_sheet(src, cfg, recipe_name="R")
        assert result.rows_written == 2

        ws = load_workbook(dest)["Out"]
        assert ws["A1"].value == "keep"
        assert ws["B1"].value == 1
        assert ws["A2"].value == "keep2"
        assert ws["B2"].value == 3


def test_pipeline_rules_then_column_selection_order():
    """
    Verify: row selection → rules → column selection (not the reverse).
    If column selection ran first, the rule column would be gone.
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[
            ["alpha", 10, "tag_a"],
            ["beta",  20, "tag_b"],
            ["alpha", 30, "tag_a"],
        ])
        # Output only col A, filter by col C (which is not in output)
        cfg = _cfg(
            dest, columns="A", rows="",
            rules=[Rule(mode="include", column="C", operator="equals", value="tag_a")],
        )
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = load_workbook(dest)["Out"]
        assert ws["A1"].value == "alpha"
        assert ws["A2"].value == "alpha"


# ============================================================
# EMPTY / EDGE CASE SOURCES
# ============================================================

def test_empty_source_file_writes_nothing():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src)  # no data rows

        cfg = _cfg(dest, columns="", rows="")
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


def test_single_cell_output():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["hello", "world"]])

        cfg = _cfg(dest, columns="A", rows="1")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        assert load_workbook(dest)["Out"]["A1"].value == "hello"


def test_source_start_row_beyond_end_of_file_writes_nothing():
    """source_start_row past end of file → empty table → 0 rows written."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["a"], ["b"], ["c"]])  # 3 rows

        cfg = _cfg(dest, src_start_row="10")  # offset past end
        result = run_sheet(src, cfg)
        assert result.rows_written == 0


def test_keep_mode_all_rows_all_cols():
    """keep mode with blank rows_spec and columns_spec — full table."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[
            ["a", "b"],
            ["c", "d"],
        ])
        cfg = _cfg(dest, columns="", rows="", mode="keep")
        result = run_sheet(src, cfg)
        assert result.rows_written == 2
        ws = load_workbook(dest)["Out"]
        assert ws["A1"].value == "a"
        assert ws["B2"].value == "d"


# ============================================================
# ERROR PATHS
# ============================================================

def test_missing_sheet_raises_sheet_not_found():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, sheet="Sheet1")

        cfg = _cfg(dest, src_sheet="DoesNotExist")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == SHEET_NOT_FOUND


def test_bad_column_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["a", "b"]])

        cfg = _cfg(dest, columns="A,??")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == BAD_SPEC


def test_bad_row_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["a", "b"]])

        cfg = _cfg(dest, rows="nope")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == BAD_SPEC


def test_bad_start_row_zero_raises_bad_spec():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["a"]])

        cfg = _cfg(dest, start_row="0")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == BAD_SPEC


# ============================================================
# APPEND — LANDING ZONE ISOLATION
# ============================================================

def test_append_column_outside_landing_zone_does_not_affect_row():
    """
    A tall value in col A must NOT push the append row for landing zone B:C.
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["val1", "val2"]])

        # Pre-populate dest: col A has data up to row 100, landing zone B:C is empty
        wb = Workbook(); ws = wb.active; ws.title = "Out"
        for i in range(1, 101):
            ws[f"A{i}"] = f"noise_{i}"
        wb.save(dest)

        cfg = _cfg(dest, columns="A,B", start_col="B", start_row="")
        result = run_sheet(src, cfg)
        assert result.rows_written == 1

        ws2 = load_workbook(dest)["Out"]
        # Landing zone B:C was empty, so append starts at row 1
        assert ws2["B1"].value == "val1"
        assert ws2["C1"].value == "val2"


def test_append_formula_cell_treated_as_unoccupied():
    """Formula-only cells (no cached value) must not block append."""
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["new_data"]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "=SUM(B1:B10)"  # formula — planner treats as unoccupied
        wb.save(dest)

        cfg = _cfg(dest, columns="A", start_col="A", start_row="")
        result = run_sheet(src, cfg)
        # Formula cell is unoccupied → append at row 1, which means collision
        # check passes and write lands at row 1 (overwriting formula)
        assert result.rows_written == 1


def test_collision_blocked_on_second_row_of_shaped_output():
    """
    Shape has 3 rows; row 1 clear, row 2 blocked → DEST_BLOCKED.
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _make_xlsx(src, data=[["r1"], ["r2"], ["r3"]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"   # row 2 of the landing zone
        wb.save(dest)

        cfg = _cfg(dest, columns="A", rows="1-3", start_row="1")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == DEST_BLOCKED


# ============================================================
# PLANNER WRITE PLAN FIELDS
# ============================================================

def test_write_plan_width_and_height_correct():
    from openpyxl import Workbook
    from core.planner import build_plan

    wb = Workbook(); ws = wb.active
    shaped = [["a", "b", "c"], ["d", "e", "f"]]  # 2 rows x 3 cols
    plan = build_plan(ws, shaped, start_col_letters="B", start_row_str="5")

    assert plan.width == 3
    assert plan.height == 2
    assert plan.start_row == 5
    assert plan.start_col == 2   # B = 2


def test_build_plan_bad_start_col_raises():
    from openpyxl import Workbook
    from core.planner import build_plan

    wb = Workbook(); ws = wb.active
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], start_col_letters="1BAD", start_row_str="1")
    assert ei.value.code == BAD_SPEC
