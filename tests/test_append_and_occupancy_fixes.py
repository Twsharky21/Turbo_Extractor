"""
tests/test_append_and_occupancy_fixes.py

Regression tests for:
  1. Unified is_occupied — formula strings now unoccupied in both io and planner.
  2. Pack mode side-by-side column extractions with different start_col values
     correctly land in parallel columns, not stacked rows.
  3. run_all shared workbook cache — sequential writes to the same dest file
     correctly see prior writes within the same batch.
  4. Keep mode collision behaviour — target columns block, gap columns do not.
     (Updated from original full-bounding-box model to target-column model,
      which enables automatic merge of overlapping bounding boxes.)
"""
from __future__ import annotations

import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet, run_all
from core.errors import AppError, DEST_BLOCKED
from core.io import is_occupied
from core.models import Destination, Rule, SheetConfig
from core.planner import build_plan, is_cell_occupied


# ── helpers ───────────────────────────────────────────────────────────────────

def _xlsx(path: str, data, sheet: str = "Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _cfg(dest, *, columns="", rows="", mode="pack", rules=None, combine="AND",
         start_col="A", start_row="", dest_sheet="Out", src_sheet="Sheet1"):
    return SheetConfig(
        name=src_sheet, workbook_sheet=src_sheet,
        source_start_row="", columns_spec=columns, rows_spec=rows,
        paste_mode=mode, rules_combine=combine, rules=rules or [],
        destination=Destination(
            file_path=dest, sheet_name=dest_sheet,
            start_col=start_col, start_row=start_row,
        ),
    )


def _ws(path: str, sheet: str = "Out"):
    return load_workbook(path)[sheet]


# ══════════════════════════════════════════════════════════════════════════════
# 1. UNIFIED OCCUPANCY — is_occupied and is_cell_occupied must agree
# ══════════════════════════════════════════════════════════════════════════════

class TestUnifiedOccupancy:
    """
    io.is_occupied and planner.is_cell_occupied share the same definition for
    all non-formula values. They intentionally differ on formula strings:

      io.is_occupied           — used for SOURCE table used-range detection.
                                 A formula string in a source cell IS real content → occupied.

      planner.is_cell_occupied — used for DESTINATION scan and collision.
                                 A formula string with no cached result means the
                                 destination cell is effectively empty → unoccupied.
    """

    _common_cases = [
        (None,    False),
        ("",      False),
        ("hello", True),
        (" ",     True),
        ("N/A",   True),
        (0,       True),
        (0.0,     True),
        (False,   True),
        (True,    True),
        (42,      True),
        (-1,      True),
        (3.14,    True),
    ]

    def test_is_occupied_and_is_cell_occupied_agree_on_non_formula_values(self):
        """Both functions must agree on all non-formula values."""
        for val, expected in self._common_cases:
            io_result = is_occupied(val)
            planner_result = is_cell_occupied(val)
            assert io_result == expected, (
                f"io.is_occupied({val!r}) expected {expected}, got {io_result}"
            )
            assert planner_result == expected, (
                f"planner.is_cell_occupied({val!r}) expected {expected}, got {planner_result}"
            )

    def test_formula_strings_differ_by_design(self):
        """
        Formula strings are intentionally treated differently:
          - io.is_occupied: formula IS occupied (source-table content detection)
          - is_cell_occupied: formula is UNOCCUPIED (dest scan, no cached result = empty)
        """
        assert is_occupied("=SUM(A1)") is True
        assert is_cell_occupied("=SUM(A1)") is False
        assert is_occupied("=A1+B1") is True
        assert is_cell_occupied("=A1+B1") is False

    def test_formula_string_not_blocking_append_scan(self):
        """A formula-only cell must not advance the append row."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1:B10)"  # bare formula, no cached value

        plan = build_plan(ws, [["data"]], "A", "")
        assert plan is not None
        assert plan.start_row == 1  # formula didn't push append row

    def test_zero_in_landing_zone_does_advance_append_row(self):
        """Numeric zero is occupied and must advance the append row."""
        wb = Workbook()
        ws = wb.active
        ws["A3"] = 0  # zero is occupied

        plan = build_plan(ws, [["data"]], "A", "")
        assert plan is not None
        assert plan.start_row == 4  # zero at row 3 → append at 4

    def test_false_in_landing_zone_does_advance_append_row(self):
        """Boolean False is occupied and must advance the append row."""
        wb = Workbook()
        ws = wb.active
        ws["B2"] = False

        plan = build_plan(ws, [["a", "b"]], "A", "")
        assert plan is not None
        assert plan.start_row == 3  # False at B2 → append at 3


# ══════════════════════════════════════════════════════════════════════════════
# 2. PACK MODE SIDE-BY-SIDE: different start_col → parallel columns
# ══════════════════════════════════════════════════════════════════════════════

class TestPackSideBySide:
    """
    Two pack extractions with non-overlapping start_col values must land in
    parallel columns on the same row, not stack vertically.
    """

    def test_pack_AC_then_BD_land_side_by_side(self):
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["alpha", "skip_B", "charlie", "skip_D"]])
            _xlsx(srcB, [["skip_A", "bravo",  "skip_C",  "delta"]])

            cfgA = _cfg(dest, columns="A,C", start_col="A")
            cfgB = _cfg(dest, columns="B,D", start_col="C")

            r1 = run_sheet(srcA, cfgA)
            r2 = run_sheet(srcB, cfgB)

            assert r1.rows_written == 1
            assert r2.rows_written == 1

            ws = _ws(dest)
            assert ws["A1"].value == "alpha"
            assert ws["B1"].value == "charlie"
            assert ws["C1"].value == "bravo"
            assert ws["D1"].value == "delta"
            assert ws["A2"].value is None
            assert ws["C2"].value is None

    def test_pack_AC_then_BD_via_run_all(self):
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["alpha", "skip", "charlie"]])
            _xlsx(srcB, [["skip",  "bravo", "skip", "delta"]])

            cfgA = _cfg(dest, columns="A,C", start_col="A")
            cfgB = _cfg(dest, columns="B,D", start_col="C")

            report = run_all([(srcA, "R1", cfgA), (srcB, "R2", cfgB)])
            assert report.ok
            assert report.results[0].rows_written == 1
            assert report.results[1].rows_written == 1

            ws = _ws(dest)
            assert ws["A1"].value == "alpha"
            assert ws["B1"].value == "charlie"
            assert ws["C1"].value == "bravo"
            assert ws["D1"].value == "delta"

    def test_pack_multi_row_AC_then_BD_side_by_side(self):
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["a1", "x", "c1"], ["a2", "x", "c2"]])
            _xlsx(srcB, [["x", "b1", "x", "d1"], ["x", "b2", "x", "d2"]])

            cfgA = _cfg(dest, columns="A,C", start_col="A")
            cfgB = _cfg(dest, columns="B,D", start_col="C")

            r1 = run_sheet(srcA, cfgA)
            r2 = run_sheet(srcB, cfgB)

            assert r1.rows_written == 2
            assert r2.rows_written == 2

            ws = _ws(dest)
            assert ws["A1"].value == "a1"
            assert ws["B1"].value == "c1"
            assert ws["A2"].value == "a2"
            assert ws["B2"].value == "c2"
            assert ws["C1"].value == "b1"
            assert ws["D1"].value == "d1"
            assert ws["C2"].value == "b2"
            assert ws["D2"].value == "d2"

    def test_pack_same_start_col_does_stack_vertically(self):
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "s1.xlsx")
            src2 = os.path.join(td, "s2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(src1, [["r1c1", "r1c2"]])
            _xlsx(src2, [["r2c1", "r2c2"]])

            cfg1 = _cfg(dest, start_col="A")
            cfg2 = _cfg(dest, start_col="A")

            run_sheet(src1, cfg1)
            run_sheet(src2, cfg2)

            ws = _ws(dest)
            assert ws["A1"].value == "r1c1"
            assert ws["A2"].value == "r2c1"


# ══════════════════════════════════════════════════════════════════════════════
# 3. run_all SHARED WORKBOOK CACHE
# ══════════════════════════════════════════════════════════════════════════════

class TestRunAllSharedCache:

    def test_run_all_three_items_same_dest_stack_in_order(self):
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "dest.xlsx")
            sources = []
            for i in range(1, 4):
                p = os.path.join(td, f"s{i}.xlsx")
                _xlsx(p, [[f"row{i}_A", f"row{i}_B"]])
                sources.append((p, f"R{i}", _cfg(dest)))

            report = run_all(sources)
            assert report.ok

            ws = _ws(dest)
            assert ws["A1"].value == "row1_A"
            assert ws["A2"].value == "row2_A"
            assert ws["A3"].value == "row3_A"

    def test_run_all_different_start_cols_same_dest_same_row(self):
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["left_1", "left_2"]])
            _xlsx(srcB, [["right_1", "right_2"]])

            cfgA = _cfg(dest, start_col="A")
            cfgB = _cfg(dest, start_col="C")

            report = run_all([(srcA, "R1", cfgA), (srcB, "R2", cfgB)])
            assert report.ok

            ws = _ws(dest)
            assert ws["A1"].value == "left_1"
            assert ws["B1"].value == "left_2"
            assert ws["C1"].value == "right_1"
            assert ws["D1"].value == "right_2"
            assert ws["A2"].value is None
            assert ws["C2"].value is None

    def test_run_all_fail_fast_does_not_corrupt_prior_writes(self):
        with TemporaryDirectory() as td:
            src1 = os.path.join(td, "s1.xlsx")
            src2 = os.path.join(td, "s2.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(src1, [["good_data"]])
            _xlsx(src2, [["will_block"]])

            wb = Workbook()
            ws = wb.active
            ws.title = "Out"
            ws["A2"] = "BLOCKER"
            wb.save(dest)

            cfg1 = _cfg(dest, start_row="1")
            cfg2 = _cfg(dest, start_row="2")

            report = run_all([(src1, "R1", cfg1), (src2, "R2", cfg2)])
            assert not report.ok
            assert report.results[0].rows_written == 1
            assert report.results[1].error_code == DEST_BLOCKED

            ws_check = _ws(dest)
            assert ws_check["A1"].value == "good_data"


# ══════════════════════════════════════════════════════════════════════════════
# 4. KEEP MODE COLLISION — target columns block, gap columns allow merge
# ══════════════════════════════════════════════════════════════════════════════

class TestKeepCollision:
    """
    Collision detection operates on TARGET COLUMNS ONLY (columns that actually
    receive data). Gap columns produced by Keep Format bounding boxes are ignored.

    This enables automatic merge: two Keep extractions whose bounding boxes
    overlap can land side-by-side as long as their data columns don't conflict.

    Contract:
      - Target column occupied → DEST_BLOCKED
      - Gap column occupied (data from another extraction) → allowed, merge succeeds
    """

    def test_keep_AC_then_BD_overlapping_bboxes_merges(self):
        """
        Keep A,C → bbox A:C, target cols A and C (B is gap).
        Keep B,D at start_col=B → bbox B:D, target cols B and D (C is gap).
        C (from run 1) is a gap for run 2 → merge succeeds, no DEST_BLOCKED.
        Final result: A,B,C,D all populated on row 1.
        """
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["a", "b", "c"]])
            _xlsx(srcB, [["p", "q", "r", "s"]])

            cfgA = _cfg(dest, columns="A,C", mode="keep", start_col="A")
            cfgB = _cfg(dest, columns="B,D", mode="keep", start_col="B")

            r1 = run_sheet(srcA, cfgA)
            r2 = run_sheet(srcB, cfgB)

            assert r1.rows_written == 1
            assert r2.rows_written == 1

            ws = _ws(dest)
            assert ws["A1"].value == "a"    # from run 1
            assert ws["B1"].value == "q"    # from run 2 (col B of srcB)
            assert ws["C1"].value == "c"    # from run 1
            assert ws["D1"].value == "s"    # from run 2 (col D of srcB)

    def test_keep_target_col_occupied_raises_dest_blocked(self):
        """
        If a target column (one that receives actual data) is already occupied,
        DEST_BLOCKED must fire regardless of mode.
        """
        with TemporaryDirectory() as td:
            src = os.path.join(td, "src.xlsx")
            dest = os.path.join(td, "dest.xlsx")
            _xlsx(src, [["p", "q", "r", "s"]])

            # Pre-occupy B1 — which IS a target col for Keep B,D at start_col=B
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["B1"] = "EXISTING"
            wb.save(dest)

            cfgB = _cfg(dest, columns="B,D", mode="keep", start_col="B", start_row="1")
            with pytest.raises(AppError) as ei:
                run_sheet(src, cfgB)
            assert ei.value.code == DEST_BLOCKED

    def test_keep_overlapping_bboxes_explicit_row_blocked(self):
        """
        When both target cols of run 2 are clear but an explicit row is used,
        and one of the target cols IS actually occupied → DEST_BLOCKED.
        Here we pre-occupy D1 (a target col for Keep B,D) → blocked.
        """
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["a", "b_gap", "c"]])
            _xlsx(srcB, [["p", "q", "r", "s"]])

            cfgA = _cfg(dest, columns="A,C", mode="keep", start_col="A")
            run_sheet(srcA, cfgA)  # writes A1="a", C1="c"; B1 is gap (not written)

            # Pre-occupy D1 — a target col for Keep B,D
            wb = load_workbook(dest)
            wb["Out"]["D1"] = "BLOCK_D"
            wb.save(dest)

            # Explicit row 1: B and D are target cols; D1 is occupied → DEST_BLOCKED
            cfgB = _cfg(dest, columns="B,D", mode="keep", start_col="B", start_row="1")
            with pytest.raises(AppError) as ei:
                run_sheet(srcB, cfgB)
            assert ei.value.code == DEST_BLOCKED

    def test_keep_AC_then_BD_non_overlapping_start_col_ok(self):
        """
        Keep A,C at start_col=A → bbox A:C, data in A and C.
        Keep B,D at start_col=D → bbox D:G, data in D and F (no overlap with A:C).
        Both succeed.
        """
        with TemporaryDirectory() as td:
            srcA = os.path.join(td, "srcA.xlsx")
            srcB = os.path.join(td, "srcB.xlsx")
            dest = os.path.join(td, "dest.xlsx")

            _xlsx(srcA, [["a", "b_gap", "c"]])
            _xlsx(srcB, [["p", "q_gap", "r", "s"]])

            cfgA = _cfg(dest, columns="A,C", mode="keep", start_col="A")
            cfgB = _cfg(dest, columns="B,D", mode="keep", start_col="D")

            r1 = run_sheet(srcA, cfgA)
            r2 = run_sheet(srcB, cfgB)

            assert r1.rows_written == 1
            assert r2.rows_written == 1

            ws = _ws(dest)
            assert ws["A1"].value == "a"
            assert ws["B1"].value is None       # gap
            assert ws["C1"].value == "c"
            # Keep B,D → bbox width 3 at start_col=D: D=q_gap(gap), E=None(gap), F=s
            assert ws["D1"].value == "q_gap"    # col B of srcB (idx 1, offset 0 of bbox)
            assert ws["E1"].value is None       # gap
            assert ws["F1"].value == "s"        # col D of srcB (idx 3, offset 2 of bbox)
