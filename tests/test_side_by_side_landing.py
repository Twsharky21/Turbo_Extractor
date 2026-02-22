"""
tests/test_side_by_side_landing.py

Tests that prove landing-zone isolation works correctly for side-by-side
extractions. Two extractions targeting different start_col values on the
same destination sheet must land on the same row, not stack vertically.

This covers the exact scenario from the screenshot bug report:
- Source 1: names + amounts → dest cols A,B
- Source 2: dates + market values → dest cols C,D
- Both in append mode → both should land at row 1 (cols independent)
"""
from __future__ import annotations

import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet, run_all
from core.errors import AppError, DEST_BLOCKED
from core.models import Destination, SheetConfig


def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)


def _cfg(dest, *, start_col="A", start_row="", dest_sheet="Out", src_sheet="Sheet1",
         columns="", rows="", mode="pack"):
    return SheetConfig(
        name=src_sheet, workbook_sheet=src_sheet,
        source_start_row="", columns_spec=columns, rows_spec=rows,
        paste_mode=mode, rules_combine="AND", rules=[],
        destination=Destination(
            file_path=dest, sheet_name=dest_sheet,
            start_col=start_col, start_row=start_row,
        ),
    )


# ══════════════════════════════════════════════════════════════════════════════
# CORE: side-by-side isolation
# ══════════════════════════════════════════════════════════════════════════════

def test_side_by_side_two_run_sheets_land_on_same_row():
    """
    Two sequential run_sheet calls with different start_col values must
    land on the same row (row 1), not stack vertically.

    This is the exact screenshot scenario:
    - Run 1: 7 rows → cols A,B (start_col=A)
    - Run 2: 6 rows → cols C,D (start_col=C)
    - Both in append mode → C,D are empty → run 2 starts at row 1
    """
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "src1.xlsx")
        src2 = os.path.join(td, "src2.xlsx")
        dest = os.path.join(td, "dest.xlsx")

        _xlsx(src1, [
            ["Jackson Brown",   25456.56],
            ["Liam Patel",      50930.89],
            ["Charlotte Clark", 213858.29],
            ["Noah Johnson",    52865.79],
            ["Michael King",    66842.41],
            ["Liam Patel",      38954.61],
            ["Liam Patel",      101039.79],
        ])
        _xlsx(src2, [
            ["2024-01-19", 104982.5],
            ["2024-01-19", 35658.55],
            ["2024-01-19", 61963.96],
            ["2024-01-22", 103981],
            ["2024-01-22", 183775.8],
            ["2024-01-26", 115904],
        ])

        r1 = run_sheet(src1, _cfg(dest, start_col="A"))
        r2 = run_sheet(src2, _cfg(dest, start_col="C"))

        assert r1.rows_written == 7
        assert r2.rows_written == 6

        ws = load_workbook(dest)["Out"]

        # Run 1 data in A,B starting row 1
        assert ws["A1"].value == "Jackson Brown"
        assert ws["B1"].value == 25456.56
        assert ws["A7"].value == "Liam Patel"

        # Run 2 data in C,D starting row 1 — NOT row 8
        assert ws["C1"].value == "2024-01-19"
        assert ws["D1"].value == 104982.5
        assert ws["C6"].value == "2024-01-26"

        # No overlap: A8, C7 are empty
        assert ws["A8"].value is None
        assert ws["C7"].value is None


def test_side_by_side_via_run_all():
    """run_all with shared cache also lands side-by-side correctly."""
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "src1.xlsx")
        src2 = os.path.join(td, "src2.xlsx")
        dest = os.path.join(td, "dest.xlsx")

        _xlsx(src1, [["name1", 100], ["name2", 200], ["name3", 300]])
        _xlsx(src2, [["date1", 111], ["date2", 222]])

        report = run_all([
            (src1, "R1", _cfg(dest, start_col="A")),
            (src2, "R2", _cfg(dest, start_col="C")),
        ])

        assert report.ok
        ws = load_workbook(dest)["Out"]

        # Both land at row 1 in their respective columns
        assert ws["A1"].value == "name1"
        assert ws["B1"].value == 100
        assert ws["C1"].value == "date1"
        assert ws["D1"].value == 111

        # Columns are independent — no vertical stacking
        assert ws["A3"].value == "name3"
        assert ws["C3"].value is None   # src2 only has 2 rows


def test_same_start_col_stacks_vertically():
    """
    When both extractions use the same start_col, they SHOULD stack
    vertically — this is correct append behaviour, not a bug.
    """
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "src1.xlsx")
        src2 = os.path.join(td, "src2.xlsx")
        dest = os.path.join(td, "dest.xlsx")

        _xlsx(src1, [["a", 1], ["b", 2]])
        _xlsx(src2, [["c", 3], ["d", 4]])

        run_sheet(src1, _cfg(dest, start_col="A"))
        run_sheet(src2, _cfg(dest, start_col="A"))

        ws = load_workbook(dest)["Out"]
        assert ws["A1"].value == "a"
        assert ws["A2"].value == "b"
        assert ws["A3"].value == "c"  # stacked below
        assert ws["A4"].value == "d"


def test_landing_zone_ignores_noise_in_other_columns():
    """
    Col A has 100 rows of data. Extraction targets cols B,C.
    Append row for B,C must be 1 (not 101).
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")

        _xlsx(src, [["val1", "val2"]])

        # Pre-populate dest: col A has 100 rows of noise
        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        for i in range(1, 101):
            ws[f"A{i}"] = f"noise_{i}"
        wb.save(dest)

        r = run_sheet(src, _cfg(dest, start_col="B"))
        assert r.rows_written == 1

        ws2 = load_workbook(dest)["Out"]
        assert ws2["B1"].value == "val1"   # landed at row 1, not row 101
        assert ws2["C1"].value == "val2"
        assert ws2["B2"].value is None


def test_collision_still_fires_when_target_cols_occupied():
    """
    Even with side-by-side config, if the target cols ARE occupied,
    DEST_BLOCKED must fire.
    """
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        _xlsx(src, [["x", "y"]])

        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["C1"] = "BLOCK"
        wb.save(dest)

        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_col="C", start_row="1"))
        assert ei.value.code == DEST_BLOCKED


def test_three_side_by_side_extractions():
    """Three extractions each in their own column group land at row 1."""
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "dest.xlsx")
        items = []
        for i, col in enumerate(["A", "C", "E"]):
            src = os.path.join(td, f"src{i}.xlsx")
            _xlsx(src, [[f"r1c{i}", f"r2c{i}"], [f"r3c{i}", f"r4c{i}"]])
            items.append((src, f"R{i}", _cfg(dest, start_col=col)))

        report = run_all(items)
        assert report.ok

        ws = load_workbook(dest)["Out"]
        assert ws["A1"].value == "r1c0"
        assert ws["C1"].value == "r1c1"
        assert ws["E1"].value == "r1c2"
        assert ws["A2"].value == "r3c0"
        assert ws["C2"].value == "r3c1"
        assert ws["E2"].value == "r3c2"
