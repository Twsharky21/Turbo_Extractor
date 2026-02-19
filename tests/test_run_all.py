\
import os
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_all
from core.models import SheetConfig, Destination
from core.errors import DEST_BLOCKED


def make_source_xlsx(path: str, sheet_name: str, tag: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = f"{tag}-alpha"
    ws["B1"] = "x"
    ws["C1"] = 1
    wb.save(path)


def test_run_all_two_items_success_same_dest_stacks():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "out.xlsx")

        make_source_xlsx(src1, "Sheet1", "S1")
        make_source_xlsx(src2, "Sheet1", "S2")

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (src1, "R1", cfg1),
            (src2, "R2", cfg2),
        ])

        assert report.ok is True
        assert len(report.results) == 2
        assert report.results[0].rows_written == 1
        assert report.results[1].rows_written == 1

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "S1-alpha"
        assert ws["C1"].value == 1
        assert ws["B2"].value == "S2-alpha"
        assert ws["C2"].value == 1


def test_run_all_fail_fast_records_error_and_stops():
    with TemporaryDirectory() as td:
        src1 = os.path.join(td, "s1.xlsx")
        src2 = os.path.join(td, "s2.xlsx")
        dest = os.path.join(td, "out.xlsx")

        make_source_xlsx(src1, "Sheet1", "S1")
        make_source_xlsx(src2, "Sheet1", "S2")

        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B1"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )

        cfg_should_not_run = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (src1, "R1", cfg_blocked),
            (src2, "R2", cfg_should_not_run),
        ])

        assert report.ok is False
        assert len(report.results) == 1
        assert report.results[0].error_code == DEST_BLOCKED

        wb2 = load_workbook(dest)
        ws2 = wb2["Out"]
        values = [ws2["B1"].value, ws2["B2"].value, ws2["B3"].value]
        assert "S2-alpha" not in values
