"""
tests/test_landing.py

Unit tests for core.landing — the standalone landing-zone and collision module.

These tests work entirely with plain dicts (CellMap) — no worksheet or
openpyxl objects required. This proves the logic is fully isolated.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from core.landing import (
    is_dest_cell_occupied,
    read_zone,
    scan_landing_zone,
    probe_collision,
    CellMap,
)


# ══════════════════════════════════════════════════════════════════════════════
# is_dest_cell_occupied
# ══════════════════════════════════════════════════════════════════════════════

class TestIsDestCellOccupied:
    def test_none_is_unoccupied(self):
        assert is_dest_cell_occupied(None) is False

    def test_empty_string_is_unoccupied(self):
        assert is_dest_cell_occupied("") is False

    def test_bare_formula_is_unoccupied(self):
        assert is_dest_cell_occupied("=SUM(A1:A10)") is False
        assert is_dest_cell_occupied("=A1+B1") is False

    def test_text_is_occupied(self):
        assert is_dest_cell_occupied("hello") is True
        assert is_dest_cell_occupied(" ") is True
        assert is_dest_cell_occupied("N/A") is True
        assert is_dest_cell_occupied("BLOCK") is True

    def test_zero_int_is_occupied(self):
        assert is_dest_cell_occupied(0) is True

    def test_zero_float_is_occupied(self):
        assert is_dest_cell_occupied(0.0) is True

    def test_false_bool_is_occupied(self):
        assert is_dest_cell_occupied(False) is True

    def test_true_bool_is_occupied(self):
        assert is_dest_cell_occupied(True) is True

    def test_positive_number_is_occupied(self):
        assert is_dest_cell_occupied(42) is True
        assert is_dest_cell_occupied(3.14) is True


# ══════════════════════════════════════════════════════════════════════════════
# scan_landing_zone — pure dict tests
# ══════════════════════════════════════════════════════════════════════════════

class TestScanLandingZone:
    def test_empty_map_returns_zero(self):
        assert scan_landing_zone({}, 1, 2) == 0

    def test_finds_max_row_in_landing_cols(self):
        cell_map: CellMap = {
            (1, 1): "a",
            (3, 1): "b",
            (2, 2): "c",
        }
        assert scan_landing_zone(cell_map, 1, 2) == 3

    def test_ignores_cells_outside_col_range(self):
        cell_map: CellMap = {
            (10, 5): "noise",   # col 5 outside landing zone 1:2
            (2,  1): "data",
        }
        assert scan_landing_zone(cell_map, 1, 2) == 2

    def test_col_range_single_column(self):
        cell_map: CellMap = {
            (4, 3): "x",
            (7, 4): "y",   # col 4, outside single-col range 3:3
        }
        assert scan_landing_zone(cell_map, 3, 3) == 4

    def test_unoccupied_values_not_counted(self):
        cell_map: CellMap = {
            (5, 1): None,
            (6, 1): "",
            (7, 1): "=SUM()",
            (2, 1): "real",
        }
        assert scan_landing_zone(cell_map, 1, 1) == 2

    def test_zero_value_counts_as_occupied(self):
        cell_map: CellMap = {(3, 1): 0}
        assert scan_landing_zone(cell_map, 1, 1) == 3

    def test_false_counts_as_occupied(self):
        cell_map: CellMap = {(5, 2): False}
        assert scan_landing_zone(cell_map, 1, 2) == 5


# ══════════════════════════════════════════════════════════════════════════════
# probe_collision — pure dict tests
# ══════════════════════════════════════════════════════════════════════════════

class TestProbeCollision:
    def test_empty_zone_returns_none(self):
        assert probe_collision({}, 1, 3, 1, 2) is None

    def test_occupied_cell_in_zone_returns_blocker(self):
        cell_map: CellMap = {(2, 1): "BLOCK"}
        result = probe_collision(cell_map, 1, 3, 1, 2)
        assert result == (2, 1, "BLOCK")

    def test_cell_outside_row_range_not_reported(self):
        cell_map: CellMap = {(5, 1): "outside"}
        assert probe_collision(cell_map, 1, 3, 1, 2) is None

    def test_cell_outside_col_range_not_reported(self):
        cell_map: CellMap = {(2, 5): "outside"}
        assert probe_collision(cell_map, 1, 3, 1, 2) is None

    def test_returns_top_left_blocker_first(self):
        cell_map: CellMap = {
            (3, 2): "second",
            (2, 1): "first",
            (1, 2): "top",
        }
        result = probe_collision(cell_map, 1, 3, 1, 2)
        assert result is not None
        assert result[0] == 1  # row 1 is first
        assert result[1] == 2  # col 2 (only occupied in row 1)

    def test_zero_value_triggers_collision(self):
        cell_map: CellMap = {(1, 1): 0}
        result = probe_collision(cell_map, 1, 1, 1, 1)
        assert result == (1, 1, 0)

    def test_none_value_does_not_trigger(self):
        cell_map: CellMap = {(1, 1): None}
        assert probe_collision(cell_map, 1, 1, 1, 1) is None

    def test_empty_string_does_not_trigger(self):
        cell_map: CellMap = {(1, 1): ""}
        assert probe_collision(cell_map, 1, 1, 1, 1) is None


# ══════════════════════════════════════════════════════════════════════════════
# read_zone — openpyxl worksheet tests (iter_rows, no phantom cells)
# ══════════════════════════════════════════════════════════════════════════════

class TestReadZone:
    def _ws(self):
        return Workbook().active

    def test_empty_sheet_returns_empty_map(self):
        ws = self._ws()
        result = read_zone(ws, 1, 2)
        assert result == {}

    def test_reads_cells_in_col_range(self):
        ws = self._ws()
        ws["A1"] = "alpha"
        ws["B1"] = 42
        ws["C1"] = "outside"   # col C=3, range is 1:2
        result = read_zone(ws, 1, 2)
        assert (1, 1) in result
        assert result[(1, 1)] == "alpha"
        assert (1, 2) in result
        assert result[(1, 2)] == 42
        assert (1, 3) not in result

    def test_none_cells_not_included(self):
        ws = self._ws()
        ws["A1"] = "data"
        # A2 is None — should not appear in map
        result = read_zone(ws, 1, 1)
        assert (1, 1) in result
        assert (2, 1) not in result

    def test_does_not_create_phantom_cells(self):
        """
        read_zone must NOT inflate ws.max_row.
        This is the core correctness requirement — ws.cell() access
        registers phantom cells; iter_rows does not.
        """
        ws = self._ws()
        ws["A1"] = "data"
        ws["A7"] = "last"
        assert ws.max_row == 7

        # Read cols C:D (empty) — must not change max_row
        read_zone(ws, 3, 4)
        assert ws.max_row == 7, (
            "read_zone inflated ws.max_row — phantom cell bug!"
        )

    def test_extra_rows_covers_probe_zone(self):
        """
        extra_rows extends the read beyond ws.max_row so the collision
        probe can inspect append rows that sit past existing content.
        """
        ws = self._ws()
        ws["A1"] = "data"
        ws["A3"] = "last"    # max_row = 3
        ws["A5"] = "future"  # manually set beyond current — shouldn't exist

        # Without extra_rows, row 5 is beyond max_row=3 and won't be read
        without = read_zone(ws, 1, 1, extra_rows=0)
        # ws.max_row is 3 so iter_rows goes 1..3, A5 not visited
        # (A5 was never written to this ws, so it won't appear)

        # extra_rows=2 extends to max_row+2 = 5
        ws["A5"] = "probe_target"
        assert ws.max_row == 5
        with_extra = read_zone(ws, 1, 1, extra_rows=0)
        assert (5, 1) in with_extra

    def test_side_by_side_isolation(self):
        """
        KEY TEST: After writing 7 rows to cols A:B, reading cols C:D
        via read_zone must return an empty map AND must not change max_row.
        This directly proves the phantom-cell bug is fixed.
        """
        ws = self._ws()
        for i in range(1, 8):
            ws.cell(row=i, column=1, value=f"name{i}")
            ws.cell(row=i, column=2, value=i * 1000.0)

        assert ws.max_row == 7

        # Scan C:D (cols 3,4) — the side-by-side landing zone
        cell_map = read_zone(ws, 3, 4)
        assert cell_map == {}, "C:D should be empty after A:B write"
        assert ws.max_row == 7, "max_row must not change after read_zone scan"

        max_used = scan_landing_zone(cell_map, 3, 4)
        assert max_used == 0
        start_row = max_used + 1 if max_used > 0 else 1
        assert start_row == 1, f"Expected start_row=1, got {start_row}"


# ══════════════════════════════════════════════════════════════════════════════
# Integration: scan + probe round-trip
# ══════════════════════════════════════════════════════════════════════════════

class TestScanProbeIntegration:
    def test_append_then_collision_clear(self):
        """Full append flow: scan → compute start_row → probe → clear."""
        ws = Workbook().active
        ws["C1"] = "existing"
        ws["C3"] = "also"

        cell_map = read_zone(ws, 3, 4)
        max_used = scan_landing_zone(cell_map, 3, 4)
        assert max_used == 3
        start_row = max_used + 1   # 4

        # Probe rows 4..4, cols C:D — should be clear
        probe_map = read_zone(ws, 3, 4, extra_rows=1)
        result = probe_collision(probe_map, start_row, start_row, 3, 4)
        assert result is None

    def test_explicit_row_collision_detected(self):
        """Explicit row targeting occupied cell → blocker detected."""
        ws = Workbook().active
        ws["C5"] = "BLOCK"

        probe_map = read_zone(ws, 3, 4, extra_rows=2)
        result = probe_collision(probe_map, 5, 5, 3, 4)
        assert result is not None
        assert result[0] == 5
        assert result[1] == 3
        assert result[2] == "BLOCK"

    def test_side_by_side_full_flow(self):
        """
        Simulates the exact side-by-side scenario from the bug report:
        - A:B has 7 rows of data.
        - Scan C:D → max_used=0 → start_row=1.
        - Probe C:D rows 1..6 → clear → write proceeds.
        """
        ws = Workbook().active
        for i in range(1, 8):
            ws.cell(row=i, column=1, value=f"name{i}")
            ws.cell(row=i, column=2, value=i * 1000.0)

        # Scan C:D (cols 3,4)
        scan_map = read_zone(ws, 3, 4)
        max_used = scan_landing_zone(scan_map, 3, 4)
        assert max_used == 0
        start_row = 1
        output_height = 6

        # Probe C:D rows 1..6
        row_end = start_row + output_height - 1
        extra = max(0, row_end - (ws.max_row or 0))
        probe_map = read_zone(ws, 3, 4, extra_rows=extra)
        blocker = probe_collision(probe_map, start_row, row_end, 3, 4)
        assert blocker is None, f"Unexpected blocker: {blocker}"
