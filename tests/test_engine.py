"""
test_engine.py — Consolidated engine integration tests.

Covers:
  - core.engine.run_sheet: XLSX/CSV extraction, pack/keep, rules, append, error paths
  - core.engine.run_all: batch execution, stacking, fail-fast, progress callbacks
  - Pipeline ordering: rules use absolute source columns (pre-column-selection)
  - Append / landing-zone isolation
  - Collision detection end-to-end
  - Engine hardening: wide sheets, unicode, mixed types, CSV edge cases
  - Pack/Keep combination matrix (30+ scenarios)
  - Rules and transform edge cases (error paths, OR/AND, operators)
  - Landing zone and side-by-side merge behaviour (core.landing)
"""
from __future__ import annotations

import csv
import os
import time
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_all, run_sheet
from core.errors import AppError, BAD_SPEC, DEST_BLOCKED, SHEET_NOT_FOUND
from core.landing import find_target_col_offsets, read_zone
from core.models import Destination, Rule, SheetConfig
from core.parsing import col_index_to_letters, col_letters_to_index


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _make_xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if data:
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)
    return path


def _make_csv(path: str, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path: str, sheet: str = "Out"):
    return load_workbook(path)[sheet]


def _pack_cfg(dest, sheet="Out", src_col="B", rows="1-1", columns="A,C"):
    return SheetConfig(
        name="Sheet1", workbook_sheet="Sheet1",
        columns_spec=columns, rows_spec=rows,
        paste_mode="pack", rules_combine="AND", rules=[],
        destination=Destination(file_path=dest, sheet_name=sheet,
                                start_col=src_col, start_row=""),
    )


# Standard 4-row source data
_STD_DATA = [
    ["alpha", "x", 1, "foo"],
    ["beta",  "y", 2, "bar"],
    ["gamma", "z", 3, "baz"],
    ["beta",  "y", 4, "qux"],
]


# ══════════════════════════════════════════════════════════════════════════════
# RUN_SHEET — BASIC EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def test_run_sheet_pack_columns_rules_append():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=_STD_DATA)
        dest = os.path.join(td, "dest.xlsx")

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B1"] = "existing"; ws["C1"] = "existing2"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="pack", rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="beta")],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="B", start_row=""),
        )
        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        ws2 = _ws(dest)
        assert ws2["B2"].value == "beta" and ws2["C2"].value == 2
        assert ws2["B3"].value == "beta" and ws2["C3"].value == 4


def test_run_sheet_keep_mode_preserves_gaps():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=_STD_DATA)
        dest = os.path.join(td, "dest.xlsx")

        cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1,3",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="A", start_row=""),
        )
        result = run_sheet(src, cfg)
        assert result.rows_written == 3  # bounding box rows 1-3

        ws2 = _ws(dest)
        assert ws2["A1"].value == "alpha"
        assert ws2["B1"].value is None   # gap
        assert ws2["C1"].value == 1


def test_run_sheet_csv_source():
    with TemporaryDirectory() as td:
        src  = _make_csv(os.path.join(td, "src.csv"), _STD_DATA)
        dest = os.path.join(td, "dest.xlsx")

        cfg = _cfg(dest, columns="A,B")
        result = run_sheet(src, cfg)
        assert result.rows_written == 4

        ws2 = _ws(dest)
        assert ws2["A1"].value == "alpha"
        assert ws2["B1"].value == "x"


def test_run_sheet_all_columns_when_blank_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b", "c"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns=""))
        assert result.rows_written == 1
        ws2 = _ws(dest)
        assert ws2["C1"].value == "c"


def test_run_sheet_result_message_ok():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.message == "OK"


def test_run_sheet_result_message_zero_rows():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, rules=[Rule(mode="include", column="A",
                                      operator="equals", value="NO_MATCH")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 0
        assert result.message == "0 rows written"


def test_run_sheet_source_start_row_offset():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["header"], ["row1"], ["row2"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, src_start_row="2"))
        assert result.rows_written == 2
        assert _ws(dest)["A1"].value == "row1"


def test_run_sheet_source_start_row_1_same_as_no_offset():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"], ["b"]])
        dest = os.path.join(td, "dest.xlsx")
        r1   = run_sheet(src, _cfg(dest, src_start_row="1"))
        dest2 = os.path.join(td, "dest2.xlsx")
        r2   = run_sheet(src, _cfg(dest2, src_start_row=""))
        assert r1.rows_written == r2.rows_written == 2


def test_run_sheet_source_start_row_non_numeric_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="abc"))


def test_run_sheet_source_start_row_zero_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="0"))


def test_run_sheet_source_start_row_negative_raises():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError):
            run_sheet(src, _cfg(dest, src_start_row="-1"))


def test_run_sheet_source_start_row_past_end_zero_rows():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"], ["b"], ["c"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, src_start_row="10"))
        assert result.rows_written == 0


def test_run_sheet_keep_mode_all_rows_all_cols():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["a", "b"], ["c", "d"]])
        dest = os.path.join(td, "dest.xlsx")
        result = run_sheet(src, _cfg(dest, columns="", rows="", mode="keep"))
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "a"
        assert ws2["B2"].value == "d"


def test_run_sheet_two_calls_same_file_different_dest_sheets():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["v1"], ["v2"]])
        dest = os.path.join(td, "dest.xlsx")

        r1 = run_sheet(src, _cfg(dest, dest_sheet="Sheet1"))
        r2 = run_sheet(src, _cfg(dest, dest_sheet="Sheet2"))

        assert r1.rows_written == 2
        assert r2.rows_written == 2
        wb = load_workbook(dest)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames


# ══════════════════════════════════════════════════════════════════════════════
# RUN_SHEET — DESTINATION SHEET MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

def test_new_dest_file_gets_default_sheet_removed(tmp_path):
    src  = _make_xlsx(str(tmp_path / "src.xlsx"), data=[["a"]])
    dest = str(tmp_path / "dest.xlsx")
    run_sheet(src, _cfg(dest, dest_sheet="MyOutput"))
    wb = load_workbook(dest)
    assert "MyOutput" in wb.sheetnames
    assert "Sheet" not in wb.sheetnames or len(wb.sheetnames) == 1


def test_run_sheet_multiple_dest_sheets_preserved(tmp_path):
    wb = Workbook()
    wb.active.title = "Existing"
    wb["Existing"]["A1"] = "keep_me"
    wb.create_sheet("Other")
    wb["Other"]["A1"] = "also_keep"
    dest = str(tmp_path / "dest.xlsx")
    wb.save(dest)

    src = _make_xlsx(str(tmp_path / "src.xlsx"), data=[["new"]])
    run_sheet(src, _cfg(dest, dest_sheet="Existing", start_row="2"))

    wb2 = load_workbook(dest)
    assert wb2["Existing"]["A1"].value == "keep_me"
    assert wb2["Existing"]["A2"].value == "new"
    assert wb2["Other"]["A1"].value == "also_keep"


# ══════════════════════════════════════════════════════════════════════════════
# RUN_SHEET — ERROR PATHS
# ══════════════════════════════════════════════════════════════════════════════

def test_missing_sheet_raises_sheet_not_found():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"))
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, src_sheet="DoesNotExist")
        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg)
        assert ei.value.code == SHEET_NOT_FOUND


def test_bad_column_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, columns="A,??"))
        assert ei.value.code == BAD_SPEC


def test_bad_row_spec_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a", "b"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, rows="nope"))
        assert ei.value.code == BAD_SPEC


def test_bad_start_row_zero_raises_bad_spec():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="0"))
        assert ei.value.code == BAD_SPEC


def test_collision_raises_dest_blocked():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["a"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="1"))
        assert ei.value.code == DEST_BLOCKED


def test_collision_blocked_on_inner_row_of_output():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"),
                          data=[["r1"], ["r2"], ["r3"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"
        wb.save(dest)
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, start_row="1"))
        assert ei.value.code == DEST_BLOCKED


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE ORDER
# ══════════════════════════════════════════════════════════════════════════════

def test_pipeline_rules_use_absolute_source_column_not_in_output():
    """Rules must run against original source columns, not post-selection columns."""
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[
            ["keep",  "x", 1, "YES"],
            ["drop",  "y", 2, "NO"],
            ["keep2", "z", 3, "YES"],
        ])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, columns="A,C",
                    rules=[Rule(mode="include", column="D",
                                operator="equals", value="YES")])
        result = run_sheet(src, cfg, recipe_name="R")
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "keep"
        assert ws2["B1"].value == 1
        assert ws2["A2"].value == "keep2"
        assert ws2["B2"].value == 3


def test_pipeline_rules_then_column_selection_order():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[
            ["alpha", "x", 1, "YES"],
            ["beta",  "y", 2, "NO"],
        ])
        dest = os.path.join(td, "dest.xlsx")
        cfg  = _cfg(dest, columns="A,B",
                    rules=[Rule(mode="include", column="D",
                                operator="equals", value="YES")])
        result = run_sheet(src, cfg)
        assert result.rows_written == 1
        assert _ws(dest)["A1"].value == "alpha"


# ══════════════════════════════════════════════════════════════════════════════
# APPEND — LANDING ZONE ISOLATION
# ══════════════════════════════════════════════════════════════════════════════

def test_append_column_outside_landing_zone_does_not_affect_row():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["val1", "val2"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        for i in range(1, 101):
            ws[f"A{i}"] = f"noise_{i}"
        wb.save(dest)

        result = run_sheet(src, _cfg(dest, columns="A,B", start_col="B"))
        assert result.rows_written == 1
        ws2 = _ws(dest)
        assert ws2["B1"].value == "val1"
        assert ws2["C1"].value == "val2"


def test_append_formula_cell_treated_as_unoccupied():
    with TemporaryDirectory() as td:
        src  = _make_xlsx(os.path.join(td, "src.xlsx"), data=[["new_data"]])
        dest = os.path.join(td, "dest.xlsx")
        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "=SUM(B1:B10)"
        wb.save(dest)
        result = run_sheet(src, _cfg(dest, columns="A", start_col="A"))
        assert result.rows_written == 1


# ══════════════════════════════════════════════════════════════════════════════
# RUN_ALL — BATCH EXECUTION
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_two_sources_stack_same_dest():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["S1", "x", 1]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["S2", "x", 2]])

        report = run_all([(s1, "R1", _pack_cfg(dest)),
                          (s2, "R2", _pack_cfg(dest))])
        assert report.ok
        ws2 = load_workbook(dest)["Out"]
        assert ws2["B1"].value == "S1" and ws2["C1"].value == 1
        assert ws2["B2"].value == "S2" and ws2["C2"].value == 2


def test_run_all_three_sources_stack_in_order():
    with TemporaryDirectory() as td:
        dest    = os.path.join(td, "out.xlsx")
        sources = []
        for i in range(1, 4):
            p = _make_xlsx(os.path.join(td, f"s{i}.xlsx"),
                           data=[[f"A{i}", "x", i]])
            sources.append((p, f"R{i}", _pack_cfg(dest)))

        report = run_all(sources)
        assert report.ok
        ws2 = load_workbook(dest)["Out"]
        for i in range(1, 4):
            assert ws2[f"B{i}"].value == f"A{i}"
            assert ws2[f"C{i}"].value == i


def test_run_all_two_different_destinations():
    with TemporaryDirectory() as td:
        d1 = os.path.join(td, "o1.xlsx")
        d2 = os.path.join(td, "o2.xlsx")
        s1 = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["A1", "x", 1]])
        s2 = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["A2", "x", 2]])

        report = run_all([(s1, "R1", _pack_cfg(d1)),
                          (s2, "R2", _pack_cfg(d2))])
        assert report.ok
        assert load_workbook(d1)["Out"]["B1"].value == "A1"
        assert load_workbook(d2)["Out"]["B1"].value == "A2"


def test_run_all_empty_iterable_returns_ok():
    report = run_all([])
    assert report.ok
    assert report.results == []


def test_run_all_generator_input_works():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["x"]])

        def gen():
            yield (s1, "R1", _pack_cfg(dest, columns="A", src_col="A"))

        report = run_all(gen())
        assert report.ok
        assert report.results[0].rows_written == 1


def test_run_all_fail_fast_on_collision():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["r1"], ["r2"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["x"]])

        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = _cfg(dest, columns="A", start_col="A", start_row="1")
        cfg_second  = _cfg(dest, columns="A", start_col="A")

        report = run_all([(s1, "R1", cfg_blocked), (s2, "R2", cfg_second)])
        assert not report.ok
        assert len(report.results) == 1
        assert report.results[0].error_code == DEST_BLOCKED


def test_run_all_fail_fast_does_not_corrupt_prior_writes():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["good"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["bad"]])

        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"
        wb.save(dest)

        cfg1 = _cfg(dest, columns="A", start_col="A")   # appends to row 3 (BLOCK at 2)
        cfg2 = _cfg(dest, columns="A", start_col="A", start_row="2")  # explicit collision

        report = run_all([(s1, "R1", cfg1), (s2, "R2", cfg2)])
        assert not report.ok
        # First write succeeded, second blocked
        assert report.results[0].rows_written == 1
        assert report.results[1].error_code == DEST_BLOCKED


def test_run_all_mixed_widths_landing_zone_awareness():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["v1", "v2", "v3"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["w1", "x", "w3"]])

        cfg1 = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,B,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="B", start_row=""),
        )
        report = run_all([(s1, "R1", cfg1), (s2, "R2", _pack_cfg(dest))])
        assert report.ok
        ws2 = load_workbook(dest)["Out"]
        assert ws2["B1"].value == "v1" and ws2["D1"].value == "v3"
        assert ws2["B2"].value == "w1"


def test_run_all_keep_then_pack_stacks_correctly():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"),
                          data=[["alpha","x",1],["beta","y",2],["gamma","z",3]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["delta","q",9]])

        cfg_keep = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="A", start_row=""),
        )
        cfg_pack = _pack_cfg(dest, columns="A,C", src_col="A")
        report = run_all([(s1, "R1", cfg_keep), (s2, "R2", cfg_pack)])
        assert report.ok


# ══════════════════════════════════════════════════════════════════════════════
# SIDE-BY-SIDE / LANDING ZONE (core.landing)
# ══════════════════════════════════════════════════════════════════════════════

def test_find_target_col_offsets_pack_all_data():
    shaped = [["a", "b", "c"], ["d", "e", "f"]]
    assert find_target_col_offsets(shaped) == [0, 1, 2]


def test_find_target_col_offsets_keep_with_gaps():
    shaped = [["a", None, "c"], [None, None, None], ["g", None, "i"]]
    result = find_target_col_offsets(shaped)
    assert 0 in result
    assert 2 in result
    assert 1 not in result


def test_find_target_col_offsets_all_none_returns_empty():
    shaped = [[None, None], [None, None]]
    assert find_target_col_offsets(shaped) == []


def test_read_zone_returns_correct_cell_map(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B2"] = 42
    ws["C3"] = None
    path = str(tmp_path / "t.xlsx")
    wb.save(path)

    wb2 = load_workbook(path, data_only=True)
    ws2 = wb2.active
    cell_map = read_zone(ws2, col_start=1, col_end=3)

    assert cell_map.get((1, 1)) == "hello"
    assert cell_map.get((2, 2)) == 42
    assert (3, 3) not in cell_map  # None not stored


def test_read_zone_does_not_inflate_max_row(tmp_path):
    """iter_rows must not register phantom cells."""
    wb = Workbook()
    ws = wb.active
    ws["A5"] = "real"
    path = str(tmp_path / "t.xlsx")
    wb.save(path)

    wb2 = load_workbook(path, data_only=True)
    ws2 = wb2.active
    before = ws2.max_row
    _ = read_zone(ws2, 1, 3)
    assert ws2.max_row == before  # must not change


def test_side_by_side_pack_AC_BD_same_dest():
    """Pack A+C → start_col=A, then Pack B+D → start_col=C land on same rows."""
    with TemporaryDirectory() as td:
        src1 = _make_xlsx(os.path.join(td, "s1.xlsx"),
                          data=[["ac1", "x", "ac2"], ["ac3", "x", "ac4"]])
        src2 = _make_xlsx(os.path.join(td, "s2.xlsx"),
                          data=[["bd1", "x", "bd2"], ["bd3", "x", "bd4"]])
        dest = os.path.join(td, "d.xlsx")

        cfg1 = _cfg(dest, columns="A,C", mode="pack", start_col="A")
        cfg2 = _cfg(dest, columns="A,C", mode="pack", start_col="C")

        r1 = run_sheet(src1, cfg1)
        r2 = run_sheet(src2, cfg2)
        assert r1.rows_written == 2
        assert r2.rows_written == 2

        ws2 = _ws(dest)
        assert ws2["A1"].value == "ac1"
        assert ws2["B1"].value == "ac2"
        assert ws2["C1"].value == "bd1"
        assert ws2["D1"].value == "bd2"


def test_side_by_side_keep_AC_BD_merge_same_rows():
    """Keep A,C → start A; Keep B,D → start B. Target cols don't overlap → same rows."""
    with TemporaryDirectory() as td:
        src1 = _make_xlsx(os.path.join(td, "s1.xlsx"),
                          data=[["a1", "x", "c1"], ["a2", "x", "c2"]])
        src2 = _make_xlsx(os.path.join(td, "s2.xlsx"),
                          data=[["b1", "x", "d1"], ["b2", "x", "d2"]])
        dest = os.path.join(td, "d.xlsx")

        cfg1 = _cfg(dest, columns="A,C", mode="keep", start_col="A")
        cfg2 = _cfg(dest, columns="A,C", mode="keep", start_col="B")

        r1 = run_sheet(src1, cfg1)
        r2 = run_sheet(src2, cfg2)
        assert r1.rows_written == 2
        assert r2.rows_written == 2

        ws2 = _ws(dest)
        assert ws2["A1"].value == "a1"
        assert ws2["C1"].value == "c1"
        assert ws2["B1"].value == "b1"
        assert ws2["D1"].value == "d1"


def test_merge_collision_on_target_col_raises_dest_blocked():
    """Collision on actual target columns must raise DEST_BLOCKED."""
    with TemporaryDirectory() as td:
        src = _make_xlsx(os.path.join(td, "s.xlsx"), data=[["x"]])
        dest = os.path.join(td, "d.xlsx")
        wb  = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)
        with pytest.raises(AppError) as ei:
            run_sheet(src, _cfg(dest, columns="A", start_col="A", start_row="1"))
        assert ei.value.code == DEST_BLOCKED


# ══════════════════════════════════════════════════════════════════════════════
# PACK / KEEP COMBINATION MATRIX
# ══════════════════════════════════════════════════════════════════════════════

def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


def test_pack_all_cols_no_rules():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["a", 1], ["b", 2], ["c", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        ws2 = _ws(dest)
        assert ws2["A1"].value == "a"


def test_pack_subset_columns_no_rules():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["a", "b", "c"], ["d", "e", "f"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,C"))
        assert result.rows_written == 2
        ws2 = _ws(dest)
        assert ws2["A1"].value == "a"
        assert ws2["B1"].value == "c"


def test_pack_include_equals_rule():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["keep", 1], ["drop", 2], ["keep", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="include", column="A", operator="equals", value="keep")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["A2"].value == "keep"


def test_pack_exclude_rule():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["alpha", 1], ["beta", 2], ["gamma", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="exclude", column="A", operator="equals", value="beta")
        ]))
        assert result.rows_written == 2
        names = [_ws(dest)[f"A{i}"].value for i in range(1, 3)]
        assert "beta" not in names


def test_pack_row_range_selection():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["r1"], ["r2"], ["r3"], ["r4"], ["r5"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rows="2-4"))
        assert result.rows_written == 3
        assert _ws(dest)["A1"].value == "r2"


def test_pack_explicit_start_row():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["only"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, start_row="999"))
        assert result.rows_written == 1
        assert _ws(dest)["A999"].value == "only"


def test_pack_include_all_rule_same_as_no_rules():
    with TemporaryDirectory() as td:
        data  = [["alpha", 1], ["beta", 2], ["gamma", 3]]
        src   = _xlsx(os.path.join(td, "s.xlsx"), data)
        dest1 = os.path.join(td, "d1.xlsx")
        dest2 = os.path.join(td, "d2.xlsx")

        r1 = run_sheet(src, _cfg(dest1, rules=[
            Rule(mode="include", column="A", operator="contains", value="")
        ]))
        r2 = run_sheet(src, _cfg(dest2))
        assert r1.rows_written == r2.rows_written == 3


def test_pack_exclude_all_rule_zero_rows():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["alpha", 1], ["beta", 2]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="exclude", column="A", operator="contains", value="")
        ]))
        assert result.rows_written == 0


def test_pack_contains_case_insensitive_subset_cols():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["ALPHA_val", 1], ["beta_val", 2], ["Alpha_val", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A", rules=[
            Rule(mode="include", column="A", operator="contains", value="alpha")
        ]))
        assert result.rows_written == 2


def test_pack_numeric_greater_than_col_in_output():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["a", 5], ["b", 15], ["c", 25]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,B", rules=[
            Rule(mode="include", column="B", operator=">", value="10")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["B1"].value == 15


def test_pack_numeric_less_than_col_not_in_output():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["a", 5, "x"], ["b", 15, "y"], ["c", 3, "z"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,C", rules=[
            Rule(mode="include", column="B", operator="<", value="10")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["A1"].value == "a"
        assert _ws(dest)["B1"].value == "x"


def test_pack_column_range_spec_with_rules_append():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["keep","x",10,"extra"],["drop","y",20,"extra"],
                      ["keep","z",30,"extra"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A-C", rules=[
            Rule(mode="include", column="A", operator="equals", value="keep")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["C1"].value == 10
        assert _ws(dest)["D1"].value is None


def test_keep_non_contiguous_cols_rows_gaps_preserved():
    with TemporaryDirectory() as td:
        data = [
            ["A1","B1","C1","D1","E1"],
            ["A2","B2","C2","D2","E2"],
            ["A3","B3","C3","D3","E3"],
            ["A4","B4","C4","D4","E4"],
            ["A5","B5","C5","D5","E5"],
        ]
        src  = _xlsx(os.path.join(td, "s.xlsx"), data)
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, columns="A,C,E", rows="1,3,5",
                                     mode="keep"))
        assert result.rows_written == 5
        ws2 = _ws(dest)
        assert ws2["A1"].value == "A1"
        assert ws2["B1"].value is None
        assert ws2["C1"].value == "C1"
        for col in ["A","B","C","D","E"]:
            assert ws2[f"{col}2"].value is None
        assert ws2["A3"].value == "A3"
        assert ws2["A5"].value == "A5"


def test_pack_then_keep_same_dest_stack():
    with TemporaryDirectory() as td:
        src1 = _xlsx(os.path.join(td, "s1.xlsx"),
                     [["pack1", 1], ["pack2", 2]])
        src2 = _xlsx(os.path.join(td, "s2.xlsx"),
                     [["keep1","x","A"],["keep2","y","B"],["keep3","z","C"]])
        dest = os.path.join(td, "d.xlsx")

        r1 = run_sheet(src1, _cfg(dest, columns="A,B", mode="pack"))
        r2 = run_sheet(src2, _cfg(dest, columns="A,C", rows="1,3", mode="keep"))
        assert r1.rows_written == 2
        assert r2.rows_written == 3

        ws2 = _ws(dest)
        assert ws2["A1"].value == "pack1"
        assert ws2["A3"].value == "keep1"
        assert ws2["C3"].value == "A"
        assert ws2["A5"].value == "keep3"
        assert ws2["C5"].value == "C"


def test_pack_multiple_and_rules_all_satisfied():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [
            ["alpha", 50, "YES"],
            ["alpha",  5, "YES"],
            ["beta",  50, "YES"],
            ["alpha", 60, "NO"],
            ["alpha", 55, "YES"],
        ])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, combine="AND", rules=[
            Rule(mode="include", column="A", operator="equals",  value="alpha"),
            Rule(mode="include", column="B", operator=">",       value="40"),
            Rule(mode="include", column="C", operator="equals",  value="YES"),
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["B1"].value == 50
        assert _ws(dest)["B2"].value == 55


def test_pack_mixed_include_exclude_or():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [
            ["alpha", 10],
            ["beta",  60],
            ["gamma", 20],
        ])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, combine="OR", rules=[
            Rule(mode="include", column="A", operator="equals", value="alpha"),
            Rule(mode="exclude", column="B", operator=">",      value="40"),
        ]))
        assert result.rows_written >= 1


# ══════════════════════════════════════════════════════════════════════════════
# ENGINE HARDENING — WIDE SHEETS, UNICODE, MIXED TYPES, CSV EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_wide_sheet_128_cols_all_cols_pack():
    with TemporaryDirectory() as td:
        src  = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        wb   = Workbook()
        ws   = wb.active; ws.title = "Sheet1"
        for c in range(1, 129):
            ws.cell(row=1, column=c, value=f"col{c}")
        wb.save(src)

        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 1
        out = _ws(dest)
        assert out.cell(row=1, column=128).value == "col128"


def test_wide_sheet_column_letter_roundtrip_boundaries():
    boundaries = [1, 26, 27, 52, 53, 702, 703, 16384]
    for n in boundaries:
        letters = col_index_to_letters(n)
        assert col_letters_to_index(letters) == n


def test_wide_sheet_zz_column_spec():
    with TemporaryDirectory() as td:
        src  = os.path.join(td, "s.xlsx")
        dest = os.path.join(td, "d.xlsx")
        wb   = Workbook()
        ws   = wb.active; ws.title = "Sheet1"
        ws.cell(row=1, column=702, value=9999)
        ws.cell(row=1, column=1,   value="first")
        wb.save(src)
        result = run_sheet(src, _cfg(dest, columns="A,ZZ"))
        assert result.rows_written == 1
        out = _ws(dest)
        assert out["A1"].value == "first"
        assert out["B1"].value == 9999


def test_unicode_values_pass_through_pack():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [
            ["café", "naïve", "résumé"],
            ["日本語", "中文", "한국어"],
            ["emoji🚀", "fire🔥", "✓check"],
        ])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 3
        out = _ws(dest)
        assert out["A1"].value == "café"
        assert out["A2"].value == "日本語"
        assert out["C3"].value == "✓check"


def test_unicode_rule_filter_equals():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["日本語", 1], ["中文", 2], ["日本語", 3]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="include", column="A", operator="equals", value="日本語")
        ]))
        assert result.rows_written == 2


def test_mixed_types_in_single_column_pack():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [
            [42, "label"], [3.14, "label"], ["text", "label"],
            [None, "label"], [True, "label"], [False, "label"],
        ])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 6
        out = _ws(dest)
        assert out["A1"].value == 42
        assert abs(out["A2"].value - 3.14) < 1e-9
        assert out["A5"].value is True
        assert out["A6"].value is False


def test_mixed_types_numeric_rule_skips_non_numeric():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [
            [100, "big"], ["text", "skip"], [None, "skip"], [200, "big"]
        ])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="include", column="A", operator=">", value="50")
        ]))
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == 100
        assert out["A2"].value == 200


def test_mixed_types_negative_numbers_in_rule():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [[-10, "a"], [-5, "b"], [0, "c"], [5, "d"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rules=[
            Rule(mode="include", column="A", operator="<", value="-3")
        ]))
        assert result.rows_written == 2
        assert _ws(dest)["A1"].value == -10


def test_very_long_string_cell_value_survives_roundtrip():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"),
                     [["x" * 10_000, "short"], ["normal", "val"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        assert len(_ws(dest)["A1"].value) == 10_000


def test_csv_quoted_fields_with_commas():
    with TemporaryDirectory() as td:
        src  = os.path.join(td, "s.csv")
        dest = os.path.join(td, "d.xlsx")
        _make_csv(src, [["Smith, John", "New York, NY", 100],
                        ["Doe, Jane",   "Austin, TX",  200]])
        result = run_sheet(src, _cfg(dest))
        assert result.rows_written == 2
        out = _ws(dest)
        assert out["A1"].value == "Smith, John"
        assert out["B1"].value == "New York, NY"


def test_dest_sheet_name_with_spaces():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["v", 1]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, dest_sheet="My Sheet Name"))
        assert result.rows_written == 1
        assert load_workbook(dest)["My Sheet Name"]["A1"].value == "v"


def test_rows_spec_beyond_used_range_ignored_gracefully():
    with TemporaryDirectory() as td:
        src  = _xlsx(os.path.join(td, "s.xlsx"), [["r1"], ["r2"]])
        dest = os.path.join(td, "d.xlsx")
        result = run_sheet(src, _cfg(dest, rows="1,3,5,100"))
        assert result.rows_written == 1   # only row 1 exists


@pytest.mark.slow
def test_200k_rows_performance():
    with TemporaryDirectory() as td:
        src  = os.path.join(td, "big.xlsx")
        dest = os.path.join(td, "out.xlsx")
        wb   = Workbook(write_only=True)
        ws   = wb.create_sheet("Sheet1")
        for i in range(200_000):
            ws.append([f"val_{i}", i, i * 0.1])
        wb.save(src)

        t0     = time.time()
        result = run_sheet(src, _cfg(dest))
        elapsed = time.time() - t0

        assert result.rows_written == 200_000
        assert elapsed < 120
