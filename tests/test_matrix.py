"""
test_matrix.py — Combinatorial matrix tests for Turbo Extractor V3.

Systematically exercises combinations that existing tests only partially cover:
  - Source type:   XLSX, CSV
  - Paste mode:    pack, keep
  - Column spec:   all, contiguous range, non-adjacent, single, wide gap
  - Row spec:      all, range, sparse list, single row
  - Rules:         none, include equals, include contains, exclude equals,
                   AND multi-rule, OR multi-rule, numeric < and >
  - Destination:   new file, pre-existing file, explicit start_row,
                   append mode, non-A start_col, mid-sheet anchor
  - Multi-source:  same dest same sheet, same dest diff sheets,
                   different dests, stacked append order
  - Collision:     explicit row blocked, non-overlapping columns safe,
                   target-col probe semantics (gap cols never block)
  - Edge cases:    empty source, all rows filtered, unicode values,
                   mixed numeric/string, None cells in source
"""
from __future__ import annotations

import csv
import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.batch import run_all
from core.errors import AppError, DEST_BLOCKED
from core.models import Destination, Rule, SheetConfig
from core.runner import run_sheet


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _xlsx(path, data, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


def _csv(path, data):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(data)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _ws(path, sheet="Out"):
    return load_workbook(path)[sheet]


def _col(ws, col_letter, max_row):
    """Return list of cell values from row 1..max_row in a given column."""
    return [ws[f"{col_letter}{r}"].value for r in range(1, max_row + 1)]


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — SOURCE TYPE × PASTE MODE (CSV vs XLSX, pack vs keep)
# ══════════════════════════════════════════════════════════════════════════════

class TestSourceTypePasteMode:

    def test_xlsx_pack_all_cols_all_rows(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["a", 1], ["b", 2], ["c", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 3
            ws = _ws(dest)
            assert ws["A1"].value == "a" and ws["B1"].value == 1
            assert ws["A3"].value == "c" and ws["B3"].value == 3

    def test_csv_pack_all_cols_all_rows(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"), [["x", "y"], ["1", "2"], ["3", "4"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 3
            ws = _ws(dest)
            assert ws["A1"].value == "x"
            assert ws["B3"].value == "4"

    def test_xlsx_keep_all_cols_all_rows(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["a", "b", "c"], ["d", "e", "f"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["C2"].value == "f"

    def test_csv_keep_all_cols_all_rows(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"), [["p", "q"], ["r", "s"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["B2"].value == "s"

    def test_xlsx_pack_non_adjacent_cols(self):
        """Pack: A and C selected → output col B gets C data, no gap."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["aa", "bb", "cc"], ["dd", "ee", "ff"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,C"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "aa"
            assert ws["B1"].value == "cc"   # no gap
            assert ws["C1"].value is None   # nothing in col C

    def test_xlsx_keep_non_adjacent_cols_preserves_gap(self):
        """Keep: A and C selected → output col B is None (gap preserved)."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["aa", "bb", "cc"], ["dd", "ee", "ff"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,C", mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "aa"
            assert ws["B1"].value is None   # gap preserved
            assert ws["C1"].value == "cc"

    def test_csv_pack_non_adjacent_cols(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["v1", "v2", "v3", "v4"], ["w1", "w2", "w3", "w4"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,D"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "v1"
            assert ws["B1"].value == "v4"   # D packed to col B
            assert ws["C1"].value is None

    def test_csv_keep_non_adjacent_wide_gap(self):
        """Keep with A and D: output width = 4, cols B and C are None."""
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["v1", "v2", "v3", "v4"], ["w1", "w2", "w3", "w4"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,D", mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "v1"
            assert ws["B1"].value is None   # gap
            assert ws["C1"].value is None   # gap
            assert ws["D1"].value == "v4"


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — ROW SELECTION × SOURCE TYPE
# ══════════════════════════════════════════════════════════════════════════════

class TestRowSelection:

    def test_xlsx_pack_row_range_middle(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["r1"], ["r2"], ["r3"], ["r4"], ["r5"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="2-4"))
            assert r.rows_written == 3
            ws = _ws(dest)
            assert ws["A1"].value == "r2"
            assert ws["A3"].value == "r4"
            assert ws["A4"].value is None

    def test_csv_pack_sparse_row_list(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["r1"], ["r2"], ["r3"], ["r4"], ["r5"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="1,3,5"))
            assert r.rows_written == 3
            ws = _ws(dest)
            assert ws["A1"].value == "r1"
            assert ws["A2"].value == "r3"
            assert ws["A3"].value == "r5"

    def test_xlsx_keep_row_range_compresses_rows(self):
        """Keep mode: selected rows 1 and 3 → output has 2 rows (no empty row gap)."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["A1", "B1"], ["A2", "B2"], ["A3", "B3"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="1,3", mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "A1"
            assert ws["A2"].value == "A3"  # row 3 follows immediately
            assert ws["A3"].value is None

    def test_xlsx_keep_non_adjacent_rows_and_cols_combo(self):
        """Keep mode: rows 1,3 + cols A,C → 2×3 output with col gap, no row gap."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["a", "b", "c"],
                          ["d", "e", "f"],
                          ["g", "h", "i"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="1,3", columns="A,C", mode="keep"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "a"
            assert ws["B1"].value is None  # column gap
            assert ws["C1"].value == "c"
            assert ws["A2"].value == "g"   # row 3 immediately follows
            assert ws["C2"].value == "i"

    def test_csv_pack_single_row(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["only"], ["skip"], ["skip"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="1"))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "only"
            assert ws["A2"].value is None


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — RULES MATRIX
# ══════════════════════════════════════════════════════════════════════════════

class TestRulesMatrix:

    # ── include / exclude basics ──────────────────────────────────────────────

    def test_include_equals_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["keep", 1], ["drop", 2], ["keep", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="A", operator="equals", value="keep")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "keep"
            assert ws["A2"].value == "keep"

    def test_include_equals_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["yes", "10"], ["no", "20"], ["yes", "30"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="A", operator="equals", value="yes")
            ]))
            assert r.rows_written == 2

    def test_exclude_equals_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["alpha", 1], ["beta", 2], ["gamma", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="exclude", column="A", operator="equals", value="beta")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            vals = [ws[f"A{i}"].value for i in range(1, 3)]
            assert "beta" not in vals

    def test_include_contains_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["apple", 1], ["banana", 2], ["apricot", 3], ["cherry", 4]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="A", operator="contains", value="ap")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            vals = [ws[f"A{i}"].value for i in range(1, 3)]
            assert "apple" in vals
            assert "apricot" in vals

    def test_include_contains_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["foo_bar"], ["baz"], ["foo_qux"], ["quux"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="A", operator="contains", value="foo")
            ]))
            assert r.rows_written == 2

    def test_numeric_greater_than_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["a", 5], ["b", 15], ["c", 25], ["d", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="B", operator=">", value="10")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            vals = [ws[f"A{i}"].value for i in range(1, 3)]
            assert "b" in vals and "c" in vals

    def test_numeric_less_than_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["x", "5"], ["y", "15"], ["z", "3"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="B", operator="<", value="10")
            ]))
            assert r.rows_written == 2

    # ── AND / OR combinator ───────────────────────────────────────────────────

    def test_and_two_include_rules_both_must_match(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["keep", "high", 50],
                          ["keep", "low",   5],
                          ["drop", "high", 50]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, combine="AND", rules=[
                Rule(mode="include", column="A", operator="equals",  value="keep"),
                Rule(mode="include", column="B", operator="equals",  value="high"),
            ]))
            assert r.rows_written == 1
            assert _ws(dest)["A1"].value == "keep"

    def test_or_two_include_rules_either_matches(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["alpha", 1], ["beta", 2], ["gamma", 3]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, combine="OR", rules=[
                Rule(mode="include", column="A", operator="equals", value="alpha"),
                Rule(mode="include", column="A", operator="equals", value="gamma"),
            ]))
            assert r.rows_written == 2

    def test_and_include_plus_exclude(self):
        """AND: include col A equals 'keep' AND exclude col B equals 'bad'."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["keep", "good"], ["keep", "bad"], ["drop", "good"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, combine="AND", rules=[
                Rule(mode="include", column="A", operator="equals", value="keep"),
                Rule(mode="exclude", column="B", operator="equals", value="bad"),
            ]))
            assert r.rows_written == 1
            assert _ws(dest)["B1"].value == "good"

    def test_or_include_plus_exclude_semantics(self):
        """OR: keep row if include matches OR exclude does not match."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["yes", "x"], ["no", "y"], ["no", "z"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, combine="OR", rules=[
                Rule(mode="include", column="A", operator="equals", value="yes"),
                Rule(mode="exclude", column="B", operator="equals", value="x"),
            ]))
            assert r.rows_written == 3

    def test_all_rows_filtered_produces_zero_rows(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["alpha", 1], ["beta", 2]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rules=[
                Rule(mode="include", column="A", operator="equals", value="NONE")
            ]))
            assert r.rows_written == 0

    def test_rules_use_absolute_source_columns_not_selected_cols(self):
        """Rule on col B must see original col B even when col A is excluded."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["drop_me", "keep", 1],
                          ["drop_me", "skip", 2]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="B,C", rules=[
                Rule(mode="include", column="B", operator="equals", value="keep")
            ]))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "keep"   # B mapped to output col A in pack mode

    def test_rules_with_keep_mode_csv(self):
        """Rules + keep mode on CSV: filtered rows don't appear, col gaps preserved."""
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["yes", "x", "1"],
                         ["no",  "y", "2"],
                         ["yes", "z", "3"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, columns="A,C", mode="keep", rules=[
                Rule(mode="include", column="A", operator="equals", value="yes")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "yes"
            assert ws["B1"].value is None    # col gap (B not selected)
            assert ws["C1"].value == "1"
            assert ws["A2"].value == "yes"
            assert ws["C2"].value == "3"


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — DESTINATION CONFIG: START ROW × START COL × APPEND
# ══════════════════════════════════════════════════════════════════════════════

class TestDestinationConfig:

    def test_explicit_start_row_1(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["val"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_row="1"))
            assert r.rows_written == 1
            assert _ws(dest)["A1"].value == "val"

    def test_explicit_start_row_mid_sheet(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["mid"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_row="10"))
            assert r.rows_written == 1
            assert _ws(dest)["A10"].value == "mid"
            assert _ws(dest)["A9"].value is None

    def test_explicit_start_col_b(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["c1", "c2"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_col="B"))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value is None
            assert ws["B1"].value == "c1"
            assert ws["C1"].value == "c2"

    def test_explicit_start_col_e(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["x", "y", "z"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_col="E"))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["E1"].value == "x"
            assert ws["F1"].value == "y"
            assert ws["G1"].value == "z"
            assert ws["D1"].value is None

    def test_explicit_start_col_and_row_combo(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["p", "q"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_col="C", start_row="5"))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["C5"].value == "p"
            assert ws["D5"].value == "q"
            assert ws["C4"].value is None

    def test_append_to_empty_dest_lands_row_1(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["first"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, start_row=""))
            assert r.rows_written == 1
            assert _ws(dest)["A1"].value == "first"

    def test_append_stacks_below_existing_data(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["second"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A1"] = "existing"
            wb.save(dest)
            r = run_sheet(src, _cfg(dest, start_row=""))
            assert r.rows_written == 1
            ws2 = _ws(dest)
            assert ws2["A1"].value == "existing"
            assert ws2["A2"].value == "second"

    def test_append_with_full_landing_zone_scans_past_all_blockers(self):
        """
        Append mode absorbs all occupied cells in the landing zone via the scan.
        The scan finds max_used_row then places at max+1, which is clear.
        This verifies the 'upside-down Tetris' behavior: no DEST_BLOCKED in
        pure append mode.
        """
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["new"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A1"] = "r1"; ws["A2"] = "r2"; ws["A3"] = "r3"
            wb.save(dest)
            r = run_sheet(src, _cfg(dest, start_row=""))
            assert r.rows_written == 1
            assert _ws(dest)["A4"].value == "new"   # placed at max+1=4

    def test_append_respects_landing_zone_columns(self):
        """Append scans only landing-zone cols; data in unrelated cols is ignored."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["new"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A5"] = "noise"   # col A has data at row 5
            ws["B1"] = "other"   # col B has data at row 1
            wb.save(dest)
            # Writing to col C — should land at row 1 (col C is empty)
            r = run_sheet(src, _cfg(dest, start_col="C", start_row=""))
            assert r.rows_written == 1
            ws2 = _ws(dest)
            assert ws2["C1"].value == "new"

    def test_append_non_a_start_col_stacks_correctly(self):
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["batch1"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["batch2"]])
            report = run_all([
                (s1, "R1", _cfg(dest, start_col="D", start_row="")),
                (s2, "R2", _cfg(dest, start_col="D", start_row="")),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["D1"].value == "batch1"
            assert ws["D2"].value == "batch2"


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — MULTI-SOURCE BATCH SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════

class TestMultiSourceBatch:

    def test_same_dest_same_sheet_stack_order(self):
        """Three sources → same dest/sheet: rows written in source order."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            srcs = [
                _xlsx(os.path.join(td, f"s{i}.xlsx"), [[f"row{i}"]]) for i in range(1, 4)
            ]
            items = [(s, f"R{i+1}", _cfg(dest)) for i, s in enumerate(srcs)]
            report = run_all(items)
            assert report.ok
            ws = _ws(dest)
            for i in range(1, 4):
                assert ws[f"A{i}"].value == f"row{i}"

    def test_same_dest_different_sheets(self):
        """Two sources writing to different sheets in the same dest file."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["sheet_a_data"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["sheet_b_data"]])
            report = run_all([
                (s1, "R1", _cfg(dest, dest_sheet="SheetA")),
                (s2, "R2", _cfg(dest, dest_sheet="SheetB")),
            ])
            assert report.ok
            wb = load_workbook(dest)
            assert wb["SheetA"]["A1"].value == "sheet_a_data"
            assert wb["SheetB"]["A1"].value == "sheet_b_data"

    def test_different_dests(self):
        """Two sources, two separate destination files."""
        with TemporaryDirectory() as td:
            d1 = os.path.join(td, "d1.xlsx")
            d2 = os.path.join(td, "d2.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["dest1_val"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["dest2_val"]])
            report = run_all([
                (s1, "R1", _cfg(d1)),
                (s2, "R2", _cfg(d2)),
            ])
            assert report.ok
            assert _ws(d1)["A1"].value == "dest1_val"
            assert _ws(d2)["A1"].value == "dest2_val"

    def test_mixed_source_types_same_dest(self):
        """XLSX and CSV sources both appending to the same destination."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            sx = _xlsx(os.path.join(td, "s.xlsx"), [["from_xlsx"]])
            sc = _csv(os.path.join(td, "s.csv"), [["from_csv"]])
            report = run_all([
                (sx, "R1", _cfg(dest)),
                (sc, "R2", _cfg(dest)),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "from_xlsx"
            assert ws["A2"].value == "from_csv"

    def test_mixed_paste_modes_same_dest(self):
        """Pack then keep, stacking to same dest."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["a", "b", "c"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["x", "y", "z"]])
            report = run_all([
                (s1, "R1", _cfg(dest, mode="pack")),
                (s2, "R2", _cfg(dest, mode="keep")),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "a"
            assert ws["A2"].value == "x"

    def test_five_sources_same_dest_correct_row_count(self):
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            items = []
            for i in range(1, 6):
                src = _xlsx(os.path.join(td, f"s{i}.xlsx"), [[f"v{i}"]])
                items.append((src, f"R{i}", _cfg(dest)))
            report = run_all(items)
            assert report.ok
            ws = _ws(dest)
            for i in range(1, 6):
                assert ws[f"A{i}"].value == f"v{i}"

    def test_same_dest_with_rules_each_source(self):
        """Each source has a different filter rule; results stack correctly."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"),
                       [["yes", 1], ["no", 2], ["yes", 3]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"),
                       [["keep", 10], ["drop", 20]])
            report = run_all([
                (s1, "R1", _cfg(dest, rules=[
                    Rule(mode="include", column="A", operator="equals", value="yes")
                ])),
                (s2, "R2", _cfg(dest, rules=[
                    Rule(mode="include", column="A", operator="equals", value="keep")
                ])),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "yes"
            assert ws["A2"].value == "yes"
            assert ws["A3"].value == "keep"

    def test_multi_source_different_start_cols_no_collision(self):
        """Two sources write to non-overlapping columns — both succeed."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["left"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["right"]])
            report = run_all([
                (s1, "R1", _cfg(dest, start_col="A")),
                (s2, "R2", _cfg(dest, start_col="E")),
            ])
            assert report.ok
            ws = _ws(dest)
            assert ws["A1"].value == "left"
            assert ws["E1"].value == "right"


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — COLLISION MATRIX
# ══════════════════════════════════════════════════════════════════════════════

class TestCollisionMatrix:

    def test_explicit_row_blocked_by_existing_data(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["new"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A5"] = "BLOCKER"
            wb.save(dest)
            with pytest.raises(AppError) as ei:
                run_sheet(src, _cfg(dest, start_row="5"))
            assert ei.value.code == DEST_BLOCKED

    def test_multi_col_write_partial_overlap_blocked(self):
        """Source has 3 cols; col B is blocked at target row → DEST_BLOCKED."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["x", "y", "z"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["B1"] = "BLOCK"
            wb.save(dest)
            with pytest.raises(AppError) as ei:
                run_sheet(src, _cfg(dest, start_row="1", start_col="A"))
            assert ei.value.code == DEST_BLOCKED

    def test_non_overlapping_start_col_safe_after_existing_data(self):
        """Writing to col D when existing data is only in cols A–C: no collision."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["safe"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A1"] = "x"; ws["B1"] = "y"; ws["C1"] = "z"
            wb.save(dest)
            r = run_sheet(src, _cfg(dest, start_col="D", start_row="1"))
            assert r.rows_written == 1
            assert _ws(dest)["D1"].value == "safe"

    def test_batch_fail_fast_stops_after_first_collision(self):
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A1"] = "BLOCK"
            wb.save(dest)
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["bad"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["good"]])
            report = run_all([
                (s1, "R1", _cfg(dest, start_row="1")),
                (s2, "R2", _cfg(dest)),
            ])
            assert not report.ok
            assert len(report.results) == 1
            assert report.results[0].error_code == DEST_BLOCKED

    def test_keep_mode_gap_col_blocker_does_not_block(self):
        """
        Keep mode produces gap columns (all-None). The planner probes target
        columns only — a blocker in a gap column is intentionally ignored.
        This verifies the 'target-cols-only' probe design.
        Source selects cols A and C (keep mode) → bounding box is A-C, col B is a gap.
        A blocker at B1 must NOT raise DEST_BLOCKED.
        """
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["aa", "bb", "cc"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["B1"] = "existing_in_gap"   # gap column — ignored by probe
            wb.save(dest)
            r = run_sheet(src, _cfg(dest, columns="A,C", mode="keep",
                                    start_row="1", start_col="A"))
            assert r.rows_written == 1
            ws2 = _ws(dest)
            assert ws2["A1"].value == "aa"
            assert ws2["C1"].value == "cc"

    def test_keep_mode_data_col_blocker_raises_dest_blocked(self):
        """
        Keep mode: a blocker in an actual data column (not a gap) raises DEST_BLOCKED.
        Source selects cols A and C → data cols are A and C.
        A blocker at C1 must raise DEST_BLOCKED.
        """
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["aa", "bb", "cc"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["C1"] = "DATA_COL_BLOCKER"   # actual data column — must block
            wb.save(dest)
            with pytest.raises(AppError) as ei:
                run_sheet(src, _cfg(dest, columns="A,C", mode="keep",
                                    start_row="1", start_col="A"))
            assert ei.value.code == DEST_BLOCKED

    def test_collision_error_includes_code_in_apperror(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["v"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "Out"
            ws["A3"] = "BLOCK"
            wb.save(dest)
            try:
                run_sheet(src, _cfg(dest, start_row="3"))
                assert False, "Expected AppError"
            except AppError as e:
                assert e.code == DEST_BLOCKED
                assert isinstance(e.details, dict)


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — EDGE CASES & DATA INTEGRITY
# ══════════════════════════════════════════════════════════════════════════════

class TestEdgeCasesAndDataIntegrity:

    def test_empty_xlsx_source_zero_rows(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 0

    def test_empty_csv_source_zero_rows(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"), [])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 0

    def test_unicode_values_preserved_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["こんにちは", "мир", "🎉"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "こんにちは"
            assert ws["B1"].value == "мир"
            assert ws["C1"].value == "🎉"

    def test_unicode_values_preserved_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"), [["αβγ", "δεζ"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "αβγ"

    def test_mixed_numeric_string_none_preserved(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [[1, "text", None, 3.14, True]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == 1
            assert ws["B1"].value == "text"
            assert ws["D1"].value == 3.14

    def test_zero_numeric_value_written_not_treated_as_empty(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [[0, 0.0, "0"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == 0
            assert ws["B1"].value == 0.0

    def test_single_cell_source_xlsx(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["solo"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            assert _ws(dest)["A1"].value == "solo"

    def test_single_cell_source_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"), [["csv_solo"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            assert _ws(dest)["A1"].value == "csv_solo"

    def test_wide_source_100_cols_pack(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [[f"col{i}" for i in range(100)]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest))
            assert r.rows_written == 1
            ws = _ws(dest)
            assert ws["A1"].value == "col0"
            assert ws.cell(row=1, column=100).value == "col99"

    def test_dest_sheet_created_when_missing_from_existing_workbook(self):
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"), [["v"]])
            dest = os.path.join(td, "d.xlsx")
            wb = Workbook(); wb.active.title = "Existing"; wb.save(dest)
            r = run_sheet(src, _cfg(dest, dest_sheet="NewSheet"))
            assert r.rows_written == 1
            assert load_workbook(dest)["NewSheet"]["A1"].value == "v"

    def test_source_start_row_skips_header(self):
        """source_start_row=2 skips row 1 (header); data starts from row 2."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["header"], ["data1"], ["data2"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, src_start_row="2"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "data1"
            assert ws["A2"].value == "data2"

    def test_source_start_row_skips_header_csv(self):
        with TemporaryDirectory() as td:
            src  = _csv(os.path.join(td, "s.csv"),
                        [["ID", "Name"], ["1", "Alice"], ["2", "Bob"]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, src_start_row="2"))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "1"
            assert ws["B2"].value == "Bob"

    def test_rows_and_cols_spec_combined_with_rules_xlsx(self):
        """rows=1-3, cols=A,C, include rule on B: pipeline order is correct."""
        with TemporaryDirectory() as td:
            src  = _xlsx(os.path.join(td, "s.xlsx"),
                         [["keep", "x", 10],
                          ["drop", "y", 20],
                          ["keep", "z", 30],
                          ["keep", "w", 40]])
            dest = os.path.join(td, "d.xlsx")
            r = run_sheet(src, _cfg(dest, rows="1-3", columns="A,C", rules=[
                Rule(mode="include", column="A", operator="equals", value="keep")
            ]))
            assert r.rows_written == 2
            ws = _ws(dest)
            assert ws["A1"].value == "keep"
            assert ws["B1"].value == 10      # col C packed to output col B
            assert ws["A2"].value == "keep"
            assert ws["B2"].value == 30

    def test_multiple_appends_same_dest_then_collision_on_explicit_row(self):
        """After two successful appends (rows 1,2), explicit start_row=1 → DEST_BLOCKED."""
        with TemporaryDirectory() as td:
            dest = os.path.join(td, "d.xlsx")
            s1 = _xlsx(os.path.join(td, "s1.xlsx"), [["first"]])
            s2 = _xlsx(os.path.join(td, "s2.xlsx"), [["second"]])
            run_all([
                (s1, "R1", _cfg(dest)),
                (s2, "R2", _cfg(dest)),
            ])
            s3 = _xlsx(os.path.join(td, "s3.xlsx"), [["collide"]])
            with pytest.raises(AppError) as ei:
                run_sheet(s3, _cfg(dest, start_row="1"))
            assert ei.value.code == DEST_BLOCKED
