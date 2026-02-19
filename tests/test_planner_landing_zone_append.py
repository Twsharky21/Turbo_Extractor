\
import pytest
from openpyxl import Workbook

from core.planner import build_plan
from core.errors import AppError, DEST_BLOCKED


def make_ws_with_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb, ws


def test_append_uses_max_used_row_across_landing_cols():
    wb, ws = make_ws_with_values()

    # Landing zone will be D:E (start_col=D, width=2)
    ws["D5"] = "x"
    ws["E10"] = "y"
    # This must NOT affect landing-zone aware append:
    ws["A100"] = "zzz"

    shaped = [["a", "b"], ["c", "d"]]  # height=2, width=2
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 11  # max used row across D/E is 10 -> append at 11
    assert plan.start_col == 4   # D
    assert plan.landing_cols == (4, 5)
    assert plan.landing_rows == (11, 12)


def test_append_collision_probe_blocks_if_any_cell_in_rectangle_occupied():
    wb, ws = make_ws_with_values()

    # D/E landing zone, width=2; max used row in D/E is 3 -> append_row 4
    ws["D3"] = "x"
    # Put a blocker inside the rectangle at the computed append row
    ws["E4"] = "BLOCK"

    shaped = [["a", "b"], ["c", "d"]]  # would write into D4:E5

    with pytest.raises(AppError) as ei:
        build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details is not None
    assert ei.value.details["first_blocker"]["col_letter"] == "E"
    assert ei.value.details["first_blocker"]["row"] == 4


def test_explicit_start_row_no_append_scan_but_still_probes_collision():
    wb, ws = make_ws_with_values()

    ws["D50"] = "BLOCK"
    shaped = [["a", "b"]]

    with pytest.raises(AppError) as ei:
        build_plan(ws, shaped, start_col_letters="D", start_row_str="50")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["target_start"] == "D50"
