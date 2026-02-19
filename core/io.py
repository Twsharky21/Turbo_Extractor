\
from __future__ import annotations

from typing import List, Any
import csv
from openpyxl import load_workbook


def is_occupied(value: Any) -> bool:
    """
    Single occupancy definition for destination + used-range detection.
    """
    if value is None:
        return False
    if isinstance(value, str) and value == "":
        return False
    return True


def normalize_table(rows: List[List[Any]]) -> List[List[Any]]:
    """
    Pad ragged rows to used width.
    """
    if not rows:
        return []

    used_width = max(len(r) for r in rows)
    return [r + [None] * (used_width - len(r)) for r in rows]


def compute_used_range(rows: List[List[Any]]) -> tuple[int, int]:
    """
    Returns (used_height, used_width).
    """
    if not rows:
        return 0, 0

    used_height = 0
    used_width = 0

    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            if is_occupied(value):
                used_height = max(used_height, r_idx + 1)
                used_width = max(used_width, c_idx + 1)

    return used_height, used_width


def load_csv(path: str) -> List[List[Any]]:
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = [list(row) for row in reader]

    rows = normalize_table(rows)
    return rows


def load_xlsx(path: str, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    rows = normalize_table(rows)
    return rows
