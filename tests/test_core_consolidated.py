# Consolidated from: test_io_basic.py, test_parsing.py, test_rules_basic.py, test_transform_basic.py, test_planner_landing_zone_append.py, test_writer_basic.py, test_errors.py
# Generated: 2026-02-19 20:40 UTC
# NOTE: Function renames applied only to avoid name collisions across original test modules.



# ---- BEGIN test_io_basic.py ----

\
from core.io import is_occupied, normalize_table, compute_used_range


def test_is_occupied():
    assert not is_occupied(None)
    assert not is_occupied("")
    assert is_occupied(" ")
    assert is_occupied(0)
    assert is_occupied("text")


def test_normalize_table():
    rows = [[1, 2], [3]]
    norm = normalize_table(rows)
    assert len(norm[1]) == 2
    assert norm[1][1] is None


def test_compute_used_range_basic():
    rows = [
        [None, None],
        [None, 5],
        [None, None],
    ]
    h, w = compute_used_range(rows)
    assert h == 2
    assert w == 2


# ---- END {f} ----



# ---- BEGIN test_parsing.py ----

\
import pytest

from core.errors import AppError, BAD_SPEC
from core.parsing import (
    col_letters_to_index,
    col_index_to_letters,
    parse_columns,
    parse_rows,
)


def test_col_letters_to_index_basic():
    assert col_letters_to_index("A") == 1
    assert col_letters_to_index("Z") == 26
    assert col_letters_to_index("AA") == 27
    assert col_letters_to_index("AZ") == 52
    assert col_letters_to_index("BA") == 53


def test_col_index_to_letters_basic():
    assert col_index_to_letters(1) == "A"
    assert col_index_to_letters(26) == "Z"
    assert col_index_to_letters(27) == "AA"
    assert col_index_to_letters(52) == "AZ"
    assert col_index_to_letters(53) == "BA"


def test_col_roundtrip_property_small():
    for n in range(1, 200):
        assert col_letters_to_index(col_index_to_letters(n)) == n


def test_parse_columns_blank_means_all_by_caller():
    assert parse_columns("") == []
    assert parse_columns("   ") == []


def test_parse_columns_singletons_and_ranges_unique_sorted():
    assert parse_columns("A,C") == [0, 2]
    assert parse_columns("C,A") == [0, 2]
    assert parse_columns("A-C") == [0, 1, 2]
    assert parse_columns("A-C,C,A") == [0, 1, 2]


def test_parse_columns_whitespace_and_trailing_commas():
    assert parse_columns(" A, C, ") == [0, 2]
    cols = parse_columns("A, C, AC-AE,")
    assert cols[0] == 0
    assert 2 in cols


def test_parse_columns_reverse_range_normalizes():
    assert parse_columns("D-B") == [1, 2, 3]


def test_parse_columns_rejects_bad_tokens():
    with pytest.raises(AppError) as ei:
        parse_columns("A,??,C")
    assert ei.value.code == BAD_SPEC


def test_parse_rows_blank_means_all_by_caller():
    assert parse_rows("") == []
    assert parse_rows("   ") == []


def test_parse_rows_singletons_and_ranges_unique_sorted():
    assert parse_rows("1,3") == [0, 2]
    assert parse_rows("3,1") == [0, 2]
    assert parse_rows("1-3") == [0, 1, 2]
    assert parse_rows("1-3,3-5,10,12-13") == [0, 1, 2, 3, 4, 9, 11, 12]


def test_parse_rows_whitespace_and_trailing_commas():
    assert parse_rows(" 1 , 2-3, ") == [0, 1, 2]


def test_parse_rows_reverse_range_normalizes():
    assert parse_rows("6-4") == [3, 4, 5]


def test_parse_rows_rejects_bad_tokens():
    with pytest.raises(AppError) as ei:
        parse_rows("1,2,nope")
    assert ei.value.code == BAD_SPEC


def test_parse_rows_rejects_nonpositive():
    with pytest.raises(AppError):
        parse_rows("0")
    with pytest.raises(AppError):
        parse_rows("-1")
    with pytest.raises(AppError):
        parse_rows("1-0")


# ---- END {f} ----



# ---- BEGIN test_rules_basic.py ----

\
from core.rules import apply_rules
from core.models import Rule


def rules_sample_rows():
    return [
        ["alpha", "10"],
        ["beta", "20"],
        ["gamma", "30"],
        ["beta", "5"],
        ["", None],
    ]


def test_include_equals():
    rows = rules_sample_rows()
    rules = [Rule(mode="include", column="A", operator="equals", value="beta")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 2


def test_exclude_equals():
    rows = rules_sample_rows()
    rules = [Rule(mode="exclude", column="A", operator="equals", value="beta")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 3


def test_contains_case_insensitive():
    rows = rules_sample_rows()
    rules = [Rule(mode="include", column="A", operator="contains", value="ALP")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 1


def test_numeric_greater_than():
    rows = rules_sample_rows()
    rules = [Rule(mode="include", column="B", operator=">", value="15")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 2


def test_numeric_less_than_safe_on_non_numeric():
    rows = rules_sample_rows()
    rules = [Rule(mode="include", column="B", operator="<", value="15")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 2


def test_and_combine():
    rows = rules_sample_rows()
    rules = [
        Rule(mode="include", column="A", operator="equals", value="beta"),
        Rule(mode="include", column="B", operator=">", value="10"),
    ]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 1


def test_or_combine():
    rows = rules_sample_rows()
    rules = [
        Rule(mode="include", column="A", operator="equals", value="alpha"),
        Rule(mode="include", column="B", operator=">", value="25"),
    ]
    result = apply_rules(rows, rules, "OR")
    assert len(result) == 2


# ---- END {f} ----



# ---- BEGIN test_transform_basic.py ----

\
from core.transform import (
    apply_row_selection,
    apply_column_selection,
    shape_pack,
    shape_keep,
)


def transform_sample_rows():
    return [
        ["A1", "B1", "C1", "D1"],
        ["A2", "B2", "C2", "D2"],
        ["A3", "B3", "C3", "D3"],
        ["A4", "B4", "C4", "D4"],
    ]


def test_row_selection_basic():
    rows = transform_sample_rows()
    result = apply_row_selection(rows, [0, 2])
    assert len(result) == 2
    assert result[1][0] == "A3"


def test_column_selection_basic():
    rows = transform_sample_rows()
    result = apply_column_selection(rows, [0, 2])
    assert result[0] == ["A1", "C1"]


def test_pack_mode_identity():
    rows = transform_sample_rows()
    packed = shape_pack(rows)
    assert packed == rows


def test_keep_mode_preserves_spacing():
    rows = transform_sample_rows()
    selected_rows = [0, 2]
    selected_cols = [0, 2]

    shaped = shape_keep(rows, selected_rows, selected_cols)

    # bounding box: rows 0-2, cols 0-2
    assert len(shaped) == 3
    assert len(shaped[0]) == 3

    # row 1 should be blank (not selected)
    assert shaped[1] == [None, None, None]

    # selected positions preserved
    assert shaped[0][0] == "A1"
    assert shaped[0][2] == "C1"
    assert shaped[2][0] == "A3"
    assert shaped[2][2] == "C3"


# ---- END {f} ----



# ---- BEGIN test_planner_landing_zone_append.py ----

\
import pytest
from openpyxl import Workbook

from core.planner import build_plan
from core.errors import AppError, DEST_BLOCKED


def make_ws_with_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb, ws


def test_append_uses_max_used_row_across_landing_cols():
    wb, ws = make_ws_with_values()

    # Landing zone will be D:E (start_col=D, width=2)
    ws["D5"] = "x"
    ws["E10"] = "y"
    # This must NOT affect landing-zone aware append:
    ws["A100"] = "zzz"

    shaped = [["a", "b"], ["c", "d"]]  # height=2, width=2
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")

    assert plan is not None
    assert plan.start_row == 11  # max used row across D/E is 10 -> append at 11
    assert plan.start_col == 4   # D
    assert plan.landing_cols == (4, 5)
    assert plan.landing_rows == (11, 12)


def test_append_skips_past_any_used_cells_in_landing_zone():
    wb, ws = make_ws_with_values()

    # D/E landing zone, width=2
    ws["D3"] = "x"
    ws["E4"] = "BLOCK"

    shaped = [["a", "b"], ["c", "d"]]  # height=2, width=2

    # Full landing-zone awareness means append after the max used row across D/E (which is 4)
    plan = build_plan(ws, shaped, start_col_letters="D", start_row_str="")
    assert plan is not None
    assert plan.start_row == 5  # max used row across D/E is 4 -> append at 5


def test_explicit_start_row_collision_probe_blocks():
    wb, ws = make_ws_with_values()

    ws["D50"] = "BLOCK"
    shaped = [["a", "b"]]

    with pytest.raises(AppError) as ei:
        build_plan(ws, shaped, start_col_letters="D", start_row_str="50")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["target_start"] == "D50"


# ---- END {f} ----



# ---- BEGIN test_writer_basic.py ----

\
from openpyxl import Workbook

from core.planner import build_plan
from core.writer import apply_write_plan


def test_writer_writes_exact_rectangle():
    wb = Workbook()
    ws = wb.active

    shaped = [
        ["a", "b"],
        ["c", "d"],
    ]

    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    rows_written = apply_write_plan(ws, shaped, plan)

    assert rows_written == 2
    assert ws["C1"].value == "a"
    assert ws["D1"].value == "b"
    assert ws["C2"].value == "c"
    assert ws["D2"].value == "d"


def test_writer_appends_after_existing_data():
    wb = Workbook()
    ws = wb.active

    ws["C1"] = "existing"
    ws["D3"] = "also existing"

    shaped = [
        ["x", "y"],
    ]

    plan = build_plan(ws, shaped, start_col_letters="C", start_row_str="")
    apply_write_plan(ws, shaped, plan)

    # max used row across C:D is 3 -> append at row 4
    assert ws["C4"].value == "x"
    assert ws["D4"].value == "y"


# ---- END {f} ----



# ---- BEGIN test_errors.py ----

\
from core.errors import AppError


def test_app_error_str_includes_code_message():
    e = AppError("X", "Nope")
    assert str(e).startswith("X: Nope")


def test_app_error_str_includes_details_when_present():
    e = AppError("X", "Nope", {"a": 1})
    s = str(e)
    assert "X: Nope" in s
    assert "a" in s


# ---- END {f} ----
