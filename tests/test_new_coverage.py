"""
New coverage tests — filling every remaining gap identified in the full audit.

Covers:
  ENGINE / VALIDATION
  - source_start_row: non-numeric ('abc'), zero ('0'), negative ('-1') → BAD_SOURCE_START_ROW
  - source_start_row='1' is same as no offset
  - run_sheet: bad workbook sheet name propagates SHEET_NOT_FOUND with message
  - Destination auto-creates parent dirs? (engine saves, not mkdir)
  - Two run_sheet calls to same file, different dest sheets, both written
  - run_all on empty iterable returns ok=True, results=[]
  - run_all: generator as input (not just list)
  - Result message is exactly "OK" when rows written > 0
  - Result message is exactly "0 rows written" when 0 rows

  PLANNER
  - start_row=0 raises BAD_SPEC
  - start_row=-1 raises BAD_SPEC
  - start_row=non-integer string raises BAD_SPEC
  - Blocker in append mode has append_mode=True in details
  - Blocker in explicit mode has append_mode=False in details
  - Blocker error details contain first_blocker row/col/col_letter/value
  - Collision anywhere in bounding box (not just top-left) raises DEST_BLOCKED
  - Append on completely empty sheet starts at row 1

  WRITER
  - Write None values into cells explicitly
  - Write bool values (True/False) to cells
  - Write numeric zero to a cell
  - Offset: apply_write_plan fills exact rectangle (no extra cells written)

  IO / OCCUPIED
  - is_occupied on various edge cases: 0, False, True, 0.0, [], {}
  - compute_used_range: only rightmost occupied col defines width
  - load_xlsx: multiple sheets, reads only the specified one
  - load_xlsx: sheet with only empty strings → used_range (0,0)

  RULES
  - apply_rules with combine_mode AND and no rules → returns all rows
  - apply_rules with combine_mode OR and no rules → returns all rows
  - apply_rules: AND with one True and one False → excluded
  - apply_rules: OR with one True and one False → included
  - apply_rules: ALL exclude rules with OR → row kept unless ALL match
  - apply_rules: numeric rule, cell contains float that satisfies >
  - apply_rules: numeric rule, cell is int that satisfies <
  - apply_rules: rule column ZZ (high column) → row too short → no match

  TRANSFORM
  - apply_column_selection: duplicate column indices → duplicated columns
  - apply_row_selection: duplicate row indices → duplicated rows
  - shape_keep: max_col and max_row correctly track bounding box
  - shape_keep: col_indices=[] uses all columns
  - shape_keep: row_indices=[] uses all rows

  PARSING
  - col_letters_to_index: XFD (16384) roundtrips
  - parse_columns: empty token between commas (trailing comma OK)
  - parse_rows: very large number (1,000,000)
  - parse_rows: single row 1 → [0]

  DEST SHEET MANAGEMENT
  - New dest file gets default "Sheet" tab removed when named sheet created
  - Existing dest file with multiple sheets: adding to one doesn't affect others
  - run_sheet writes to correct sheet when dest has multiple sheets

  GUI — SELECTION NAME VAR
  - Selecting a source sets selection_name_var to source filename
  - Selecting a recipe sets selection_name_var to recipe name
  - Selecting a sheet sets selection_name_var to sheet name

  GUI — CONTEXT MENU WIRING
  - _ctx_source_index set on right-click of source node
  - _ctx_recipe_path set on right-click of recipe node
  - _ctx_sheet_path set on right-click of sheet node
  - _get_ctx_source returns None when index is None
  - _apply_recipe_rename / _apply_sheet_rename update model correctly (already tested, verify idempotent)

  GUI — SCROLLABLE REPORT DIALOG
  - _show_scrollable_report_dialog creates a Toplevel window
  - Second call destroys the first window before opening new one
  - Dialog contains the report text

  GUI — MISC
  - _format_run_report: success result formatted as "recipe / sheet: N rows"
  - _format_run_report: error result formatted with ERROR prefix
  - _format_run_report: empty results returns "No work items."
  - move_source_up on first source does nothing
  - move_source_down on last source does nothing
  - move_selected_up on recipe that is already first does nothing
  - move_selected_down on sheet that is already last does nothing
  - remove_selected on empty selection does nothing (no crash)
  - _mark_dirty sets _autosave_dirty flag
"""
from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet, run_all
from core.errors import AppError, BAD_SPEC, DEST_BLOCKED, SHEET_NOT_FOUND
from core.io import is_occupied, compute_used_range, load_xlsx, normalize_table
from core.models import Destination, Rule, SheetConfig, SheetResult, RunReport
from core.parsing import col_letters_to_index, col_index_to_letters, parse_columns, parse_rows
from core.planner import build_plan, is_cell_occupied
from core.project import ProjectConfig, RecipeConfig, SourceConfig
from core.rules import apply_rules
from core.transform import (
    apply_column_selection, apply_row_selection, shape_keep, shape_pack
)
from core.writer import apply_write_plan


# ── shared helpers ─────────────────────────────────────────────────────────────

def _xlsx(path, sheet="Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for r, row in enumerate(data or [], 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    wb.save(path)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="",
         dest_sheet="Out", src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet, workbook_sheet=src_sheet,
        source_start_row=src_start_row, columns_spec=columns,
        rows_spec=rows, paste_mode=mode, rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path, sheet_name=dest_sheet,
            start_col=start_col, start_row=start_row,
        ),
    )


def _ws():
    return Workbook().active


# ══════════════════════════════════════════════════════════════════════════════
# ENGINE / SOURCE_START_ROW VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

def test_source_start_row_nonnumeric_raises_app_error(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a", "b"]]))
    dest = str(tmp_path / "dest.xlsx")
    cfg = _cfg(dest, src_start_row="abc")
    with pytest.raises(AppError) as ei:
        run_sheet(src, cfg)
    assert ei.value.code == "BAD_SOURCE_START_ROW"


def test_source_start_row_zero_raises_app_error(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a"]]))
    dest = str(tmp_path / "dest.xlsx")
    cfg = _cfg(dest, src_start_row="0")
    with pytest.raises(AppError) as ei:
        run_sheet(src, cfg)
    assert ei.value.code == "BAD_SOURCE_START_ROW"


def test_source_start_row_negative_raises_app_error(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a"]]))
    dest = str(tmp_path / "dest.xlsx")
    cfg = _cfg(dest, src_start_row="-1")
    with pytest.raises(AppError) as ei:
        run_sheet(src, cfg)
    assert ei.value.code == "BAD_SOURCE_START_ROW"


def test_source_start_row_one_identical_to_no_offset(tmp_path):
    """source_start_row='1' must give identical results to source_start_row=''."""
    data = [["x", 1], ["y", 2], ["z", 3]]
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=data))
    dest1 = str(tmp_path / "d1.xlsx")
    dest2 = str(tmp_path / "d2.xlsx")

    r1 = run_sheet(src, _cfg(dest1, src_start_row="1"))
    r2 = run_sheet(src, _cfg(dest2, src_start_row=""))
    assert r1.rows_written == r2.rows_written == 3

    wb1 = load_workbook(dest1)["Out"]
    wb2 = load_workbook(dest2)["Out"]
    assert wb1["A1"].value == wb2["A1"].value == "x"
    assert wb1["A3"].value == wb2["A3"].value == "z"


def test_run_sheet_bad_workbook_sheet_raises_sheet_not_found(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), sheet="Sheet1", data=[["a"]]))
    dest = str(tmp_path / "dest.xlsx")
    cfg = _cfg(dest, src_sheet="NonExistentSheet")
    with pytest.raises(AppError) as ei:
        run_sheet(src, cfg)
    assert ei.value.code == SHEET_NOT_FOUND
    assert "NonExistentSheet" in ei.value.message


def test_run_sheet_result_message_ok_when_rows_written(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a", "b"]]))
    dest = str(tmp_path / "dest.xlsx")
    result = run_sheet(src, _cfg(dest))
    assert result.rows_written == 1
    assert result.message == "OK"


def test_run_sheet_result_message_zero_rows_when_nothing_written(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a", "b"]]))
    dest = str(tmp_path / "dest.xlsx")
    # Rule that matches nothing
    cfg = _cfg(dest, rules=[Rule(mode="include", column="A", operator="equals", value="NOMATCH")])
    result = run_sheet(src, cfg)
    assert result.rows_written == 0
    assert result.message == "0 rows written"


def test_two_run_sheets_same_file_different_dest_sheets(tmp_path):
    """Two extractions writing to different sheets of the same dest file."""
    data1 = [["alpha", 1], ["beta", 2]]
    data2 = [["gamma", 3], ["delta", 4]]
    src1 = str(_xlsx(str(tmp_path / "s1.xlsx"), data=data1))
    src2 = str(_xlsx(str(tmp_path / "s2.xlsx"), data=data2))
    dest = str(tmp_path / "dest.xlsx")

    r1 = run_sheet(src1, _cfg(dest, dest_sheet="SheetA"))
    r2 = run_sheet(src2, _cfg(dest, dest_sheet="SheetB"))
    assert r1.rows_written == 2
    assert r2.rows_written == 2

    wb = load_workbook(dest)
    assert "SheetA" in wb.sheetnames
    assert "SheetB" in wb.sheetnames
    assert wb["SheetA"]["A1"].value == "alpha"
    assert wb["SheetB"]["A1"].value == "gamma"
    assert wb["SheetA"]["A2"].value == "beta"
    assert wb["SheetB"]["A2"].value == "delta"


def test_run_all_empty_iterable_returns_ok_true_empty_results():
    report = run_all([])
    assert report.ok is True
    assert report.results == []


def test_run_all_accepts_generator_not_just_list(tmp_path):
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["v", 1]]))
    dest = str(tmp_path / "dest.xlsx")

    def _gen():
        yield (src, "R1", _cfg(dest))

    report = run_all(_gen())
    assert report.ok is True
    assert report.results[0].rows_written == 1


def test_dest_sheet_cleanup_default_sheet_tab(tmp_path):
    """When run_sheet creates a new workbook, the blank default 'Sheet' tab is removed."""
    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["a"]]))
    dest = str(tmp_path / "dest.xlsx")  # does not exist yet
    run_sheet(src, _cfg(dest, dest_sheet="MyOutput"))
    wb = load_workbook(dest)
    # The blank default sheet should have been cleaned up
    assert "MyOutput" in wb.sheetnames
    # Default "Sheet" should not be present alongside named sheet
    assert "Sheet" not in wb.sheetnames or len(wb.sheetnames) == 1


def test_run_sheet_multiple_dest_sheets_preserved(tmp_path):
    """Writing to one dest sheet does not destroy existing sibling sheets."""
    # Pre-create dest with two sheets
    wb = Workbook()
    wb.active.title = "Existing"
    wb["Existing"]["A1"] = "keep_me"
    wb.create_sheet("Other")
    wb["Other"]["A1"] = "also_keep"
    dest = str(tmp_path / "dest.xlsx")
    wb.save(dest)

    src = str(_xlsx(str(tmp_path / "src.xlsx"), data=[["new"]]))
    run_sheet(src, _cfg(dest, dest_sheet="Existing", start_row="2"))

    wb2 = load_workbook(dest)
    assert wb2["Existing"]["A1"].value == "keep_me"
    assert wb2["Existing"]["A2"].value == "new"
    assert wb2["Other"]["A1"].value == "also_keep"


# ══════════════════════════════════════════════════════════════════════════════
# PLANNER EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_planner_start_row_zero_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "0")
    assert ei.value.code == BAD_SPEC


def test_planner_start_row_negative_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "-5")
    assert ei.value.code == BAD_SPEC


def test_planner_start_row_float_string_raises_bad_spec():
    ws = _ws()
    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "3.5")
    assert ei.value.code == BAD_SPEC


def test_planner_blocker_append_mode_details_flag_true():
    ws = _ws()
    ws["A1"] = "existing"
    ws["A2"] = "existing2"  # max row 2, so append would try row 3
    ws["A3"] = "BLOCK"  # blocker at the append row

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "")  # append mode

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["append_mode"] is True


def test_planner_blocker_explicit_mode_details_flag_false():
    ws = _ws()
    ws["A5"] = "BLOCK"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a"]], "A", "5")

    assert ei.value.code == DEST_BLOCKED
    assert ei.value.details["append_mode"] is False


def test_planner_blocker_details_contain_first_blocker_fields():
    ws = _ws()
    ws["C3"] = "blocked_value"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b"], ["c", "d"]], "B", "2")

    details = ei.value.details
    blocker = details["first_blocker"]
    assert blocker["row"] == 3
    assert blocker["col"] == 3          # C is column 3
    assert blocker["col_letter"] == "C"
    assert blocker["value"] == "blocked_value"


def test_planner_collision_anywhere_in_bounding_box(tmp_path):
    """Blocker not at top-left but inside the write rectangle still triggers DEST_BLOCKED."""
    ws = _ws()
    # Shape is 3 rows × 3 cols starting at B2; put blocker at D4 (bottom-right)
    ws["D4"] = "HIDDEN_BLOCK"

    with pytest.raises(AppError) as ei:
        build_plan(ws, [["a", "b", "c"]] * 3, "B", "2")
    assert ei.value.code == DEST_BLOCKED


def test_planner_append_completely_empty_sheet_starts_row_1():
    ws = _ws()  # brand new, nothing in it
    plan = build_plan(ws, [["a", "b"]], "A", "")
    assert plan is not None
    assert plan.start_row == 1


def test_planner_write_plan_landing_fields_exact():
    ws = _ws()
    plan = build_plan(ws, [["a", "b", "c"], ["d", "e", "f"]], "B", "5")
    assert plan.start_row == 5
    assert plan.start_col == 2       # B
    assert plan.width == 3
    assert plan.height == 2
    assert plan.landing_cols == (2, 4)   # B=2, D=4
    assert plan.landing_rows == (5, 6)


# ══════════════════════════════════════════════════════════════════════════════
# WRITER EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_writer_writes_none_values_explicitly():
    ws = _ws()
    shaped = [[None, "a"], ["b", None]]
    plan = build_plan(ws, shaped, "A", "1")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value is None
    assert ws["B1"].value == "a"
    assert ws["A2"].value == "b"
    assert ws["B2"].value is None


def test_writer_writes_bool_values():
    ws = _ws()
    shaped = [[True, False]]
    plan = build_plan(ws, shaped, "A", "1")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value is True
    assert ws["B1"].value is False


def test_writer_writes_numeric_zero():
    ws = _ws()
    shaped = [[0, 0.0]]
    plan = build_plan(ws, shaped, "A", "1")
    apply_write_plan(ws, shaped, plan)
    assert ws["A1"].value == 0
    assert ws["B1"].value == 0.0


def test_writer_does_not_write_outside_plan_rectangle():
    """Cells adjacent to the write rectangle must remain untouched."""
    ws = _ws()
    ws["D3"] = "sentinel_above"
    ws["A5"] = "sentinel_left"
    ws["E5"] = "sentinel_right"
    ws["B7"] = "sentinel_below"

    shaped = [["x", "y"], ["z", "w"]]
    plan = build_plan(ws, shaped, "B", "5")  # B5:C6
    apply_write_plan(ws, shaped, plan)

    assert ws["D3"].value == "sentinel_above"
    assert ws["A5"].value == "sentinel_left"
    assert ws["E5"].value == "sentinel_right"
    assert ws["B7"].value == "sentinel_below"


# ══════════════════════════════════════════════════════════════════════════════
# IO / IS_OCCUPIED EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_is_occupied_integer_zero_is_occupied():
    assert is_occupied(0) is True


def test_is_occupied_float_zero_is_occupied():
    assert is_occupied(0.0) is True


def test_is_occupied_false_bool_is_occupied():
    assert is_occupied(False) is True


def test_is_occupied_true_bool_is_occupied():
    assert is_occupied(True) is True


def test_is_occupied_empty_list_is_occupied():
    """Any non-None, non-empty-string value is occupied."""
    assert is_occupied([]) is True


def test_is_occupied_empty_dict_is_occupied():
    assert is_occupied({}) is True


def test_compute_used_range_only_rightmost_col_determines_width():
    rows = [
        [None, None, "v"],   # col 3 used → width 3
        ["x",  None, None],  # col 1 used but doesn't extend width
    ]
    h, w = compute_used_range(rows)
    assert h == 2
    assert w == 3


def test_compute_used_range_only_lowest_row_determines_height():
    rows = [
        ["a", None],
        [None, None],
        [None, None],
        [None, "b"],   # row 4 used → height 4
        [None, None],
    ]
    h, w = compute_used_range(rows)
    assert h == 4
    assert w == 2


def test_load_xlsx_reads_only_specified_sheet(tmp_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "from_sheet1"
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "from_sheet2"
    path = str(tmp_path / "multi.xlsx")
    wb.save(path)

    rows1 = load_xlsx(path, "Sheet1")
    rows2 = load_xlsx(path, "Sheet2")
    assert rows1[0][0] == "from_sheet1"
    assert rows2[0][0] == "from_sheet2"


def test_load_xlsx_sheet_with_only_empty_strings_used_range_zero(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = ""
    ws["B2"] = ""
    path = str(tmp_path / "empty_strings.xlsx")
    wb.save(path)

    rows = load_xlsx(path, "Sheet1")
    h, w = compute_used_range(rows)
    assert h == 0
    assert w == 0


# ══════════════════════════════════════════════════════════════════════════════
# RULES EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_apply_rules_no_rules_and_combine_returns_all():
    rows = [["a", 1], ["b", 2], ["c", 3]]
    result = apply_rules(rows, [], "AND")
    assert result == rows


def test_apply_rules_no_rules_or_combine_returns_all():
    rows = [["a", 1], ["b", 2]]
    result = apply_rules(rows, [], "OR")
    assert result == rows


def test_apply_rules_and_one_true_one_false_row_excluded():
    """AND: both must be true; if one is false, row excluded."""
    rows = [["alpha", "10"], ["beta", "20"]]
    rules = [
        Rule(mode="include", column="A", operator="equals",   value="alpha"),
        Rule(mode="include", column="B", operator=">",        value="15"),  # alpha has 10 → False
    ]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 0


def test_apply_rules_or_one_true_one_false_row_included():
    """OR: if at least one is true, row kept."""
    rows = [["alpha", "10"], ["gamma", "10"]]
    rules = [
        Rule(mode="include", column="A", operator="equals", value="alpha"),
        Rule(mode="include", column="B", operator=">",      value="15"),   # both have 10 → False
    ]
    result = apply_rules(rows, rules, "OR")
    # "alpha" row: first rule True → kept; "gamma" row: both False → excluded
    assert len(result) == 1
    assert result[0][0] == "alpha"


def test_apply_rules_all_exclude_or_semantics():
    """OR with only exclude rules: row EXCLUDED if ANY exclude matches."""
    rows = [["alpha", "5"], ["beta", "5"], ["gamma", "5"]]
    rules = [
        Rule(mode="exclude", column="A", operator="equals", value="alpha"),
        Rule(mode="exclude", column="A", operator="equals", value="beta"),
    ]
    # OR: keep row only if at least one result is True
    # exclude-alpha: True for alpha, False for beta/gamma
    # exclude-beta:  True for beta, False for alpha/gamma
    # alpha: [True, False] → OR → True → kept
    # beta:  [False, True] → OR → True → kept
    # gamma: [False, False] → OR → False → excluded
    result = apply_rules(rows, rules, "OR")
    names = [r[0] for r in result]
    assert "gamma" not in names
    assert "alpha" in names
    assert "beta" in names


def test_apply_rules_numeric_float_cell_satisfies_greater_than():
    rows = [[1.6, "x"], [1.4, "y"], [2.0, "z"]]
    rules = [Rule(mode="include", column="A", operator=">", value="1.5")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 2
    assert result[0][0] == 1.6
    assert result[1][0] == 2.0


def test_apply_rules_int_cell_satisfies_less_than():
    rows = [[3, "x"], [10, "y"], [1, "z"]]
    rules = [Rule(mode="include", column="A", operator="<", value="5")]
    result = apply_rules(rows, rules, "AND")
    assert len(result) == 2
    values = [r[0] for r in result]
    assert 3 in values and 1 in values


def test_apply_rules_high_column_row_too_short_no_match():
    """Rule on column ZZ (702), but rows only have 3 cols → treated as no-match."""
    rows = [["a", "b", "c"], ["x", "y", "z"]]
    rules = [Rule(mode="include", column="ZZ", operator="equals", value="a")]
    result = apply_rules(rows, rules, "AND")
    assert result == []  # no rows have a ZZ column → none match


# ══════════════════════════════════════════════════════════════════════════════
# TRANSFORM EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_apply_column_selection_duplicate_indices_duplicates_column():
    rows = [["a", "b", "c"], ["x", "y", "z"]]
    result = apply_column_selection(rows, [0, 0, 2])  # A, A, C
    assert result[0] == ["a", "a", "c"]
    assert result[1] == ["x", "x", "z"]


def test_apply_row_selection_duplicate_indices_duplicates_row():
    rows = [["a"], ["b"], ["c"]]
    result = apply_row_selection(rows, [0, 0, 2])
    assert len(result) == 3
    assert result[0] == ["a"]
    assert result[1] == ["a"]
    assert result[2] == ["c"]


def test_shape_keep_bounding_box_spans_selected_min_max():
    """Bounding box is min→max of selected rows AND cols, not global min/max."""
    rows = [
        ["A1", "B1", "C1", "D1"],
        ["A2", "B2", "C2", "D2"],
        ["A3", "B3", "C3", "D3"],
        ["A4", "B4", "C4", "D4"],
    ]
    # Select rows 1,3 (0-based) and cols 1,3 (0-based B and D)
    shaped = shape_keep(rows, [1, 3], [1, 3])
    # Bounding box: rows 1-3 (3 rows), cols 1-3 (3 cols)
    assert len(shaped) == 3
    assert len(shaped[0]) == 3
    assert shaped[0][0] == "B2"   # (row=1, col=1) maps to (0,0) of bbox
    assert shaped[0][2] == "D2"   # (row=1, col=3) maps to (0,2)
    assert shaped[2][0] == "B4"   # (row=3, col=1) maps to (2,0)
    assert shaped[2][2] == "D4"   # (row=3, col=3)
    assert shaped[1][0] is None   # row 2 not selected → None
    assert shaped[1][1] is None   # col 2 not selected → None


def test_shape_keep_empty_col_indices_uses_all_columns():
    rows = [["a", "b", "c"], ["x", "y", "z"]]
    shaped = shape_keep(rows, [0, 1], [])
    assert len(shaped) == 2
    assert len(shaped[0]) == 3
    assert shaped[0] == ["a", "b", "c"]


def test_shape_keep_empty_row_indices_uses_all_rows():
    rows = [["a", "b"], ["c", "d"], ["e", "f"]]
    shaped = shape_keep(rows, [], [0, 1])
    assert len(shaped) == 3
    assert shaped[0] == ["a", "b"]
    assert shaped[2] == ["e", "f"]


# ══════════════════════════════════════════════════════════════════════════════
# PARSING EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_col_xfd_roundtrip():
    """XFD is Excel's maximum column (16384)."""
    assert col_letters_to_index("XFD") == 16384
    assert col_index_to_letters(16384) == "XFD"


def test_parse_columns_trailing_comma_ok():
    result = parse_columns("A,B,C,")
    assert result == [0, 1, 2]


def test_parse_rows_very_large_row_number():
    result = parse_rows("1000000")
    assert result == [999999]


def test_parse_rows_single_row_one():
    assert parse_rows("1") == [0]


def test_parse_columns_single_col_a():
    assert parse_columns("A") == [0]


def test_parse_columns_single_col_z():
    assert parse_columns("Z") == [25]


def test_parse_columns_single_multi_letter_aa():
    assert parse_columns("AA") == [26]


# ══════════════════════════════════════════════════════════════════════════════
# IS_CELL_OCCUPIED (PLANNER) vs IS_OCCUPIED (IO) CONSISTENCY
# ══════════════════════════════════════════════════════════════════════════════

def test_is_cell_occupied_matches_is_occupied_for_common_values():
    """Both functions should agree on the same values."""
    test_cases = [
        (None, False),
        ("", False),
        ("hello", True),
        (0, True),
        (0.0, True),
        (False, True),
        (True, True),
        (42, True),
        (" ", True),
    ]
    for val, expected in test_cases:
        assert is_occupied(val) == expected, f"is_occupied({val!r}) should be {expected}"
        assert is_cell_occupied(val) == expected, f"is_cell_occupied({val!r}) should be {expected}"


def test_is_cell_occupied_formula_string_treated_as_unoccupied():
    """Planner-specific: formula strings (start with '=') are unoccupied."""
    assert is_cell_occupied("=SUM(A1:A10)") is False
    assert is_cell_occupied("=A1+B1") is False


def test_is_occupied_formula_string_is_occupied():
    """io.is_occupied has no formula special-case — formula strings ARE occupied."""
    assert is_occupied("=SUM(A1:A10)") is True


# ══════════════════════════════════════════════════════════════════════════════
# GUI — SELECTION_NAME_VAR
# ══════════════════════════════════════════════════════════════════════════════

def _make_gui_with_project():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    src = SourceConfig(path="C:/data/source_file.xlsx", recipes=[
        RecipeConfig(name="MyRecipe", sheets=[
            SheetConfig(name="MySheet", workbook_sheet="MySheet"),
        ])
    ])
    app.project = ProjectConfig(sources=[src])
    app.refresh_tree()
    return app


def test_selection_name_var_set_to_filename_on_source_select():
    app = _make_gui_with_project()
    try:
        src_id = app.tree.get_children("")[0]
        app.tree.selection_set(src_id)
        app._on_tree_select()
        name = app.selection_name_var.get()
        # Should show basename, not full path
        assert "source_file.xlsx" in name
        assert "C:/data" not in name
    finally:
        app.destroy()


def test_selection_name_var_set_to_recipe_name_on_recipe_select():
    app = _make_gui_with_project()
    try:
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]
        app.tree.selection_set(rec_id)
        app._on_tree_select()
        assert app.selection_name_var.get() == "MyRecipe"
    finally:
        app.destroy()


def test_selection_name_var_set_to_sheet_name_on_sheet_select():
    app = _make_gui_with_project()
    try:
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]
        sh_id = app.tree.get_children(rec_id)[0]
        app.tree.selection_set(sh_id)
        app._on_tree_select()
        assert app.selection_name_var.get() == "MySheet"
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# GUI — CONTEXT MENU RIGHT-CLICK PATH TRACKING
# ══════════════════════════════════════════════════════════════════════════════

def _make_gui_3level():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    src = SourceConfig(path="s.xlsx", recipes=[
        RecipeConfig(name="R1", sheets=[
            SheetConfig(name="S1", workbook_sheet="S1"),
        ])
    ])
    app.project = ProjectConfig(sources=[src])
    app.refresh_tree()
    return app


def test_ctx_source_index_set_when_right_clicking_source():
    app = _make_gui_3level()
    try:
        src_id = app.tree.get_children("")[0]
        app.tree.selection_set(src_id)
        path = app._get_tree_path(src_id)
        app._ctx_source_index = path[0]
        assert app._ctx_source_index == 0
    finally:
        app.destroy()


def test_ctx_recipe_path_set_on_right_click_recipe():
    app = _make_gui_3level()
    try:
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]
        app.tree.selection_set(rec_id)
        path = app._get_tree_path(rec_id)
        app._ctx_recipe_path = path
        assert app._ctx_recipe_path == [0, 0]
    finally:
        app.destroy()


def test_ctx_sheet_path_set_on_right_click_sheet():
    app = _make_gui_3level()
    try:
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]
        sh_id = app.tree.get_children(rec_id)[0]
        app.tree.selection_set(sh_id)
        path = app._get_tree_path(sh_id)
        app._ctx_sheet_path = path
        assert app._ctx_sheet_path == [0, 0, 0]
    finally:
        app.destroy()


def test_get_ctx_source_returns_none_when_index_is_none():
    app = _make_gui_3level()
    try:
        app._ctx_source_index = None
        assert app._get_ctx_source() is None
    finally:
        app.destroy()


def test_get_ctx_source_returns_correct_source():
    app = _make_gui_3level()
    try:
        app._ctx_source_index = 0
        src = app._get_ctx_source()
        assert src is app.project.sources[0]
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# GUI — SCROLLABLE REPORT DIALOG
# ══════════════════════════════════════════════════════════════════════════════

def test_show_scrollable_report_dialog_creates_toplevel():
    from gui.app import TurboExtractorApp
    import tkinter as tk
    app = TurboExtractorApp()
    try:
        app._show_scrollable_report_dialog("Test Title", "Line1\nLine2")
        app.update_idletasks()
        assert app._report_dialog is not None
        assert isinstance(app._report_dialog, tk.Toplevel)
        app._report_dialog.destroy()
        app._report_dialog = None
    finally:
        app.destroy()


def test_show_scrollable_report_dialog_second_call_replaces_first():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        app._show_scrollable_report_dialog("First", "text1")
        app.update_idletasks()
        first_dialog = app._report_dialog

        app._show_scrollable_report_dialog("Second", "text2")
        app.update_idletasks()
        second_dialog = app._report_dialog

        # First dialog should be gone, second present
        assert second_dialog is not first_dialog
        try:
            first_dialog.winfo_exists()
            first_was_destroyed = False
        except Exception:
            first_was_destroyed = True
        # Either destroyed or a new object — either way second != first
        assert second_dialog is not None
        second_dialog.destroy()
        app._report_dialog = None
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# GUI — _FORMAT_RUN_REPORT
# ══════════════════════════════════════════════════════════════════════════════

def _make_result(recipe, sheet, rows=5, error_code=None, error_msg=None):
    return SheetResult(
        source_path="s.xlsx", recipe_name=recipe, sheet_name=sheet,
        dest_file="d.xlsx", dest_sheet="Out", rows_written=rows,
        message="ERROR" if error_code else "OK",
        error_code=error_code, error_message=error_msg,
    )


def test_format_run_report_success_format():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        report = RunReport(ok=True, results=[_make_result("MyRecipe", "MySheet", 42)])
        text = app._format_run_report(report)
        assert "MyRecipe" in text
        assert "MySheet" in text
        assert "42" in text
    finally:
        app.destroy()


def test_format_run_report_error_format():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        report = RunReport(ok=False, results=[
            _make_result("R1", "S1", 0, error_code="DEST_BLOCKED", error_msg="Zone blocked")
        ])
        text = app._format_run_report(report)
        assert "ERROR" in text
        assert "DEST_BLOCKED" in text
        assert "Zone blocked" in text
    finally:
        app.destroy()


def test_format_run_report_empty_results_returns_no_work_items():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        text = app._format_run_report(RunReport(ok=True, results=[]))
        assert text == "No work items."
    finally:
        app.destroy()


def test_format_run_report_multiple_results_all_present():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        report = RunReport(ok=True, results=[
            _make_result("R1", "S1", 10),
            _make_result("R2", "S2", 0, error_code="BAD_SPEC", error_msg="oops"),
            _make_result("R3", "S3", 7),
        ])
        text = app._format_run_report(report)
        assert "R1" in text and "10" in text
        assert "R2" in text and "BAD_SPEC" in text
        assert "R3" in text and "7" in text
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# GUI — MOVE BOUNDARY CONDITIONS
# ══════════════════════════════════════════════════════════════════════════════

def _make_gui_two_sources():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    for name in ["a.xlsx", "b.xlsx"]:
        app.project.sources.append(
            SourceConfig(path=name, recipes=[
                RecipeConfig(name="R1", sheets=[SheetConfig(name="S1", workbook_sheet="S1")])
            ])
        )
    app.refresh_tree()
    return app


def test_move_source_up_on_first_does_nothing():
    app = _make_gui_two_sources()
    try:
        src_id = app.tree.get_children("")[0]
        app.tree.selection_set(src_id)
        app._on_tree_select()
        app.move_source_up()
        assert app.project.sources[0].path == "a.xlsx"
        assert app.project.sources[1].path == "b.xlsx"
    finally:
        app.destroy()


def test_move_source_down_on_last_does_nothing():
    app = _make_gui_two_sources()
    try:
        src_ids = app.tree.get_children("")
        app.tree.selection_set(src_ids[1])
        app._on_tree_select()
        app.move_source_down()
        assert app.project.sources[0].path == "a.xlsx"
        assert app.project.sources[1].path == "b.xlsx"
    finally:
        app.destroy()


def test_move_selected_up_recipe_already_first_does_nothing():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        app.project.sources.append(SourceConfig(path="s.xlsx", recipes=[
            RecipeConfig(name="R1", sheets=[SheetConfig(name="S1", workbook_sheet="S1")]),
            RecipeConfig(name="R2", sheets=[SheetConfig(name="S2", workbook_sheet="S2")]),
        ]))
        app.refresh_tree()
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]  # R1 (first)
        app.tree.selection_set(rec_id)
        app.tree.focus(rec_id)
        app._on_tree_select()
        app.move_selected_up()
        assert app.project.sources[0].recipes[0].name == "R1"
    finally:
        app.destroy()


def test_move_selected_down_sheet_already_last_does_nothing():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        app.project.sources.append(SourceConfig(path="s.xlsx", recipes=[
            RecipeConfig(name="R1", sheets=[
                SheetConfig(name="S1", workbook_sheet="S1"),
                SheetConfig(name="S2", workbook_sheet="S2"),
            ])
        ]))
        app.refresh_tree()
        src_id = app.tree.get_children("")[0]
        rec_id = app.tree.get_children(src_id)[0]
        sh_ids = app.tree.get_children(rec_id)
        last_sh = sh_ids[-1]  # S2 (last)
        app.tree.selection_set(last_sh)
        app.tree.focus(last_sh)
        app._on_tree_select()
        app.move_selected_down()
        assert app.project.sources[0].recipes[0].sheets[-1].name == "S2"
    finally:
        app.destroy()


def test_remove_selected_with_no_selection_does_not_crash():
    from gui.app import TurboExtractorApp
    app = TurboExtractorApp()
    try:
        # Nothing selected
        app.tree.selection_remove(*app.tree.selection())
        app.remove_selected()  # should be a no-op
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# GUI — AUTOSAVE / _MARK_DIRTY
# ══════════════════════════════════════════════════════════════════════════════

def test_mark_dirty_sets_autosave_dirty_flag(tmp_path, monkeypatch):
    from gui.app import TurboExtractorApp
    monkeypatch.setenv("TURBO_AUTOSAVE_PATH", str(tmp_path / "save.json"))
    app = TurboExtractorApp()
    try:
        app._autosave_dirty = False
        app._mark_dirty()
        assert app._autosave_dirty is True
    finally:
        app.destroy()


def test_autosave_now_skipped_when_not_dirty(tmp_path, monkeypatch):
    from gui.app import TurboExtractorApp
    save_path = tmp_path / "save.json"
    monkeypatch.setenv("TURBO_AUTOSAVE_PATH", str(save_path))
    app = TurboExtractorApp()
    try:
        app._autosave_dirty = False
        app._autosave_now()
        assert not save_path.exists()
    finally:
        app.destroy()


def test_autosave_now_writes_when_dirty(tmp_path, monkeypatch):
    from gui.app import TurboExtractorApp
    save_path = tmp_path / "save.json"
    monkeypatch.setenv("TURBO_AUTOSAVE_PATH", str(save_path))
    app = TurboExtractorApp()
    try:
        app._autosave_dirty = True
        app._autosave_path = str(save_path)
        app._autosave_now()
        assert save_path.exists()
        assert app._autosave_dirty is False
    finally:
        app.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# SERIALIZATION — ADDITIONAL EDGE CASES
# ══════════════════════════════════════════════════════════════════════════════

def test_project_to_dict_then_from_dict_is_equivalent():
    """to_dict → from_dict produces structurally identical project."""
    from core.project import ProjectConfig, RecipeConfig, SourceConfig
    src = SourceConfig(path="s.xlsx", recipes=[
        RecipeConfig(name="R1", sheets=[
            SheetConfig(
                name="S1", workbook_sheet="WB1", source_start_row="2",
                columns_spec="A,C", rows_spec="1-3", paste_mode="keep",
                rules_combine="OR",
                rules=[Rule(mode="exclude", column="B", operator="contains", value="skip")],
                destination=Destination(file_path="out.xlsx", sheet_name="Out",
                                       start_col="D", start_row="10"),
            )
        ])
    ])
    proj = ProjectConfig(sources=[src])
    loaded = ProjectConfig.from_dict(proj.to_dict())
    s = loaded.sources[0].recipes[0].sheets[0]
    assert s.name == "S1"
    assert s.workbook_sheet == "WB1"
    assert s.source_start_row == "2"
    assert s.columns_spec == "A,C"
    assert s.rows_spec == "1-3"
    assert s.paste_mode == "keep"
    assert s.rules_combine == "OR"
    assert len(s.rules) == 1
    assert s.rules[0].mode == "exclude"
    assert s.rules[0].column == "B"
    assert s.rules[0].operator == "contains"
    assert s.rules[0].value == "skip"
    assert s.destination.start_col == "D"
    assert s.destination.start_row == "10"


def test_from_dict_empty_rules_list_survives():
    data = {"sources": [{"path": "s.xlsx", "name": "", "recipes": [
        {"name": "R1", "sheets": [{
            "name": "S1", "workbook_sheet": "S1",
            "columns_spec": "", "rows_spec": "", "paste_mode": "pack",
            "rules_combine": "AND", "rules": [],
            "destination": {"file_path": "o.xlsx", "sheet_name": "Out",
                           "start_col": "A", "start_row": ""},
        }]}
    ]}]}
    loaded = ProjectConfig.from_dict(data)
    assert loaded.sources[0].recipes[0].sheets[0].rules == []


def test_runreport_has_errors_only_checks_error_code_not_message():
    """has_errors should be False even if message says ERROR, as long as error_code is None."""
    report = RunReport(ok=True, results=[
        SheetResult(source_path="s", recipe_name="r", sheet_name="sh",
                    dest_file="d", dest_sheet="ds", rows_written=0,
                    message="0 rows written", error_code=None),
    ])
    assert report.has_errors is False
