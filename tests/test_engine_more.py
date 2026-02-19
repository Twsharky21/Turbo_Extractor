\
import os
import csv
import pytest
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet
from core.models import SheetConfig, Destination, Rule
from core.errors import AppError, DEST_BLOCKED


def make_source_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    data = [
        ["alpha", "x", 1, "foo"],
        ["beta",  "y", 2, "bar"],
        ["gamma", "z", 3, "baz"],
        ["beta",  "y", 4, "qux"],
    ]
    for r, row in enumerate(data, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)


def make_source_csv(path: str):
    data = [
        ["alpha", "x", 1, "foo"],
        ["beta",  "y", 2, "bar"],
        ["gamma", "z", 3, "baz"],
        ["beta",  "y", 4, "qux"],
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in data:
            w.writerow(row)


def read_cells(dest_path: str, sheet: str, addr_list):
    wb = load_workbook(dest_path)
    ws = wb[sheet]
    return [ws[a].value for a in addr_list]


def test_engine_csv_smoke_pack_rules_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.csv")
        dest = os.path.join(td, "dest.xlsx")

        make_source_csv(src)

        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B1"] = "existing"
        ws["C1"] = "existing2"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",  # ignored for CSV
            columns_spec="A,C",
            rows_spec="",
            paste_mode="pack",
            rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="beta")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        vals = read_cells(dest, "Out", ["B2", "C2", "B3", "C3"])
        assert vals[0] == "beta"
        # CSV numeric may be string "2"
        assert vals[1] in (2, "2")
        assert vals[2] == "beta"
        assert vals[3] in (4, "4")


def test_engine_keep_mode_preserves_gaps_no_rules():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        make_source_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",   # gap at B
            rows_spec="1,3",      # gap at row 2
            paste_mode="keep",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )

        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 3  # keep bounding box rows 1-3 => 3 rows

        wb = load_workbook(dest)
        ws = wb["Out"]

        # Output bounding box width=3 (A..C) mapped to start_col B => B..D
        # Row 1 has A1 and C1
        assert ws["B1"].value == "alpha"
        assert ws["C1"].value is None
        assert ws["D1"].value == 1

        # Row 2 (gap) should be all None
        assert ws["B2"].value is None
        assert ws["C2"].value is None
        assert ws["D2"].value is None

        # Row 3 has A3 and C3
        assert ws["B3"].value == "gamma"
        assert ws["C3"].value is None
        assert ws["D3"].value == 3


def test_engine_explicit_start_row_and_col_lands_exactly():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        make_source_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-2",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="D", start_row="50"),
        )

        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["D50"].value == "alpha"
        assert ws["E50"].value == 1
        assert ws["D51"].value == "beta"
        assert ws["E51"].value == 2


def test_engine_dest_blocked_explicit_raises():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        make_source_xlsx(src)

        # Create destination with a blocker at the exact landing cell
        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["D50"] = "BLOCK"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="D", start_row="50"),
        )

        with pytest.raises(AppError) as ei:
            run_sheet(src, cfg, recipe_name="R1")
        assert ei.value.code == DEST_BLOCKED


def test_engine_append_two_runs_stack_by_landing_zone_cols():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        make_source_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        r1 = run_sheet(src, cfg, recipe_name="R1")
        r2 = run_sheet(src, cfg, recipe_name="R1")
        assert r1.rows_written == 1
        assert r2.rows_written == 1

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "alpha"
        assert ws["C1"].value == 1
        assert ws["B2"].value == "alpha"
        assert ws["C2"].value == 1


def test_engine_rules_filter_to_zero_rows_writes_nothing():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")
        make_source_xlsx(src)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="",
            paste_mode="pack",
            rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 0

        wb = load_workbook(dest)
        assert "Out" in wb.sheetnames
        ws = wb["Out"]
        # should remain empty
        assert ws["B1"].value in (None, "")
