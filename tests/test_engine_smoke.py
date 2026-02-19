\
import os
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_sheet
from core.models import SheetConfig, Destination, Rule


def make_source_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 4 cols x 4 rows
    data = [
        ["alpha", "x", 1, "foo"],
        ["beta",  "y", 2, "bar"],
        ["gamma", "z", 3, "baz"],
        ["beta",  "y", 4, "qux"],
    ]
    for r, row in enumerate(data, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)


def test_engine_smoke_pack_columns_and_rules_and_append():
    with TemporaryDirectory() as td:
        src = os.path.join(td, "src.xlsx")
        dest = os.path.join(td, "dest.xlsx")

        make_source_xlsx(src)

        # Pre-populate destination in landing columns B:C to force append to row 2
        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B1"] = "existing"
        ws["C1"] = "existing2"
        wb.save(dest)

        cfg = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",   # take alpha/beta/gamma + numeric col
            rows_spec="",         # ALL used
            paste_mode="pack",
            rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="beta")],
            destination=Destination(
                file_path=dest,
                sheet_name="Out",
                start_col="B",
                start_row="",       # append mode (landing-zone aware)
            ),
        )

        result = run_sheet(src, cfg, recipe_name="R1")
        assert result.rows_written == 2

        out = load_workbook(dest)
        ws = out["Out"]

        # Append should be at row 2 (since max used row in B:C was 1)
        assert ws["B2"].value == "beta"
        assert ws["C2"].value == 2
        assert ws["B3"].value == "beta"
        assert ws["C3"].value == 4
