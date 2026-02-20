"""Tests for core.engine.run_all — batch execution, stacking, fail-fast, progress callbacks."""
import os
import csv
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_all
from core.models import SheetConfig, Destination, Rule
from core.errors import DEST_BLOCKED


# ---- Helpers ----

def _make_xlsx(path: str, sheet: str, rows):
    wb = Workbook(); ws = wb.active; ws.title = sheet
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)


def _make_csv(path: str, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def _pack_cfg(dest, sheet="Out", src_col="B", rows="1-1", columns="A,C"):
    return SheetConfig(
        name="Sheet1", workbook_sheet="Sheet1",
        columns_spec=columns, rows_spec=rows,
        paste_mode="pack", rules_combine="AND", rules=[],
        destination=Destination(file_path=dest, sheet_name=sheet, start_col=src_col, start_row=""),
    )


# ---- Stacking ----

def test_run_all_two_sources_stack_same_dest():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["S1", "x", 1]])
        _make_xlsx(s2, "Sheet1", [["S2", "x", 2]])

        report = run_all([(s1, "R1", _pack_cfg(dest)), (s2, "R2", _pack_cfg(dest))])
        assert report.ok
        assert report.results[0].rows_written == 1
        assert report.results[1].rows_written == 1

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "S1" and ws["C1"].value == 1
        assert ws["B2"].value == "S2" and ws["C2"].value == 2


def test_run_all_three_sources_stack_in_order():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        sources = []
        for i in range(1, 4):
            p = os.path.join(td, f"s{i}.xlsx")
            _make_xlsx(p, "Sheet1", [[f"A{i}", "x", i]])
            sources.append((p, f"R{i}", _pack_cfg(dest)))

        report = run_all(sources)
        assert report.ok
        ws = load_workbook(dest)["Out"]
        for i in range(1, 4):
            assert ws[f"B{i}"].value == f"A{i}"
            assert ws[f"C{i}"].value == i


def test_run_all_two_different_destinations():
    with TemporaryDirectory() as td:
        d1, d2 = os.path.join(td, "o1.xlsx"), os.path.join(td, "o2.xlsx")
        s1, s2 = os.path.join(td, "s1.xlsx"), os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        _make_xlsx(s2, "Sheet1", [["A2", "x", 2]])

        report = run_all([(s1, "R1", _pack_cfg(d1)), (s2, "R2", _pack_cfg(d2))])
        assert report.ok
        assert load_workbook(d1)["Out"]["B1"].value == "A1"
        assert load_workbook(d2)["Out"]["B1"].value == "A2"


def test_run_all_mixed_widths_landing_zone_awareness():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td, "s1.xlsx"), os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["v1", "v2", "v3"]])
        _make_xlsx(s2, "Sheet1", [["w1", "x", "w3"]])

        cfg1 = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,B,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        report = run_all([(s1, "R1", cfg1), (s2, "R2", _pack_cfg(dest))])
        assert report.ok

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "v1" and ws["C1"].value == "v2" and ws["D1"].value == "v3"
        assert ws["B2"].value == "w1" and ws["C2"].value == "w3"


def test_run_all_keep_then_pack_stacks_correctly():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td, "s1.xlsx"), os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["alpha","x",1],["beta","y",2],["gamma","z",3]])
        _make_xlsx(s2, "Sheet1", [["delta","q",9]])

        cfg_keep = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1,3",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        report = run_all([(s1, "R1", cfg_keep), (s2, "R2", _pack_cfg(dest))])
        assert report.ok

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "alpha" and ws["D1"].value == 1
        assert ws["B2"].value is None   # gap row
        assert ws["B3"].value == "gamma" and ws["D3"].value == 3
        assert ws["B4"].value == "delta" and ws["C4"].value == 9


# ---- Fail-fast ----

def test_run_all_fail_fast_on_first_item():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td, "s1.xlsx"), os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["S1", "x", 1]])
        _make_xlsx(s2, "Sheet1", [["S2", "x", 2]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B1"] = "BLOCK"; wb.save(dest)

        blocked_cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )
        report = run_all([(s1, "R1", blocked_cfg), (s2, "R2", _pack_cfg(dest))])

        assert not report.ok
        assert len(report.results) == 1
        assert report.results[0].error_code == DEST_BLOCKED
        assert "S2" not in [load_workbook(dest)["Out"][f"B{i}"].value for i in range(1,5)]


def test_run_all_fail_fast_on_second_item():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2, s3 = [os.path.join(td, f"s{i}.xlsx") for i in range(1,4)]
        for i, p in enumerate([s1,s2,s3], 1):
            _make_xlsx(p, "Sheet1", [[f"A{i}","x",i]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B50"] = "BLOCK"; wb.save(dest)

        blocked_cfg = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="50"),
        )
        report = run_all([(s1,"R1",_pack_cfg(dest)), (s2,"R2",blocked_cfg), (s3,"R3",_pack_cfg(dest))])

        assert not report.ok
        assert len(report.results) == 2
        assert report.results[1].error_code == DEST_BLOCKED
        colB = [load_workbook(dest)["Out"][f"B{i}"].value for i in range(1,10)]
        assert "A3" not in colB


# ---- Zero rows ----

def test_run_all_zero_rows_then_next_still_runs():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td, "s1.xlsx"), os.path.join(td, "s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["alpha","x",1],["beta","y",2]])
        _make_xlsx(s2, "Sheet1", [["gamma","z",3]])

        cfg_zero = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="pack", rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        report = run_all([(s1,"R1",cfg_zero), (s2,"R2",_pack_cfg(dest))])
        assert report.ok
        assert report.results[0].rows_written == 0
        assert report.results[1].rows_written == 1

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "gamma" and ws["C1"].value == 3


# ---- CSV + XLSX mix ----

def test_run_all_csv_then_xlsx_stacks():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td, "s1.csv"), os.path.join(td, "s2.xlsx")
        _make_csv(s1, [["csv1","x",1]])
        _make_xlsx(s2, "Sheet1", [["xls1","y",2]])

        report = run_all([(s1,"R1",_pack_cfg(dest)), (s2,"R2",_pack_cfg(dest))])
        assert report.ok

        ws = load_workbook(dest)["Out"]
        assert ws["B1"].value == "csv1"
        assert ws["C1"].value in (1, "1")
        assert ws["B2"].value == "xls1"
        assert ws["C2"].value == 2


# ---- Progress callbacks ----

def test_run_all_progress_callback_success_order():
    events = []

    def cb(ev, payload): events.append((ev, payload))

    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td,"s1.xlsx"), os.path.join(td,"s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["S1","x",1]])
        _make_xlsx(s2, "Sheet1", [["S2","x",2]])

        report = run_all([(s1,"R1",_pack_cfg(dest)), (s2,"R2",_pack_cfg(dest))], on_progress=cb)
        assert report.ok

        kinds = [e[0] for e in events]
        assert kinds[:4] == ["start","result","start","result"]
        assert kinds[-1] == "done"
        assert getattr(events[1][1], "rows_written") == 1
        assert getattr(events[3][1], "rows_written") == 1


def test_run_all_progress_callback_fail_fast_order():
    events = []

    def cb(ev, payload): events.append((ev, payload))

    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1, s2 = os.path.join(td,"s1.xlsx"), os.path.join(td,"s2.xlsx")
        _make_xlsx(s1, "Sheet1", [["S1","x",1]])
        _make_xlsx(s2, "Sheet1", [["S2","x",2]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["B1"] = "BLOCK"; wb.save(dest)

        blocked = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="1"),
        )
        report = run_all([(s1,"R1",blocked), (s2,"R2",_pack_cfg(dest))], on_progress=cb)
        assert not report.ok

        kinds = [e[0] for e in events]
        assert kinds[0] == "start"
        assert kinds[1] == "error"
        assert "start" not in kinds[2:-1]
        assert kinds[-1] == "done"
