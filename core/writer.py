"""
core/writer.py — Applies a validated WritePlan to a worksheet.

Critical rule: None values are NEVER written to cells. Writing None explicitly
via ws.cell(value=None) registers a phantom cell in openpyxl, inflating
ws.max_row and ws.max_column. This corrupts subsequent append scans.

Gap cells (None) from Keep Format bounding boxes are simply skipped —
their absence in the destination is correct and intentional.
"""
from __future__ import annotations

from typing import Any, List

from openpyxl.worksheet.worksheet import Worksheet

from .planner import WritePlan


def apply_write_plan(
    ws: Worksheet,
    shaped: List[List[Any]],
    plan: WritePlan,
) -> int:
    """
    Write shaped data to the worksheet according to plan.

    Skips any cell whose value is None — these are Keep Format gap cells
    and must not be written (writing None registers a phantom cell that
    poisons ws.max_row for future append scans).

    Returns the number of rows written.
    """
    if not shaped:
        return 0

    for r_offset, row in enumerate(shaped):
        for c_offset, value in enumerate(row):
            if value is None:
                continue          # gap cell — do not write
            ws.cell(
                row=plan.start_row + r_offset,
                column=plan.start_col + c_offset,
                value=value,
            )

    return len(shaped)
