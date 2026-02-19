\
import os
import csv
from tempfile import TemporaryDirectory
from openpyxl import Workbook, load_workbook

from core.engine import run_all
from core.models import SheetConfig, Destination, Rule
from core.errors import DEST_BLOCKED


def make_source_xlsx(path: str, sheet_name: str, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)


def make_source_csv(path: str, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(row)


def test_run_all_three_items_stacks_rows_in_order_same_dest():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")
        s3 = os.path.join(td, "s3.xlsx")

        make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])
        make_source_xlsx(s3, "Sheet1", [["A3", "x", 3]])

        cfg = lambda: SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg()),
            (s2, "R2", cfg()),
            (s3, "R3", cfg()),
        ])

        assert report.ok is True
        assert len(report.results) == 3

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "A1"
        assert ws["C1"].value == 1
        assert ws["B2"].value == "A2"
        assert ws["C2"].value == 2
        assert ws["B3"].value == "A3"
        assert ws["C3"].value == 3


def test_run_all_mixed_widths_append_uses_full_landing_zone_awareness():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        # s1 writes width=3 (A,B,C) -> landing zone B:D
        make_source_xlsx(s1, "Sheet1", [["v1", "v2", "v3"]])
        # s2 writes width=2 (A,C) -> landing zone B:C
        make_source_xlsx(s2, "Sheet1", [["w1", "x", "w3"]])

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,B,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg1),
            (s2, "R2", cfg2),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]
        # first row from cfg1 occupies B1:D1
        assert ws["B1"].value == "v1"
        assert ws["C1"].value == "v2"
        assert ws["D1"].value == "v3"
        # second run uses landing zone B:C; max used row across B:C is 1 => append at 2
        assert ws["B2"].value == "w1"
        assert ws["C2"].value == "w3"


def test_run_all_two_different_destinations_both_created():
    with TemporaryDirectory() as td:
        dest1 = os.path.join(td, "out1.xlsx")
        dest2 = os.path.join(td, "out2.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])

        cfg1 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest1, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg2 = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest2, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg1),
            (s2, "R2", cfg2),
        ])
        assert report.ok is True

        wb1 = load_workbook(dest1)
        wb2 = load_workbook(dest2)
        assert wb1["Out"]["B1"].value == "A1"
        assert wb2["Out"]["B1"].value == "A2"


def test_run_all_keep_mode_then_pack_mode_stacks():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        # 3 rows so keep-mode bounding box height=3 when rows_spec=1,3
        make_source_xlsx(s1, "Sheet1", [
            ["alpha", "x", 1],
            ["beta",  "y", 2],
            ["gamma", "z", 3],
        ])
        make_source_xlsx(s2, "Sheet1", [["delta", "q", 9]])

        cfg_keep = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1,3",
            paste_mode="keep",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_pack = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_keep),
            (s2, "R2", cfg_pack),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]

        # keep run occupies B1:D3 (width=3, height=3)
        assert ws["B1"].value == "alpha"
        assert ws["D1"].value == 1
        assert ws["B2"].value is None
        assert ws["B3"].value == "gamma"
        assert ws["D3"].value == 3

        # pack run uses landing zone B:C; max used row across B:C is 3 => append at 4
        assert ws["B4"].value == "delta"
        assert ws["C4"].value == 9


def test_run_all_fail_fast_on_second_item_dest_blocked():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")
        s3 = os.path.join(td, "s3.xlsx")

        make_source_xlsx(s1, "Sheet1", [["A1", "x", 1]])
        make_source_xlsx(s2, "Sheet1", [["A2", "x", 2]])
        make_source_xlsx(s3, "Sheet1", [["A3", "x", 3]])

        cfg_ok = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        # Second item uses explicit start row that is blocked
        wb = Workbook()
        ws = wb.active
        ws.title = "Out"
        ws["B50"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row="50"),
        )

        report = run_all([
            (s1, "R1", cfg_ok),
            (s2, "R2", cfg_blocked),
            (s3, "R3", cfg_ok),
        ])

        assert report.ok is False
        assert len(report.results) == 2
        assert report.results[1].error_code == DEST_BLOCKED

        wb2 = load_workbook(dest)
        ws2 = wb2["Out"]
        # Third item should not run
        colB = [ws2[f"B{i}"].value for i in range(1, 10)]
        assert "A3" not in colB


def test_run_all_csv_then_xlsx_same_dest_stacks():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.csv")
        s2 = os.path.join(td, "s2.xlsx")

        make_source_csv(s1, [["csv1", "x", 1]])
        make_source_xlsx(s2, "Sheet1", [["xls1", "y", 2]])

        cfg_csv = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_xls = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_csv),
            (s2, "R2", cfg_xls),
        ])
        assert report.ok is True

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "csv1"
        assert ws["C1"].value in (1, "1")
        assert ws["B2"].value == "xls1"
        assert ws["C2"].value == 2


def test_run_all_rules_filter_zero_rows_then_next_item_runs_and_writes():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1 = os.path.join(td, "s1.xlsx")
        s2 = os.path.join(td, "s2.xlsx")

        make_source_xlsx(s1, "Sheet1", [["alpha", "x", 1], ["beta", "y", 2]])
        make_source_xlsx(s2, "Sheet1", [["gamma", "z", 3]])

        cfg_zero = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="",
            paste_mode="pack",
            rules_combine="AND",
            rules=[Rule(mode="include", column="A", operator="equals", value="NO_MATCH")],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )
        cfg_write = SheetConfig(
            name="Sheet1",
            workbook_sheet="Sheet1",
            columns_spec="A,C",
            rows_spec="1-1",
            paste_mode="pack",
            rules_combine="AND",
            rules=[],
            destination=Destination(file_path=dest, sheet_name="Out", start_col="B", start_row=""),
        )

        report = run_all([
            (s1, "R1", cfg_zero),
            (s2, "R2", cfg_write),
        ])
        assert report.ok is True
        assert report.results[0].rows_written == 0
        assert report.results[1].rows_written == 1

        wb = load_workbook(dest)
        ws = wb["Out"]
        assert ws["B1"].value == "gamma"
        assert ws["C1"].value == 3
