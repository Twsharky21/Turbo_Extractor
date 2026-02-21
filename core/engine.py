from __future__ import annotations

from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple
import os

from openpyxl import Workbook, load_workbook

from .errors import AppError, SOURCE_READ_FAILED, SHEET_NOT_FOUND
from .models import SheetConfig, SheetResult, RunReport
from .parsing import parse_columns, parse_rows
from .io import load_csv, load_xlsx, compute_used_range, normalize_table
from .rules import apply_rules
from .transform import apply_row_selection, apply_column_selection, shape_pack, shape_keep
from .planner import build_plan
from .writer import apply_write_plan


def _apply_source_start_row(table: List[List[Any]], source_start_row: str) -> List[List[Any]]:
    """Apply a 1-based source start row offset. Blank means no offset."""
    s = (source_start_row or "").strip()
    if not s:
        return table
    try:
        n = int(s)
    except ValueError:
        raise AppError("BAD_SOURCE_START_ROW", f"Source Start Row must be a number (got '{source_start_row}')")
    if n < 1:
        raise AppError("BAD_SOURCE_START_ROW", f"Source Start Row must be >= 1 (got {n})")
    offset = n - 1
    if offset <= 0:
        return table
    return table[offset:]


def _load_source_table(source_path: str, workbook_sheet: str) -> List[List[Any]]:
    ext = os.path.splitext(source_path)[1].lower()
    try:
        if ext == ".csv":
            return load_csv(source_path)
        return load_xlsx(source_path, workbook_sheet)
    except ValueError as e:
        raise AppError(SHEET_NOT_FOUND, str(e))
    except AppError:
        raise
    except Exception as e:
        raise AppError(SOURCE_READ_FAILED, f"Failed to read source: {e}")


def _open_or_create_dest(dest_path: str) -> Workbook:
    if dest_path and os.path.exists(dest_path):
        return load_workbook(dest_path)
    return Workbook()


def _get_or_create_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        default = wb["Sheet"]
        if default.max_row == 1 and default.max_column == 1 and default["A1"].value in (None, ""):
            wb.remove(default)
    return ws


def run_sheet(
    source_path: str,
    sheet_cfg: SheetConfig,
    recipe_name: str = "Recipe",
    _wb_cache: Optional[Dict[str, Workbook]] = None,
) -> SheetResult:
    """
    Execute a single sheet extraction.

    _wb_cache: optional dict keyed by dest_path holding in-memory Workbook objects.
               When provided (by run_all), the workbook is NOT saved to disk here —
               run_all manages saves after each item. When None (standalone call),
               open, write, and save here.
    """
    table = _load_source_table(source_path, sheet_cfg.workbook_sheet)
    table = _apply_source_start_row(table, getattr(sheet_cfg, "source_start_row", ""))

    used_h, used_w = compute_used_range(table)
    table = normalize_table(table)

    row_indices = parse_rows(sheet_cfg.rows_spec)
    if not row_indices:
        row_indices = list(range(used_h))

    table_rows = apply_row_selection(table, row_indices)
    table_rows = apply_rules(table_rows, sheet_cfg.rules, sheet_cfg.rules_combine)

    col_indices = parse_columns(sheet_cfg.columns_spec)
    if not col_indices:
        col_indices = list(range(used_w))

    selected = apply_column_selection(table_rows, col_indices)

    if sheet_cfg.paste_mode == "keep":
        shaped = shape_keep(table, row_indices, col_indices)
    else:
        shaped = shape_pack(selected)

    dest_path = sheet_cfg.destination.file_path
    standalone = _wb_cache is None

    if standalone:
        wb = _open_or_create_dest(dest_path)
    else:
        if dest_path not in _wb_cache:
            _wb_cache[dest_path] = _open_or_create_dest(dest_path)
        wb = _wb_cache[dest_path]

    ws = _get_or_create_sheet(wb, sheet_cfg.destination.sheet_name)

    plan = build_plan(ws, shaped, sheet_cfg.destination.start_col, sheet_cfg.destination.start_row)
    rows_written = 0
    if plan is not None:
        rows_written = apply_write_plan(ws, shaped, plan)

    if standalone and dest_path:
        wb.save(dest_path)

    return SheetResult(
        source_path=source_path,
        recipe_name=recipe_name,
        sheet_name=sheet_cfg.name,
        dest_file=dest_path,
        dest_sheet=sheet_cfg.destination.sheet_name,
        rows_written=rows_written,
        message="OK" if rows_written > 0 else "0 rows written",
    )


RunItem = Tuple[str, str, SheetConfig]
"""(source_path, recipe_name, sheet_cfg)"""


def run_all(
    items: Iterable[RunItem],
    on_progress: Optional[Callable[[str, Any], None]] = None,
) -> RunReport:
    """
    Execute all items in tree order using a shared in-memory workbook cache
    per dest file. This ensures successive runs to the same destination see
    each other's writes immediately without disk round-trips causing stale
    append-row calculations.

    Each workbook is saved to disk after every successful item for crash safety.
    Fail-fast on first error: remaining items are not executed.
    """
    results: List[SheetResult] = []
    ok = True
    wb_cache: Dict[str, Workbook] = {}

    def _emit(event: str, payload: Any) -> None:
        if on_progress is not None:
            try:
                on_progress(event, payload)
            except Exception:
                pass  # Progress callbacks must never break execution.

    for (source_path, recipe_name, sheet_cfg) in items:
        _emit("start", {
            "source_path": source_path,
            "recipe_name": recipe_name,
            "sheet_name": sheet_cfg.name,
        })

        try:
            result = run_sheet(source_path, sheet_cfg, recipe_name, _wb_cache=wb_cache)
        except AppError as e:
            result = SheetResult(
                source_path=source_path,
                recipe_name=recipe_name,
                sheet_name=sheet_cfg.name,
                dest_file=sheet_cfg.destination.file_path,
                dest_sheet=sheet_cfg.destination.sheet_name,
                rows_written=0,
                message=str(e),
                error_code=e.code,
                error_message=e.message,
                error_details=e.details,
            )
            results.append(result)
            _emit("error", result)
            ok = False
            # Fail-fast: save any workbooks written so far, then stop.
            for path, wb in wb_cache.items():
                if path:
                    try:
                        wb.save(path)
                    except Exception:
                        pass
            break

        # Successful write: persist to disk so file reflects current state.
        dest_path = sheet_cfg.destination.file_path
        if dest_path and dest_path in wb_cache:
            try:
                wb_cache[dest_path].save(dest_path)
            except Exception as save_err:
                result = SheetResult(
                    source_path=source_path,
                    recipe_name=recipe_name,
                    sheet_name=sheet_cfg.name,
                    dest_file=dest_path,
                    dest_sheet=sheet_cfg.destination.sheet_name,
                    rows_written=0,
                    message=f"Save failed: {save_err}",
                    error_code="SAVE_FAILED",
                    error_message=str(save_err),
                )
                results.append(result)
                _emit("error", result)
                ok = False
                break

        results.append(result)
        _emit("result", result)

    _emit("done", RunReport(ok=ok, results=results))
    return RunReport(ok=ok, results=results)
