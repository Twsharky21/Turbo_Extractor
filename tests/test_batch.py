"""
test_batch.py — Unit tests for core.batch (run_all batch execution).

Covers:
  - run_all: stacking to same destination, multiple destinations
  - Fail-fast on collision / error
  - Shared workbook cache correctness (no stale append-row between items)
  - Progress callback emission
  - Generator input
  - Empty iterable
  - Save-after-each-item crash safety
  - Mixed Keep/Pack stacking
"""
from __future__ import annotations

import os
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook, load_workbook

from core.batch import run_all, RunItem
from core.errors import AppError, DEST_BLOCKED
from core.models import Destination, Rule, SheetConfig


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _make_xlsx(path: str, sheet: str = "Sheet1", data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    if data:
        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)
    return path


def _cfg(dest_path, *, columns="", rows="", mode="pack", rules=None,
         combine="AND", start_col="A", start_row="", dest_sheet="Out",
         src_sheet="Sheet1", src_start_row=""):
    return SheetConfig(
        name=src_sheet,
        workbook_sheet=src_sheet,
        source_start_row=src_start_row,
        columns_spec=columns,
        rows_spec=rows,
        paste_mode=mode,
        rules_combine=combine,
        rules=rules or [],
        destination=Destination(
            file_path=dest_path,
            sheet_name=dest_sheet,
            start_col=start_col,
            start_row=start_row,
        ),
    )


def _pack_cfg(dest, *, start_col="B"):
    """Write all source columns starting at destination col B."""
    return _cfg(dest, columns="", start_col=start_col)


def _ws(path, sheet="Out"):
    return load_workbook(path)[sheet]


# ══════════════════════════════════════════════════════════════════════════════
# BASIC STACKING
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_two_sources_stack_same_dest():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["S1", "x", 1]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["S2", "x", 2]])

        report = run_all([(s1, "R1", _pack_cfg(dest)),
                          (s2, "R2", _pack_cfg(dest))])
        assert report.ok
        ws2 = _ws(dest)
        assert ws2["B1"].value == "S1" and ws2["C1"].value == "x"
        assert ws2["B2"].value == "S2" and ws2["C2"].value == "x"


def test_run_all_three_sources_stack_in_order():
    with TemporaryDirectory() as td:
        dest    = os.path.join(td, "out.xlsx")
        sources = []
        for i in range(1, 4):
            p = _make_xlsx(os.path.join(td, f"s{i}.xlsx"),
                           data=[[f"A{i}", i]])
            sources.append((p, f"R{i}", _pack_cfg(dest)))

        report = run_all(sources)
        assert report.ok
        ws2 = _ws(dest)
        for i in range(1, 4):
            assert ws2[f"B{i}"].value == f"A{i}"
            assert ws2[f"C{i}"].value == i


def test_run_all_two_different_destinations():
    with TemporaryDirectory() as td:
        d1 = os.path.join(td, "o1.xlsx")
        d2 = os.path.join(td, "o2.xlsx")
        s1 = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["A1", "x"]])
        s2 = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["A2", "x"]])

        report = run_all([(s1, "R1", _pack_cfg(d1)),
                          (s2, "R2", _pack_cfg(d2))])
        assert report.ok
        assert _ws(d1)["B1"].value == "A1"
        assert _ws(d2)["B1"].value == "A2"


# ══════════════════════════════════════════════════════════════════════════════
# EMPTY / GENERATOR INPUT
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_empty_iterable_returns_ok():
    report = run_all([])
    assert report.ok
    assert report.results == []


def test_run_all_generator_input_works():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["x"]])

        def gen():
            yield (s1, "R1", _cfg(dest, columns="A", start_col="A"))

        report = run_all(gen())
        assert report.ok
        assert report.results[0].rows_written == 1


# ══════════════════════════════════════════════════════════════════════════════
# FAIL-FAST
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_fail_fast_on_collision():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["r1"], ["r2"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["x"]])

        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)

        cfg_blocked = _cfg(dest, columns="A", start_col="A", start_row="1")
        cfg_second  = _cfg(dest, columns="A", start_col="A")

        report = run_all([(s1, "R1", cfg_blocked), (s2, "R2", cfg_second)])
        assert not report.ok
        assert len(report.results) == 1
        assert report.results[0].error_code == DEST_BLOCKED


def test_run_all_fail_fast_does_not_corrupt_prior_writes():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["good"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["bad"]])

        wb   = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A2"] = "BLOCK"
        wb.save(dest)

        cfg1 = _cfg(dest, columns="A", start_col="A")                    # appends to row 3
        cfg2 = _cfg(dest, columns="A", start_col="A", start_row="2")     # explicit collision

        report = run_all([(s1, "R1", cfg1), (s2, "R2", cfg2)])
        assert not report.ok
        assert report.results[0].rows_written == 1
        assert report.results[1].error_code == DEST_BLOCKED


# ══════════════════════════════════════════════════════════════════════════════
# PROGRESS CALLBACKS
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_progress_callback_called_for_each_item():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["a"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["b"]])

        events = []

        run_all([(s1, "R1", _cfg(dest, columns="A")),
                 (s2, "R2", _cfg(dest, columns="A"))],
                on_progress=lambda e, p: events.append(e))

        assert events.count("start")  == 2
        assert events.count("result") == 2
        assert "done" in events


def test_run_all_progress_callback_error_event_on_failure():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["a"]])

        wb = Workbook(); ws = wb.active; ws.title = "Out"
        ws["A1"] = "BLOCK"
        wb.save(dest)

        events = []
        run_all([(s1, "R1", _cfg(dest, columns="A", start_row="1"))],
                on_progress=lambda e, p: events.append(e))

        assert "error" in events
        assert "done" in events


def test_run_all_crashing_callback_does_not_abort_execution():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["a"]])

        def bad_cb(event, payload):
            raise RuntimeError("callback exploded")

        report = run_all([(s1, "R1", _cfg(dest, columns="A"))],
                         on_progress=bad_cb)
        assert report.ok


# ══════════════════════════════════════════════════════════════════════════════
# MIXED KEEP/PACK STACKING
# ══════════════════════════════════════════════════════════════════════════════

def test_run_all_keep_then_pack_stacks_correctly():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"),
                          data=[["alpha", "x", 1],
                                ["beta",  "y", 2],
                                ["gamma", "z", 3]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"),
                          data=[["pack_row", 99]])

        cfg_keep = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,C", rows_spec="",
            paste_mode="keep", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="A", start_row=""),
        )
        cfg_pack = _cfg(dest, columns="A,B", start_col="A")

        report = run_all([(s1, "R1", cfg_keep), (s2, "R2", cfg_pack)])
        assert report.ok
        ws2 = _ws(dest)
        # keep wrote 3 rows (bounding box A:C), pack stacks after
        assert ws2["A4"].value == "pack_row"


def test_run_all_mixed_widths_landing_zone_awareness():
    with TemporaryDirectory() as td:
        dest = os.path.join(td, "out.xlsx")
        s1   = _make_xlsx(os.path.join(td, "s1.xlsx"), data=[["v1", "v2", "v3"]])
        s2   = _make_xlsx(os.path.join(td, "s2.xlsx"), data=[["w1", "w2"]])

        cfg1 = SheetConfig(
            name="Sheet1", workbook_sheet="Sheet1",
            columns_spec="A,B,C", rows_spec="1-1",
            paste_mode="pack", rules_combine="AND", rules=[],
            destination=Destination(file_path=dest, sheet_name="Out",
                                    start_col="B", start_row=""),
        )
        cfg2 = _cfg(dest, columns="A,B", start_col="B")

        report = run_all([(s1, "R1", cfg1), (s2, "R2", cfg2)])
        assert report.ok
        ws2 = _ws(dest)
        assert ws2["B1"].value == "v1"
        assert ws2["D1"].value == "v3"
        assert ws2["B2"].value == "w1"
